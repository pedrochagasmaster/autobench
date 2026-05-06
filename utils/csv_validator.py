#!/usr/bin/env python
"""
CSV to Excel Rate Validation Script

Validates that the balanced totals in the exported CSV file correctly produce
the approval rates and fraud rates shown in the Excel benchmark report.

Usage:
    python utils/csv_validator.py <excel_file> <csv_file> [--tolerance PERCENT]

Example:
    python utils/csv_validator.py carrefour_with_time.xlsx carrefour_with_time_balanced.csv --tolerance 0.01
"""

import sys
import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import pandas as pd
from openpyxl import load_workbook


def load_csv_data(csv_path: Path) -> pd.DataFrame:
    """Load the balanced CSV export file."""
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    df = pd.read_csv(csv_path)
    
    # Validate required columns
    required_cols = ['Dimension', 'Category']
    if not all(col in df.columns for col in required_cols):
        raise ValueError(f"CSV must contain columns: {required_cols}")
    
    return df


def load_excel_data(excel_path: Path) -> Dict[str, pd.DataFrame]:
    """Load dimension sheets from Excel benchmark report.
    
    Returns dictionary mapping dimension name to DataFrame.
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    wb = load_workbook(excel_path, data_only=True)
    
    # Skip metadata sheets
    skip_sheets = {
        'Summary',
        'Metadata',
        'Peer Weights',
        'Weight Methods',
        'Privacy Validation',
        'Preset Comparison',
        'Impact Detail',
        'Impact Summary',
        'Subset Search',
        'Structural Summary',
        'Structural Detail',
        'Rank Changes',
    }
    
    dimension_data = {}
    
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        
        ws = wb[sheet_name]
        
        # Read sheet into DataFrame
        data = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Find header row (usually row 3, contains "Category")
            if headers is None:
                if row and 'Category' in [str(cell) for cell in row if cell is not None]:
                    headers = [str(cell) if cell is not None else f'Unnamed_{i}' 
                              for i, cell in enumerate(row)]
                continue
            
            # Skip empty rows
            if not any(cell is not None for cell in row):
                continue
            
            data.append(row)
        
        if headers and data:
            df = pd.DataFrame(data, columns=headers)
            dimension_data[sheet_name] = df
    
    return dimension_data


def extract_summary_metadata(excel_path: Path) -> Dict[str, str]:
    """Extract key/value metadata from Summary sheet."""
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    if 'Summary' not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb['Summary']
    metadata: Dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if len(row) > 1 and row[1] is not None:
            metadata[key] = str(row[1]).strip()
    wb.close()
    return metadata


def calculate_rate_from_csv(
    csv_df: pd.DataFrame,
    dimension: str,
    category: str,
    time_period: Optional[str],
    rate_type: str,
    total_col: str,
    approval_col: Optional[str],
    fraud_col: Optional[str],
    time_col: Optional[str],
) -> Optional[float]:
    """Calculate rate from CSV balanced totals.
    
    Args:
        csv_df: CSV DataFrame
        dimension: Dimension name
        category: Category value
        time_period: Time period value (or None if no time column)
        rate_type: 'approval', 'fraud', or 'share'
    
    Returns:
        Calculated rate as decimal (e.g., 0.75 for 75%)
    """
    # Filter to matching row
    mask = (csv_df['Dimension'] == dimension) & (csv_df['Category'] == str(category))
    
    if time_period is not None and time_col and time_col in csv_df.columns:
        # Convert time_period to match CSV format if needed
        mask = mask & (csv_df[time_col] == time_period)
    
    matched = csv_df[mask]
    
    if len(matched) == 0:
        return None
    
    if len(matched) > 1:
        print(f"WARNING: Multiple matches found for {dimension}/{category}/{time_period}")
        return None
    
    row = matched.iloc[0]

    # Handle Share Analysis (Direct Rate)
    if rate_type == 'share':
        # total_col holds the column name for share %
        if total_col in row:
            val = row[total_col]
            return val / 100.0 if pd.notna(val) else None
        return None
    
    # Get balanced totals
    balanced_total = row.get(total_col)
    
    if rate_type == 'approval':
        if approval_col is None:
            return None
        balanced_numerator = row.get(approval_col)
    elif rate_type == 'fraud':
        if fraud_col is None:
            return None
        balanced_numerator = row.get(fraud_col)
    else:
        raise ValueError(f"Unknown rate_type: {rate_type}")
    
    # Calculate rate
    if pd.isna(balanced_total) or pd.isna(balanced_numerator) or balanced_total == 0:
        return None
    
    rate = balanced_numerator / balanced_total
    return rate


def _normalize_time_value(value: object) -> str:
    """Normalize time value for comparison."""
    if value is None:
        return ""
    try:
        parsed = pd.to_datetime(value, errors='coerce')
        if pd.isna(parsed):
            return str(value)
        return parsed.date().isoformat()
    except Exception:
        return str(value)


def extract_rate_from_excel(
    excel_df: pd.DataFrame,
    category: str,
    rate_type: str,
    time_period: Optional[str] = None,
    time_col_hint: Optional[str] = None,
) -> Optional[float]:
    """Extract rate value from Excel dimension sheet.
    
    Args:
        excel_df: DataFrame from Excel dimension sheet
        category: Category value
        rate_type: 'approval' or 'fraud'
        time_period: Time period value (or None if no time column)
    
    Returns:
        Rate value as decimal (e.g., 0.75 for 75%)
    """
    # Find the rate column - try multiple patterns
    rate_cols = []
    
    # Pattern 1: Multi-rate format "Approval_Balanced Peer Average (%)"
    pattern1 = f"{rate_type.capitalize()}_Balanced Peer Average"
    rate_cols = [col for col in excel_df.columns if pattern1 in str(col) and '%' in str(col)]
    
    if not rate_cols and rate_type == 'share':
         if 'Balanced_Share_%' in excel_df.columns:
             rate_cols = ['Balanced_Share_%']
    
    if not rate_cols:
        # Pattern 2: Single rate format "Peer_Balanced_Approval_%"
        pattern2 = f"Peer_Balanced_{rate_type.capitalize()}"
        rate_cols = [col for col in excel_df.columns if pattern2 in str(col) and '%' in str(col)]
    
    if not rate_cols:
        # Pattern 3: "Balanced_Approval_Rate"
        pattern3 = f"Balanced_{rate_type.capitalize()}_Rate"
        rate_cols = [col for col in excel_df.columns if pattern3 in str(col)]
    
    if not rate_cols:
        # Pattern 4: Generic single-rate format "Balanced Peer Average (%)"
        if 'Balanced Peer Average (%)' in excel_df.columns:
            rate_cols = ['Balanced Peer Average (%)']

    if not rate_cols:
        return None
    
    rate_col = rate_cols[0]
    
    # Find time column for multi-rate format if needed
    time_col = None
    if time_period is not None:
        if time_col_hint and time_col_hint in excel_df.columns:
            time_col = time_col_hint
        else:
            time_col_pattern = f"{rate_type.capitalize()}_year_month"
            time_cols = [col for col in excel_df.columns if time_col_pattern in str(col)]
            if time_cols:
                time_col = time_cols[0]
            elif 'Time_Period' in excel_df.columns:
                time_col = 'Time_Period'
            else:
                # Common time column names
                for candidate in ['year_month', 'ano_mes', 'time_period', 'period', 'Time', 'Date']:
                    if candidate in excel_df.columns:
                        time_col = candidate
                        break
                # If still unknown, infer a single non-numeric, non-category column
                if time_col is None:
                    potential_cols = [
                        col for col in excel_df.columns
                        if col not in {'Category', rate_col}
                        and not pd.api.types.is_numeric_dtype(excel_df[col])
                    ]
                    if len(potential_cols) == 1:
                        time_col = potential_cols[0]
    
    # Filter to matching row
    # Convert both to strings for comparison (Excel may have numeric categories)
    excel_df_copy = excel_df.copy()
    excel_df_copy['Category'] = excel_df_copy['Category'].astype(str)
    mask = excel_df_copy['Category'] == str(category)

    if time_period is not None and time_col is not None:
        time_period_norm = _normalize_time_value(time_period)
        excel_df_copy[time_col] = excel_df_copy[time_col].apply(_normalize_time_value)
        mask = mask & (excel_df_copy[time_col] == time_period_norm)
    
    matched = excel_df_copy[mask]  # Use the copy with string conversions
    
    if len(matched) == 0:
        return None
    
    if len(matched) > 1:
        print(f"WARNING: Multiple matches in Excel for category {category}")
        return None
    
    row = matched.iloc[0]
    rate_value = row[rate_col]
    
    # Convert percentage to decimal if needed
    if pd.isna(rate_value):
        return None
    
    # Excel stores values as percentages (e.g., 75.5 for 75.5%)
    # Always divide by 100 to get decimal
    rate_value = rate_value / 100.0
    
    return rate_value


def validate_dimension(
    dimension: str,
    csv_df: pd.DataFrame,
    excel_df: pd.DataFrame,
    rate_types: List[str],
    tolerance: float,
    time_col: Optional[str],
    total_col: str,
    approval_col: Optional[str],
    fraud_col: Optional[str],
) -> Dict[str, any]:
    """Validate all categories in a dimension.
    
    Returns:
        Dictionary with validation results
    """
    results = {
        'dimension': dimension,
        'total_checks': 0,
        'passed': 0,
        'failed': 0,
        'skipped': 0,
        'failures': []
    }
    
    # Get unique categories from CSV
    csv_categories = csv_df[csv_df['Dimension'] == dimension]['Category'].unique()
    
    for category in csv_categories:
        # Get time periods if applicable
        time_periods = [None]
        if time_col and time_col in csv_df.columns:
            mask = (csv_df['Dimension'] == dimension) & (csv_df['Category'] == str(category))
            time_periods = csv_df[mask][time_col].unique()
        
        for time_period in time_periods:
            for rate_type in rate_types:
                results['total_checks'] += 1
                
                # Calculate rate from CSV
                csv_rate = calculate_rate_from_csv(
                    csv_df,
                    dimension,
                    category,
                    time_period,
                    rate_type,
                    total_col,
                    approval_col,
                    fraud_col,
                    time_col,
                )
                
                # Extract rate from Excel
                excel_rate = extract_rate_from_excel(
                    excel_df,
                    category,
                    rate_type,
                    time_period,
                    time_col_hint=time_col,
                )
                
                if csv_rate is None or excel_rate is None:
                    results['skipped'] += 1
                    continue
                
                # Compare rates
                diff = abs(csv_rate - excel_rate)
                
                if diff <= tolerance:
                    results['passed'] += 1
                else:
                    results['failed'] += 1
                    results['failures'].append({
                        'category': category,
                        'time_period': time_period,
                        'rate_type': rate_type,
                        'csv_rate': csv_rate,
                        'excel_rate': excel_rate,
                        'difference': diff,
                        'difference_pct': diff * 100
                    })
    
    return results


def main():
    parser = argparse.ArgumentParser(
        description='Validate CSV balanced totals against Excel rates',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
EXAMPLES:
  # Basic validation
  python utils/csv_validator.py report.xlsx report_balanced.csv
  
  # With custom tolerance (0.1% = 0.001)
  python utils/csv_validator.py report.xlsx report_balanced.csv --tolerance 0.001
  
  # Verbose output
  python utils/csv_validator.py report.xlsx report_balanced.csv --verbose
        """
    )
    
    parser.add_argument('excel_file', help='Path to Excel benchmark report')
    parser.add_argument('csv_file', help='Path to balanced CSV export')
    parser.add_argument('--tolerance', type=float, default=0.0001,
                       help='Maximum allowed difference in rates (default: 0.0001 = 0.01%%)')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Show detailed validation results')
    
    args = parser.parse_args()
    
    excel_path = Path(args.excel_file)
    csv_path = Path(args.csv_file)
    
    print(f"{'='*80}")
    print("CSV TO EXCEL RATE VALIDATION")
    print(f"{'='*80}")
    print(f"Excel File: {excel_path}")
    print(f"CSV File: {csv_path}")
    print(f"Tolerance: {args.tolerance:.6f} ({args.tolerance*100:.4f}%)")
    print(f"{'='*80}\n")
    
    # Load data
    try:
        print("Loading CSV data...")
        csv_df = load_csv_data(csv_path)
        print(f"  OK Loaded {len(csv_df)} rows")

        print("Loading Excel data...")
        excel_data = load_excel_data(excel_path)
        print(f"  OK Loaded {len(excel_data)} dimension sheets")

        summary_metadata = extract_summary_metadata(excel_path)

    except Exception as e:
        print(f"ERROR loading data: {e}")
        return 1

    # Detect CSV schema and rate types
    rate_types: List[str] = []
    total_col = None
    approval_col = None
    fraud_col = None

    # Detect Share Analysis (Dynamic metric name)
    # Look for column: Balanced_{metric}_Share_%
    share_col = None
    for col in csv_df.columns:
        if col.startswith("Balanced_") and col.endswith("_Share_%"):
            share_col = col
            break
            
    if 'Balanced_Total' in csv_df.columns:
        total_col = 'Balanced_Total'
        if 'Balanced_Approval_Total' in csv_df.columns:
            approval_col = 'Balanced_Approval_Total'
            rate_types.append('approval')
        if 'Balanced_Fraud_Total' in csv_df.columns:
            fraud_col = 'Balanced_Fraud_Total'
            rate_types.append('fraud')
    elif share_col:
        # Share analysis
        rate_types.append('share')
        total_col = share_col # Use this as the "value" column for share rate
    else:
        summary_total = summary_metadata.get('Total Column') or summary_metadata.get('Total Column (Shared Denominator)')
        summary_approval = summary_metadata.get('Approval Column')
        summary_fraud = summary_metadata.get('Fraud Column')
        if summary_total and summary_total in csv_df.columns:
            total_col = summary_total
        if summary_approval and summary_approval in csv_df.columns:
            approval_col = summary_approval
            rate_types.append('approval')
        if summary_fraud and summary_fraud in csv_df.columns:
            fraud_col = summary_fraud
            rate_types.append('fraud')
        if total_col is None:
            candidate_numeric_cols = [
                col for col in csv_df.columns
                if col not in {'Dimension', 'Category'} and pd.api.types.is_numeric_dtype(csv_df[col])
            ]
            if candidate_numeric_cols:
                total_col = candidate_numeric_cols[0]
            if approval_col is None:
                for candidate in ['approved', 'approval', 'approved_count']:
                    if candidate in csv_df.columns:
                        approval_col = candidate
                        rate_types.append('approval')
                        break
            if fraud_col is None:
                for candidate in ['fraud', 'fraud_count', 'fraud_amount']:
                    if candidate in csv_df.columns:
                        fraud_col = candidate
                        rate_types.append('fraud')
                        break

    if not total_col or not rate_types:
        print("ERROR No rate columns found in CSV")
        print("  Expected Balanced_* columns or summary-based metric columns.")
        return 1

    print(f"Rate Types: {', '.join(rate_types)}")

    # Check for time column
    time_col = None
    time_name_candidates = ['year_month', 'ano_mes', 'time_period', 'period']
    for name in time_name_candidates:
        if name in csv_df.columns:
            time_col = name
            break
    if time_col is None:
        non_value_cols = {'Dimension', 'Category', total_col}
        if approval_col:
            non_value_cols.add(approval_col)
        if fraud_col:
            non_value_cols.add(fraud_col)
        potential_time_cols = [
            c for c in csv_df.columns
            if c not in non_value_cols and not pd.api.types.is_numeric_dtype(csv_df[c])
        ]
        if len(potential_time_cols) == 1:
            time_col = potential_time_cols[0]

    if time_col:
        print(f"Time-Aware: Yes ({time_col} column detected)")
    else:
        print("Time-Aware: No")

    print()
    # Validate each dimension
    all_results = []
    
    for dimension in csv_df['Dimension'].unique():
        # Find matching Excel sheet
        excel_df = None
        for sheet_name, df in excel_data.items():
            normalized_sheet = sheet_name
            if normalized_sheet.startswith("Metric_"):
                parts = normalized_sheet.split("_", 2)
                if len(parts) == 3:
                    normalized_sheet = parts[2]
            normalized_candidates = {sheet_name, normalized_sheet}
            for candidate in list(normalized_candidates):
                if "_" in candidate:
                    for idx in range(len(candidate.split("_"))):
                        suffix = "_".join(candidate.split("_")[idx:])
                        if suffix == dimension:
                            normalized_candidates.add(dimension)
                            break
            # Match by sheet name (exact or sanitized)
            if (
                dimension in normalized_candidates
                or any(candidate.replace('_', '/') == dimension for candidate in normalized_candidates)
            ):
                excel_df = df
                break
        
        if excel_df is None:
            print(f"FAIL {dimension}: No matching Excel sheet found")
            all_results.append({
                "dimension": dimension,
                "total_checks": 0,
                "passed": 0,
                "failed": 1,
                "skipped": 0,
                "failures": [],
            })
            continue
        
        print(f"Validating dimension: {dimension}")
        results = validate_dimension(
            dimension,
            csv_df,
            excel_df,
            rate_types,
            args.tolerance,
            time_col,
            total_col,
            approval_col,
            fraud_col,
        )
        all_results.append(results)
        
        # Print summary
        status = "PASS" if results['failed'] == 0 else "FAIL"
        print(f"  {status} - {results['passed']}/{results['total_checks']} checks passed")
        
        if results['failed'] > 0:
            print(f"    Failed: {results['failed']}")
            if args.verbose:
                for failure in results['failures']:
                    print(f"      - {failure['category']} ({failure['rate_type']}): "
                          f"CSV={failure['csv_rate']:.4%} vs Excel={failure['excel_rate']:.4%} "
                          f"(Delta={failure['difference_pct']:.4f}%)")
        
        if results['skipped'] > 0:
            print(f"    Skipped: {results['skipped']} (missing data)")
        
        print()
    
    # Overall summary
    total_checks = sum(r['total_checks'] for r in all_results)
    total_passed = sum(r['passed'] for r in all_results)
    total_failed = sum(r['failed'] for r in all_results)
    total_skipped = sum(r['skipped'] for r in all_results)
    
    print(f"{'='*80}")
    print("VALIDATION SUMMARY")
    print(f"{'='*80}")
    print(f"Dimensions Validated: {len(all_results)}")
    print(f"Total Checks: {total_checks}")
    if total_checks > 0:
        print(f"Passed: {total_passed} ({total_passed/total_checks*100:.1f}%)")
        print(f"Failed: {total_failed} ({total_failed/total_checks*100:.1f}%)")
        print(f"Skipped: {total_skipped} ({total_skipped/total_checks*100:.1f}%)")
    else:
        print(f"Passed: {total_passed} (0.0%)")
        print(f"Failed: {total_failed} (100.0%)")
        print(f"Skipped: {total_skipped} (0.0%)")
    
    if total_failed == 0:
        print("\nALL VALIDATIONS PASSED")
        print(f"  CSV balanced totals correctly produce Excel rates within {args.tolerance*100:.4f}% tolerance")
        print(f"{'='*80}\n")
        return 0
    else:
        print("\nVALIDATION FAILED")
        print(f"  {total_failed} rate calculations do not match within tolerance")
        print("  Review failures above or run with --verbose for details")
        print(f"{'='*80}\n")
        return 1


if __name__ == '__main__':
    sys.exit(main())
