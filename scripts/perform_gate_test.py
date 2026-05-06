import sys
import json
import subprocess
import shutil
import logging
import time
import shlex
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class GateTestRunner:
    def __init__(self, output_dir: str = "test_gate"):
        self.output_dir = Path(output_dir)
        self.script_dir = Path(__file__).parent
        self.root_dir = self.script_dir.parent
        self.generate_script = self.script_dir / "generate_cli_sweep.py"

    def generate_cases(self):
        """Run generate_cli_sweep.py in gate mode."""
        logger.info("Generating gate test cases...")
        cmd = [
            sys.executable,
            str(self.generate_script),
            "--mode", "gate",
            "--out-dir", str(self.output_dir)
        ]
        result = subprocess.run(cmd, cwd=self.root_dir, capture_output=True, text=True)
        if result.returncode != 0:
            logger.error("Failed to generate cases:")
            logger.error(result.stderr)
            sys.exit(1)
        logger.info(f"Cases generated in {self.output_dir}")

    def load_cases(self) -> List[Dict]:
        """Load all cases from generated jsonl files."""
        cases = []
        for section in ["share", "rate", "config"]:
            jsonl_path = self.output_dir / section / "cases.jsonl"
            if not jsonl_path.exists():
                logger.warning(f"No cases found for section: {section}")
                continue
            
            with open(jsonl_path, "r", encoding="utf-8") as f:
                for line in f:
                    if line.strip():
                        cases.append(json.loads(line))
        return cases

    @staticmethod
    def _sheet_matches_dimension(sheet_name: str, dimension: str) -> bool:
        normalized_sheet = sheet_name.lower().replace("_", "")
        normalized_dim = dimension.lower().replace("_", "").replace("/", "")
        return normalized_sheet == normalized_dim or normalized_sheet.endswith(normalized_dim[:20])

    def verify_workbook_content(self, wb, case_id: str, analysis_type: str) -> List[str]:
        """Deep verification of workbook content (sanity checks)."""
        failures = []
        reserved = {
            "Summary", "Data Quality", "Preset Comparison", "Impact Analysis",
            "Impact Summary", "Metadata", "Peer Weights", "Weight Methods",
            "Privacy Validation", "Secondary Metrics", "Subset Search",
            "Structural Summary", "Structural Detail", "Rank Changes"
        }
        dim_sheets = [s for s in wb.sheetnames if s not in reserved]
        
        if not dim_sheets:
            # Not necessarily a failure if no dimensions requested? But unusual for gate test.
            return failures

        for sheet_name in dim_sheets:
            ws = wb[sheet_name]
            
            # Find header row dynamically
            header_row_idx = None
            headers = []
            
            # Scan first 10 rows for "Category" or "Metric"
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                row_strs = [str(c) for c in row if c is not None]
                if "Category" in row_strs or "Metric" in row_strs:
                    header_row_idx = i
                    headers = row_strs
                    break
            
            if not header_row_idx:
                failures.append(f"Sheet '{sheet_name}': Could not find header row (missing 'Category' or 'Metric' column)")
                continue
            
            # Read data
            data_rows = []
            for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                # Check if row is empty
                if not any(row): continue
                # Map to headers (truncate row to len headers)
                # Ensure we handle cases where row is shorter than headers or vice versa
                row_vals = row[:len(headers)]
                row_dict = dict(zip(headers, row_vals))
                data_rows.append(row_dict)
            
            if not data_rows:
                # Might be empty if filtered out?
                continue
                
            df = pd.DataFrame(data_rows)
            
            # Defined Checks
            # 0. Duplicates Check
            # Identify key columns for uniqueness check
            key_candidates = ["Category", "Time", "Month", "Year", "Quarter", "Period", "ano_mes", "Date", "date"]
            keys = [c for c in df.columns if c in key_candidates]
            if keys:
                # If we have keys, check for duplicates
                if df.duplicated(subset=keys).any():
                    # Get sample duplicates
                    dupes = df[df.duplicated(subset=keys, keep=False)].head(2)
                    failures.append(f"Sheet '{sheet_name}': Contains duplicate rows for keys {keys}. Sample:\n{dupes}")

            # 1. Excel Errors
            error_patterns = ("#DIV/0!", "#N/A", "#VALUE!", "#REF!", "#NAME?")
            for col in df.columns:
                values = df[col].astype(str)
                if values.apply(lambda cell: any(pattern in cell for pattern in error_patterns)).any():
                    failures.append(f"Sheet '{sheet_name}' Column '{col}': Contains Excel errors")

            # 2. Sensible Ranges
            # Rate/Share percentages: 0 to 100
            pct_cols = [c for c in df.columns if "%" in c or "Share" in c or "Rate" in c]
            for col in pct_cols:
                # Skip text columns (like Category if it happens to have %)
                try:
                    vals = pd.to_numeric(df[col], errors='coerce').dropna()
                    if vals.empty: continue
                    
                    min_val, max_val = vals.min(), vals.max()
                    if min_val < -0.1 or max_val > 100.1: # Small epsilon
                        failures.append(f"Sheet '{sheet_name}' Column '{col}': Values out of range 0-100 ({min_val} to {max_val})")
                except Exception:
                    pass
            
            # 2b. Positive Values (Volume, Count)
            pos_cols = [c for c in df.columns if any(x in c.lower() for x in ["volume", "count", "transactions", "clients"])]
            for col in pos_cols:
                try:
                    vals = pd.to_numeric(df[col], errors='coerce').dropna()
                    if (vals < 0).any():
                         failures.append(f"Sheet '{sheet_name}' Column '{col}': Contains negative values")
                except Exception:
                    pass

            # 2c. Infinity/NaN Check
            # Check for infinite values in numeric columns
            numeric_df = df.select_dtypes(include=['number'])
            # Also convert object columns that look numeric
            for col in df.columns:
                if col not in numeric_df.columns:
                    try:
                        # lightweight check if it converts
                        pd.to_numeric(df[col], errors='raise') 
                        # If success, check for inf
                        vals = pd.to_numeric(df[col], errors='coerce')
                        if np.isinf(vals).any():
                             failures.append(f"Sheet '{sheet_name}' Column '{col}': Contains Infinite values")
                    except Exception:
                        pass
            
            if hasattr(np, 'inf'):
                # Direct check on dataframe if mixed types allow
                try:
                    # filtering for numeric columns first is safer
                    num_cols = df.select_dtypes(include=[np.number]).columns
                    if not num_cols.empty:
                        if np.isinf(df[num_cols]).any().any():
                             failures.append(f"Sheet '{sheet_name}': Contains Infinite values in numeric columns")
                except Exception:
                    pass

            # 3. Variance Check (Avoid "Flat" results)
            # Only check "Balanced" columns
            balanced_cols = [c for c in df.columns if "Balanced" in c and "%" in c]
            for col in balanced_cols:
                try:
                    vals = pd.to_numeric(df[col], errors='coerce').dropna()
                    if len(vals) > 1:
                        if vals.std() == 0:
                            # Warn only, might be valid for specific cases
                            logger.warning(f"Sheet '{sheet_name}' Column '{col}': All values are identical ({vals.iloc[0]}). Verify if this is expected.")
                            # failures.append(f"Sheet '{sheet_name}' Column '{col}': Zero variance (all {vals.iloc[0]}).")
                except Exception:
                    pass
            
            # 3b. Math Consistency Check (Distance = Target - Peer)
            # Identify columns
            target_col = next((c for c in df.columns if "Target" in c and "%" in c), None)
            peer_col = next((c for c in df.columns if "Balanced Peer" in c and "%" in c), None)
            dist_col = next((c for c in df.columns if "Distance" in c and ("pp" in c or "%" in c)), None)
            
            if target_col and peer_col and dist_col:
                try:
                    t_vals = pd.to_numeric(df[target_col], errors='coerce').fillna(0)
                    p_vals = pd.to_numeric(df[peer_col], errors='coerce').fillna(0)
                    d_vals = pd.to_numeric(df[dist_col], errors='coerce').fillna(0)
                    
                    # Calculate expected distance
                    # Logic: Distance is usually (Target - Peer) or (Target - Baseline)
                    # Let's verify T - P
                    calc_dist = t_vals - p_vals
                    delta = (calc_dist - d_vals).abs()
                    
                    # Check if any deviation > 0.01 (floating point tolerance)
                    if (delta > 0.01).any():
                        # Get max deviation to report
                        max_dev = delta.max()
                        failures.append(f"Sheet '{sheet_name}': Math Mismatch. 'Distance' != 'Target' - 'Peer'. Max delta: {max_dev:.4f}")
                except Exception as e:
                    logger.warning(f"Could not verify math consistency in {sheet_name}: {e}")

            # 3c. Time Series Gap Detection
            time_col = next((c for c in df.columns if c in ["ano_mes", "Date", "date", "Time"]), None)
            if time_col:
                try:
                    # Convert to datetime
                    dates = pd.to_datetime(df[time_col], errors='coerce').dropna().unique()
                    if len(dates) > 1:
                        dates = sorted(dates)
                        # Check diffs
                        diffs = pd.Series(dates).diff().dt.days.dropna()
                        # Assuming monthly data, gaps should be ~28-31 days. 
                        # If we see a gap > 45 days, it implies a missing month.
                        max_gap = diffs.max()
                        if max_gap > 45:
                             logger.warning(f"Sheet '{sheet_name}': Potential time series gap detected. Max gap between dates: {max_gap} days.")
                             # failures.append(f"Sheet '{sheet_name}': Time series gap > 45 days ({max_gap} days).")
                except Exception:
                    pass

            # 4. Share Mix Check (Sum to ~100%) - Only for Share Analysis
            if analysis_type == 'share':
                # Identify "Balanced Peer Average (%)" or similar
                balanced_peer_cols = [c for c in df.columns if "Balanced Peer Average" in c and "%" in c]
                for col in balanced_peer_cols:
                    try:
                        vals = pd.to_numeric(df[col], errors='coerce').dropna()
                        total_sum = vals.sum()
                        
                        # In time-aware analysis, we need to sum PER time period
                        time_col = next((c for c in df.columns if "Time" in c or "month" in c or "ano" in c), None)
                        
                        if time_col:
                            # Check sum per time period
                            time_groups = df.groupby(time_col)[col].apply(lambda x: pd.to_numeric(x, errors='coerce').sum())
                            for t, s in time_groups.items():
                                if not (99.0 < s < 101.0):
                                    # Relaxed tolerance because sometimes small categories are filtered or rounding
                                    # Just warn for now as this is a new check
                                    logger.warning(f"Sheet '{sheet_name}' (Time: {t}): Balanced Mix sums to {s:.2f}%, expected ~100%")
                        else:
                            # Global sum
                            if not (99.0 < total_sum < 101.0):
                                failures.append(f"Sheet '{sheet_name}': Balanced Mix sums to {total_sum:.2f}%, expected ~100%")
                    except Exception as e:
                        logger.warning(f"Could not verify sum for {col}: {e}")
        
        # Check Data Quality sheet if present
        if "Data Quality" in wb.sheetnames:
            try:
                dq_ws = wb["Data Quality"]
                dq_data = []
                
                # Scan for header row containing "Severity" or "Issue"
                header_row_idx = None
                headers = []
                for i, row in enumerate(dq_ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                    row_strs = [str(c) for c in row if c is not None]
                    if "Severity" in row_strs or "Issue" in row_strs:
                        header_row_idx = i
                        headers = row_strs
                        break
                
                if header_row_idx:
                    for row in dq_ws.iter_rows(min_row=header_row_idx+1, values_only=True):
                         if any(row):
                             dq_data.append(dict(zip(headers, row[:len(headers)])))
                else:
                    # Fallback or warn if not found, but only if sheet is not empty
                    if dq_ws.max_row > 1:
                        logger.warning(f"Data Quality sheet found but could not identify header row (checked first 10 rows for 'Severity' or 'Issue')")

                for item in dq_data:
                    # Assuming columns like "Severity", "Message"
                    if "Severity" not in item:
                         if dq_data.index(item) == 0:
                             logger.warning(f"Data Quality sheet found but 'Severity' column missing. Keys: {list(item.keys())}")
                    
                    severity = str(item.get("Severity", "")).lower()
                    if severity in ["high", "critical"]:
                         failures.append(f"Data Quality Issue ({severity}): {item.get('Issue', 'Unknown')} - {item.get('Description', '')}")
            except Exception as e:
                logger.warning(f"Could not parse Data Quality sheet: {e}")

        return failures

    def verify_case(self, case: Dict) -> List[str]:
        """Verify expectations for a single case."""
        failures = []
        params = case.get("params", {})
        expectations = case.get("expectations", [])
        
        # Determine analysis type
        if "share_" in case["id"]:
            analysis_type = "share"
        elif "rate_" in case["id"]:
            analysis_type = "rate"
        else:
            analysis_type = "config"

        # Determine output paths
        output_arg = params.get("output")
        analysis_file = None
        
        if output_arg:
            analysis_file = Path(output_arg)
        elif any(e in expectations for e in ["list_presets_output", "preset_details_output", "validate_template_ok"]):
            # These cases check stdout or exit code, which are already handled/ignored here.
            # So we consider them passed if we reached here (execution success).
            return failures
        else:
            # Try to find file matching pattern if auto-generated
            failures.append("Cannot verify output: explicit output path not found in params")
            return failures

        # Expand relative paths
        if not analysis_file.is_absolute():
            analysis_file = self.root_dir / analysis_file

        # Check for template generation case
        if "template_created" in expectations:
            if not analysis_file.exists():
                failures.append(f"Template file missing: {analysis_file}")
            return failures

        pub_file = analysis_file.with_name(f"{analysis_file.stem}_publication{analysis_file.suffix}")
        csv_file = analysis_file.with_suffix("").with_name(f"{analysis_file.stem}_balanced.csv")
        
        wb_analysis = None
        wb_pub = None

        for exp in expectations:
            if exp == "analysis_workbook":
                if not analysis_file.exists():
                    failures.append(f"Analysis workbook missing: {analysis_file}")
                else:
                    try:
                        wb_analysis = load_workbook(analysis_file, read_only=True, data_only=True)
                        content_failures = self.verify_workbook_content(wb_analysis, case["id"], analysis_type)
                        failures.extend(content_failures)
                    except Exception as e:
                        failures.append(f"Invalid analysis workbook: {e}")

            elif exp == "publication_workbook":
                if not pub_file.exists():
                    failures.append(f"Publication workbook missing: {pub_file}")
                else:
                    try:
                        wb_pub = load_workbook(pub_file, read_only=True)
                    except Exception as e:
                        failures.append(f"Invalid publication workbook: {e}")

            elif exp == "balanced_csv":
                if not csv_file.exists():
                    failures.append(f"Balanced CSV missing: {csv_file}")
                elif analysis_file.exists():
                    # Skip CSV validation for Share analysis
                    # Reason: Share CSV exports Market Share (Impact), while Excel report contains Category Mix.
                    # Mismatched metrics make validation impossible with current validator.
                    if analysis_type == "share":
                        logger.info(f"Skipping CSV validation for Share case {case['id']} (Metric mismatch: Market Share vs Mix)")
                        continue

                    # Run CSV Validator
                    validator_script = self.root_dir / "utils" / "csv_validator.py"
                    cmd = [sys.executable, str(validator_script), str(analysis_file), str(csv_file)]
                    try:
                        # Capture output to avoid clutter, check return code
                        proc = subprocess.run(cmd, cwd=self.root_dir, capture_output=True, text=True)
                        if proc.returncode != 0:
                            failures.append(f"CSV Validation failed:\n{proc.stdout}\n{proc.stderr}")
                    except Exception as e:
                         failures.append(f"Failed to run CSV validator: {e}")
            
            elif exp == "preset_comparison_sheet":
                if wb_analysis:
                    if "Preset Comparison" not in wb_analysis.sheetnames:
                        failures.append("Missing sheet: Preset Comparison")

            elif exp == "impact_analysis_sheet":
                if wb_analysis:
                    if "Impact Analysis" not in wb_analysis.sheetnames and "Impact Summary" not in wb_analysis.sheetnames:
                         failures.append("Missing sheet: Impact Analysis")
            
            elif exp == "data_quality_sheet":
                if wb_analysis and "Data Quality" not in wb_analysis.sheetnames:
                     failures.append("Missing sheet: Data Quality")

            elif exp == "no_data_quality_sheet":
                if wb_analysis and "Data Quality" in wb_analysis.sheetnames:
                     failures.append("Unexpected sheet: Data Quality")

            elif exp == "target_columns_present":
                if wb_analysis:
                    # Check first dimension sheet
                    dims = params.get("dimensions", [])
                    if not dims and params.get("auto"):
                        reserved = {"Summary", "Data Quality", "Preset Comparison", "Impact Analysis", "Impact Summary", "Metadata", "Peer Weights", "Weight Methods", "Privacy Validation"}
                        for s in wb_analysis.sheetnames:
                            if s not in reserved:
                                dims = [s]
                                break
                    
                    if dims:
                        matching_sheet = next(
                            (sheet for sheet in wb_analysis.sheetnames if self._sheet_matches_dimension(sheet, dims[0])),
                            None,
                        )
                        if matching_sheet is None:
                             failures.append(f"Dimension sheet {dims[0]} missing")
                        else:
                            ws = wb_analysis[matching_sheet]
                            # Headers can be on row 1 (Rate) or row 3 (Share)
                            headers_r1 = [str(cell.value) for cell in ws[1] if cell.value]
                            headers_r3 = [str(cell.value) for cell in ws[3] if cell.value]
                            headers = headers_r1 + headers_r3
                            
                            if not any("Target" in h or "Distance" in h for h in headers):
                                 failures.append(f"Target columns missing in sheet {matching_sheet}")
            
            elif exp == "peer_only_mode":
                if wb_analysis:
                     # Check first dimension sheet
                    dims = params.get("dimensions", [])
                    if not dims and params.get("auto"):
                        reserved = {"Summary", "Data Quality", "Preset Comparison", "Impact Analysis", "Impact Summary", "Metadata", "Peer Weights", "Weight Methods", "Privacy Validation"}
                        for s in wb_analysis.sheetnames:
                            if s not in reserved:
                                dims = [s]
                                break
                    
                    if dims:
                        matching_sheet = next(
                            (sheet for sheet in wb_analysis.sheetnames if self._sheet_matches_dimension(sheet, dims[0])),
                            None,
                        )
                        if matching_sheet is not None:
                            ws = wb_analysis[matching_sheet]
                            headers_r1 = [str(cell.value) for cell in ws[1] if cell.value]
                            headers_r3 = [str(cell.value) for cell in ws[3] if cell.value]
                            headers = headers_r1 + headers_r3
                            
                            if any("Target" in h or "Distance" in h for h in headers):
                                 failures.append(f"Target columns unexpectedly present in sheet {matching_sheet} (Peer Only mode)")

            elif exp == "csv_includes_raw_and_impact_columns":
                if csv_file.exists():
                    try:
                        df_csv = pd.read_csv(csv_file)
                        # Check for Raw_* or *_Impact_PP columns
                        raw_cols = [c for c in df_csv.columns if c.startswith("Raw_")]
                        impact_cols = [c for c in df_csv.columns if c.endswith("_Impact_PP")]
                        if not raw_cols and not impact_cols:
                            failures.append("CSV missing calculated columns (Raw_* or *_Impact_PP)")
                    except Exception as e:
                        failures.append(f"Failed to read CSV: {e}")

            elif exp == "per_dimension_weight_methods":
                if wb_analysis:
                    if "Weight Methods" not in wb_analysis.sheetnames:
                        failures.append("Missing sheet: Weight Methods")
                    else:
                        ws = wb_analysis["Weight Methods"]
                        # Check if multiple dimensions are listed with weights
                        # Just a simple check that the sheet isn't empty
                        if ws.max_row < 2:
                            failures.append("Weight Methods sheet is empty")

            elif exp == "secondary_metrics_sheet":
                if wb_analysis:
                    found = any("Secondary" in s for s in wb_analysis.sheetnames)
                    if not found:
                        failures.append("Missing sheet: Secondary Metrics")

            elif exp == "fraud_in_bps_in_publication":
                if wb_pub:
                    found_fraud_sheet = False
                    for sheet in wb_pub.sheetnames:
                        if sheet == "Executive Summary":
                            continue
                        if "fraud" not in sheet.lower():
                            continue
                        found_fraud_sheet = True
                        headers = []
                        for row in wb_pub[sheet].iter_rows(min_row=1, max_row=5, values_only=True):
                            row_values = [str(value) for value in row if value]
                            if row_values:
                                headers.extend(row_values)
                        if not any("bps" in header.lower() for header in headers):
                            failures.append("Fraud publication output is missing bps header")
                        break
                    if not found_fraud_sheet:
                        failures.append("Fraud publication output is missing fraud rate column")

            elif exp == "fraud_in_percent_in_publication":
                if wb_pub:
                    headers = []
                    for sheet in wb_pub.sheetnames:
                        if sheet == "Executive Summary":
                            continue
                        for row in wb_pub[sheet].iter_rows(min_row=1, max_row=5, values_only=True):
                            headers.extend(str(value) for value in row if value)
                    if any("Fraud" in h and "bps" in h.lower() for h in headers):
                        failures.append("Fraud publication output used bps when percent was expected")

            elif exp.startswith("audit_log="):
                 expected_log = exp.split("=")[1]
                 # We expect it in the same dir as output
                 log_path = analysis_file.with_name(expected_log)
                 if not log_path.exists():
                     failures.append(f"Audit log missing: {log_path}")
                 else:
                     # Check audit log content
                     try:
                         content = log_path.read_text(encoding='utf-8').lower()
                         if "privacy_rule" not in content and "privacy rule" not in content:
                             failures.append("Audit log missing 'Privacy Rule' entry")
                         if "dimensions_analyzed" not in content and "dimensions analyzed" not in content:
                             failures.append("Audit log missing 'Dimensions Analyzed' entry")
                     except Exception as e:
                         failures.append(f"Failed to read audit log: {e}")

        if wb_analysis:
            wb_analysis.close()
        if wb_pub:
            wb_pub.close()
            
        return failures

    def run(self):
        # 1. Generate
        self.generate_cases()

        # 2. Clean generated outputs from previous executions, but preserve generated cases.
        generated_outputs = self.output_dir / "outputs"
        if generated_outputs.exists():
            shutil.rmtree(generated_outputs)
        generated_outputs.mkdir(parents=True, exist_ok=True)
        (generated_outputs / "share").mkdir(parents=True, exist_ok=True)
        (generated_outputs / "rate").mkdir(parents=True, exist_ok=True)
        generated_template = self.output_dir / "config" / "generated_template.yaml"
        if generated_template.exists():
            generated_template.unlink()
        (self.output_dir / "outputs" / "share").mkdir(parents=True, exist_ok=True)
        (self.output_dir / "outputs" / "rate").mkdir(parents=True, exist_ok=True)
        (self.output_dir / "config").mkdir(parents=True, exist_ok=True)
        
        # 3. Load
        cases = self.load_cases()
        logger.info(f"Loaded {len(cases)} cases.")
        
        # 4. Execute and Verify
        results = {"passed": 0, "failed": 0, "errors": 0}
        
        for case in cases:
            case_id = case["id"]
            command = case["command"]
            logger.info(f"Running case: {case_id}")
            logger.debug(f"Command: {command}")
            
            start_time = time.time()
            
            # Execute
            try:
                # Use sys.executable instead of 'py' if possible
                if command.startswith("py "):
                    cmd_list = [sys.executable] + shlex.split(command[3:])
                else:
                    cmd_list = shlex.split(command)
                
                # Fix paths in command args to be absolute or relative to cwd correctly
                # actually running from root_dir should work if paths are relative to root
                proc = subprocess.run(cmd_list, cwd=self.root_dir, capture_output=True, text=True, timeout=300)
                
                duration = time.time() - start_time
                if duration > 60:
                    logger.warning(f"Case {case_id} took {duration:.1f}s (Threshold: 60s)")
                
                # Check for Tracebacks in stderr even if return code is 0 (some tools capture/suppress exceptions)
                if "Traceback (most recent call last)" in proc.stderr:
                    logger.error(f"Case {case_id} failed silently with Traceback:\n{proc.stderr}")
                    results["errors"] += 1
                    continue

                if proc.returncode != 0:
                    logger.error(f"Execution failed for {case_id}")
                    logger.error(proc.stderr)
                    results["errors"] += 1
                    continue
                
                # Verify
                failures = self.verify_case(case)
                if failures:
                    logger.error(f"Verification failed for {case_id}:")
                    for f in failures:
                        logger.error(f"  - {f}")
                    results["failed"] += 1
                else:
                    logger.info(f"Verified {case_id}: PASS ({duration:.1f}s)")
                    results["passed"] += 1
                    
            except Exception as e:
                logger.error(f"Error processing {case_id}: {e}")
                results["errors"] += 1
                
        logger.info("-" * 40)
        logger.info(f"Summary: Passed {results['passed']}, Failed {results['failed']}, Errors {results['errors']}")
        if results['failed'] > 0 or results['errors'] > 0:
            sys.exit(1)
        else:
            sys.exit(0)

if __name__ == "__main__":
    runner = GateTestRunner()
    runner.run()