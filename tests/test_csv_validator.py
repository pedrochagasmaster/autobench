"""Unit tests for utils/csv_validator.py."""

from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from utils import csv_validator


def _write_minimal_rate_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "approval_card_type"
    ws["A3"] = "Category"
    ws["B3"] = "Balanced Peer Average (%)"
    ws["C3"] = "Target Rate (%)"
    ws["A4"] = "CREDIT"
    ws["B4"] = 92.0
    ws["C4"] = 93.0
    wb.save(path)


def _write_minimal_share_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "card_type"
    ws["A3"] = "Category"
    ws["B3"] = "Balanced Peer Average (%)"
    ws["C3"] = "Target Share (%)"
    ws["A4"] = "CREDIT"
    ws["B4"] = 12.5
    ws["C4"] = 10.0
    wb.save(path)


def test_validator_passes_matching_rate_csv_and_excel(tmp_path: Path) -> None:
    excel_path = tmp_path / "report.xlsx"
    csv_path = tmp_path / "report_balanced.csv"
    _write_minimal_rate_workbook(excel_path)
    pd.DataFrame(
        [
            {
                "Dimension": "card_type",
                "Category": "CREDIT",
                "total": 1000.0,
                "approved": 920.0,
            }
        ]
    ).to_csv(csv_path, index=False)

    results = csv_validator.validate_dimension(
        "card_type",
        pd.read_csv(csv_path),
        pd.read_excel(excel_path, sheet_name="approval_card_type", header=2),
        ["approval"],
        tolerance=0.05,
        time_col=None,
        total_col="total",
        approval_col="approved",
        fraud_col=None,
    )
    assert results["failed"] == 0
    assert results["total_checks"] > 0


def test_validator_flags_drift(tmp_path: Path) -> None:
    excel_path = tmp_path / "report.xlsx"
    csv_path = tmp_path / "report_balanced.csv"
    _write_minimal_rate_workbook(excel_path)
    pd.DataFrame(
        [
            {
                "Dimension": "card_type",
                "Category": "CREDIT",
                "total": 1000.0,
                "approved": 500.0,
            }
        ]
    ).to_csv(csv_path, index=False)

    results = csv_validator.validate_dimension(
        "card_type",
        pd.read_csv(csv_path),
        pd.read_excel(excel_path, sheet_name="approval_card_type", header=2),
        ["approval"],
        tolerance=0.01,
        time_col=None,
        total_col="total",
        approval_col="approved",
        fraud_col=None,
    )
    assert results["failed"] > 0


def test_validator_passes_matching_share_csv_and_excel(tmp_path: Path) -> None:
    excel_path = tmp_path / "share_report.xlsx"
    csv_path = tmp_path / "share_report_balanced.csv"
    _write_minimal_share_workbook(excel_path)
    pd.DataFrame(
        [
            {
                "Dimension": "card_type",
                "Category": "CREDIT",
                "Balanced_txn_cnt": 1000.0,
                "Balanced_txn_cnt_Share_%": 12.5,
            }
        ]
    ).to_csv(csv_path, index=False)

    results = csv_validator.validate_dimension(
        "card_type",
        pd.read_csv(csv_path),
        pd.read_excel(excel_path, sheet_name="card_type", header=2),
        ["share"],
        tolerance=0.0001,
        time_col=None,
        total_col="Balanced_txn_cnt_Share_%",
        approval_col=None,
        fraud_col=None,
    )
    assert results["failed"] == 0
    assert results["passed"] == 1


def test_main_fails_when_no_dimensions_match(tmp_path: Path, monkeypatch) -> None:
    excel_path = tmp_path / "report.xlsx"
    csv_path = tmp_path / "report_balanced.csv"
    _write_minimal_rate_workbook(excel_path)
    pd.DataFrame([{"Dimension": "missing_dim", "Category": "A", "total": 1, "approved": 1}]).to_csv(
        csv_path, index=False
    )
    monkeypatch.setattr(
        sys,
        "argv",
        ["csv_validator.py", str(excel_path), str(csv_path)],
    )
    assert csv_validator.main() == 1


def test_main_accepts_share_export_csv(tmp_path: Path, monkeypatch) -> None:
    excel_path = tmp_path / "share_report.xlsx"
    csv_path = tmp_path / "share_balanced.csv"
    _write_minimal_share_workbook(excel_path)
    pd.DataFrame(
        [
            {
                "Dimension": "card_type",
                "Category": "CREDIT",
                "Balanced_txn_cnt": 1000.0,
                "Balanced_txn_cnt_Share_%": 12.5,
            }
        ]
    ).to_csv(csv_path, index=False)
    monkeypatch.setattr(
        sys,
        "argv",
        ["csv_validator.py", str(excel_path), str(csv_path)],
    )
    assert csv_validator.main() == 0
    assert csv_validator.is_share_export_csv(pd.read_csv(csv_path))
