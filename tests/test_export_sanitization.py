"""Tests for Excel/CSV formula-injection sanitization on export."""

from __future__ import annotations

import logging
import re
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from benchmark import run_share_analysis
from core.export_sanitizer import sanitize_cell


def test_sanitize_cell_formula_prefixes() -> None:
    assert sanitize_cell("=HYPERLINK('http://evil')") == "'=HYPERLINK('http://evil')"
    assert sanitize_cell("+cmd") == "'+cmd"
    assert sanitize_cell("-discount") == "'-discount"
    assert sanitize_cell("@SUM(A1)") == "'@SUM(A1)"
    assert sanitize_cell("\tleak") == "'\tleak"
    assert sanitize_cell("\rleak") == "'\rleak"


def test_sanitize_cell_safe_values_unchanged() -> None:
    assert sanitize_cell("BANCO SANTANDER") == "BANCO SANTANDER"
    assert sanitize_cell(12.5) == 12.5
    assert sanitize_cell(-3) == -3


def _malicious_gate_demo_csv(tmp_path: Path, old: str = "Online", new: str = "=2+5") -> Path:
    source = Path("tests/fixtures/gate_demo.csv").read_text(encoding="utf-8")
    malicious = source.replace(old, new)
    csv_path = tmp_path / "gate_demo_malicious.csv"
    csv_path.write_text(malicious, encoding="utf-8")
    return csv_path


def _share_args(output: Path, csv_path: Path) -> SimpleNamespace:
    return SimpleNamespace(
        csv=str(csv_path),
        df=None,
        metric="txn_cnt",
        secondary_metrics=None,
        entity="Target",
        entity_col="issuer_name",
        output=str(output),
        dimensions=["card_type", "channel"],
        auto=False,
        time_col="year_month",
        config=None,
        preset=None,
        debug=True,
        log_level="INFO",
        per_dimension_weights=False,
        export_balanced_csv=True,
        validate_input=False,
        compare_presets=False,
        analyze_distortion=False,
        output_format="analysis",
        include_calculated=False,
        auto_subset_search=None,
        subset_search_max_tests=None,
        trigger_subset_on_slack=None,
        max_cap_slack=None,
        compliance_posture=None,
        acknowledge_accuracy_first=False,
    )


def _workbook_has_raw_formula_string(workbook_path: Path) -> bool:
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        return True
        return False
    finally:
        workbook.close()


def test_share_export_neutralizes_malicious_category(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"
    csv_path = _malicious_gate_demo_csv(tmp_path)

    args = _share_args(output, csv_path)
    args.compliance_posture = "best_effort"
    result = run_share_analysis(
        args,
        logging.getLogger("test_export_sanitization"),
    )

    assert result == 0
    assert output.exists()

    assert not _workbook_has_raw_formula_string(output)

    balanced_csv = tmp_path / "share_balanced.csv"
    assert balanced_csv.exists()
    csv_text = balanced_csv.read_text(encoding="utf-8")
    assert "'=2+5" in csv_text
    assert not re.search(r"(?:^|,)=2\+5(?:,|$)", csv_text)


def test_share_export_neutralizes_malicious_entity_name(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"
    csv_path = _malicious_gate_demo_csv(tmp_path, old="Target", new="=2+5TARGET")

    args = _share_args(output, csv_path)
    args.entity = "=2+5TARGET"
    args.output_format = "both"
    args.compliance_posture = "best_effort"

    result = run_share_analysis(args, logging.getLogger("test_export_sanitization_entity"))

    assert result == 0
    assert output.exists()
    assert not _workbook_has_raw_formula_string(output)

    publication = tmp_path / "share_publication.xlsx"
    assert publication.exists()
    assert not _workbook_has_raw_formula_string(publication)
