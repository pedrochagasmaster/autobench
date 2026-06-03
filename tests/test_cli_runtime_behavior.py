import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.dimensional_analyzer import DimensionalAnalyzer
from core.privacy_validator import PrivacyValidator
from tests.fixtures.mock_benchmark_data import (
    write_insufficient_peer_csv,
    write_mock_benchmark_csv,
)


def test_mock_share_cli_exposes_missing_publication_and_diagnostic_sheets(tmp_path: Path) -> None:
    csv_path = write_mock_benchmark_csv(tmp_path / "mock.csv")
    output = tmp_path / "share_both.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            "benchmark.py",
            "share",
            "--metric",
            "txn_cnt",
            "--csv",
            str(csv_path),
            "--entity-col",
            "issuer_name",
            "--entity",
            "Target",
            "--dimensions",
            "card_type",
            "channel",
            "--time-col",
            "year_month",
            "--output",
            str(output),
            "--output-format",
            "both",
            "--debug",
            "--validate-input",
        ],
        cwd=Path.cwd(),
        capture_output=True,
        text=True,
        timeout=120,
    )

    assert result.returncode == 0
    assert output.exists()
    assert (tmp_path / "share_both_publication.xlsx").exists()
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Privacy Validation" in workbook.sheetnames
        assert "Peer Weights" in workbook.sheetnames
        assert "Weight Methods" in workbook.sheetnames
    finally:
        workbook.close()


def test_privacy_validator_matches_weighted_mock_behavior(tmp_path: Path) -> None:
    csv_path = write_mock_benchmark_csv(tmp_path / "mock.csv")
    df = pd.read_csv(csv_path)
    peers_df = df[df["issuer_name"] != "Target"]
    raw_online = peers_df[
        (peers_df["year_month"] == "2024-01") & (peers_df["channel"] == "Online")
    ].groupby("issuer_name", as_index=False)["txn_cnt"].sum()
    validator = PrivacyValidator(rule_name="6/30")

    raw_ok, raw_warnings = validator.validate_peer_group(raw_online, ["txn_cnt"], "issuer_name")

    assert raw_ok is False
    assert any("P1" in warning and "41.30%" in warning for warning in raw_warnings)

    analyzer = DimensionalAnalyzer(
        target_entity="Target",
        entity_column="issuer_name",
        time_column="year_month",
        debug_mode=True,
        consistent_weights=True,
    )
    analyzer.calculate_global_privacy_weights(df, "txn_cnt", ["card_type", "channel"])
    weights = {peer: data["multiplier"] for peer, data in analyzer.global_weights.items()}
    weighted_online = raw_online.copy()
    weighted_online["txn_cnt"] = weighted_online.apply(
        lambda row: row["txn_cnt"] * weights.get(row["issuer_name"], 1.0),
        axis=1,
    )

    weighted_ok, weighted_warnings = validator.validate_peer_group(
        weighted_online, ["txn_cnt"], "issuer_name"
    )
    validation_df = analyzer.build_privacy_validation_dataframe(df, "txn_cnt", ["card_type", "channel"])

    assert weighted_ok is True
    assert weighted_warnings == []
    assert int((validation_df["Compliant"] == "No").sum()) == 0
    assert int((validation_df["Dimension"] == "_TIME_TOTAL_").sum()) > 0


def test_mock_rate_cli_produces_expected_outputs(tmp_path: Path) -> None:
    csv_path = write_mock_benchmark_csv(tmp_path / "mock.csv")
    output = tmp_path / "rate_output.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            "benchmark.py",
            "rate",
            "--total-col",
            "total",
            "--approved-col",
            "approved",
            "--fraud-col",
            "fraud",
            "--privacy-basis",
            "clearing_spend",
            "--csv",
            str(csv_path),
            "--entity-col",
            "issuer_name",
            "--entity",
            "Target",
            "--dimensions",
            "card_type",
            "channel",
            "--time-col",
            "year_month",
            "--output",
            str(output),
            "--debug",
        ],
        cwd=Path.cwd(),
        capture_output=True,
        text=True,
        timeout=120,
    )

    assert result.returncode == 0
    assert output.exists()
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Privacy Validation" in workbook.sheetnames
        assert not any(name.startswith("Metric_") and "txn_cnt" in name for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_case01_regression_fixture_is_publishable_with_validation_enabled(tmp_path: Path) -> None:
    output = tmp_path / "case01_regression.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            "benchmark.py",
            "rate",
            "--csv",
            "tests/fixtures/case01_regression.csv",
            "--entity",
            "Target",
            "--entity-col",
            "issuer_name",
            "--dimensions",
            "product_group_year_month",
            "product_group_function_year_month",
            "--total-col",
            "amount_txn_pure_local_currency",
            "--approved-col",
            "amount_approved_local_currency",
            "--fraud-col",
            "amount_fraud_local_currency",
            "--privacy-basis",
            "clearing_spend",
            "--secondary-metrics",
            "count_txn_pure",
            "qt_declined",
            "--preset",
            "compliance_strict",
            "--output",
            str(output),
        ],
        cwd=Path.cwd(),
        capture_output=True,
        text=True,
        timeout=120,
    )

    assert result.returncode == 0, result.stderr
    assert "Compliance Verdict: fully_compliant" in result.stdout
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Data Quality" in workbook.sheetnames
        assert "Secondary Metrics" in workbook.sheetnames

        summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
        assert ("Compliance Verdict:", "fully_compliant") in summary_rows
        assert ("Input Validation:", "pass") in summary_rows

        dq_rows = list(workbook["Data Quality"].iter_rows(values_only=True))
        headers = dq_rows[6]
        assert headers[:5] == ("Severity", "Category", "Message", "Details", "Sample Rows")
        low_denom = next(row for row in dq_rows if row[1] == "low_denominator")
        assert '"affected_categories"' in low_denom[3]
        assert "<missing>" in low_denom[3]

        validation = pd.read_excel(output, sheet_name="Privacy Validation")
        assert int((validation["Compliant"] == "No").sum()) == 0
    finally:
        workbook.close()


def test_insufficient_peer_cli_aborts_even_without_input_validation(tmp_path: Path) -> None:
    csv_path = write_insufficient_peer_csv(tmp_path / "insufficient.csv")
    output = tmp_path / "insufficient.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            "benchmark.py",
            "share",
            "--metric",
            "txn_cnt",
            "--csv",
            str(csv_path),
            "--entity-col",
            "issuer_name",
            "--entity",
            "Target",
            "--dimensions",
            "card_type",
            "channel",
            "--output",
            str(output),
            "--output-format",
            "analysis",
            "--debug",
            "--no-validate-input",
        ],
        cwd=Path.cwd(),
        capture_output=True,
        text=True,
        timeout=120,
    )

    assert result.returncode == 1
    assert not output.exists()
    assert "Insufficient peers" in result.stderr or "Insufficient peers" in result.stdout
