import shlex
from pathlib import Path
from types import SimpleNamespace

import pytest

from scripts.perform_gate_test import GateTestRunner


def test_gate_command_parser_preserves_quoted_entity_names() -> None:
    command = 'py benchmark.py share --entity "BANCO SANTANDER" --csv data/readme_demo.csv'

    parsed = ["/usr/bin/python3"] + shlex.split(command[3:])

    assert parsed[3] == "--entity"
    assert parsed[4] == "BANCO SANTANDER"


def test_gate_generation_uses_portable_fixture(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    captured = {}

    def fake_run(cmd, cwd, capture_output, text):
        captured["cmd"] = cmd
        captured["cwd"] = cwd
        captured["capture_output"] = capture_output
        captured["text"] = text
        return SimpleNamespace(returncode=0, stderr="")

    monkeypatch.setattr("scripts.perform_gate_test.subprocess.run", fake_run)

    runner = GateTestRunner(output_dir=str(tmp_path / "gate"))
    runner.generate_cases()

    cmd = captured["cmd"]
    assert "--csv" in cmd
    csv_path = Path(cmd[cmd.index("--csv") + 1])
    assert not csv_path.is_absolute()
    assert csv_path == Path("tests") / "fixtures" / "gate_demo.csv"
    assert csv_path.name == "gate_demo.csv"
    assert csv_path.exists()


def test_verify_workbook_content_detects_excel_error_strings(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    workbook_path = tmp_path / "error_sheet.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "card_type"
    ws["A1"] = "Metric: card_type"
    ws["A3"] = "Category"
    ws["B3"] = "Target Share %"
    ws["A4"] = "A"
    ws["B4"] = "#DIV/0!"
    wb.save(workbook_path)

    runner = GateTestRunner(output_dir=str(tmp_path / "gate"))
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        failures = runner.verify_workbook_content(workbook, "share_gate_error", "share", {})
    finally:
        workbook.close()

    assert any("Contains Excel errors" in failure for failure in failures)


def test_verify_case_requires_bps_headers_for_fraud_publication(tmp_path: Path) -> None:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    output_dir = tmp_path / "gate"
    output_dir.mkdir()
    analysis_path = tmp_path / "rate.xlsx"
    publication_path = tmp_path / "rate_publication.xlsx"

    for path in (analysis_path, publication_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary" if "publication" in path.name else "Summary"
        detail = wb.create_sheet("approval_card_type")
        detail["A3"] = "Category"
        detail["B3"] = "Fraud Rate"
        detail["A4"] = "A"
        detail["B4"] = 1.25
        wb.save(path)

    runner = GateTestRunner(output_dir=str(output_dir))
    case = {
        "id": "rate_gate_bps",
        "params": {"output": str(analysis_path), "fraud_col": "fraud", "dimensions": ["approval_card_type"]},
        "expectations": ["analysis_workbook", "publication_workbook", "fraud_in_bps_in_publication"],
    }

    failures = runner.verify_case(case)

    assert "Fraud publication output is missing bps header" in failures


def test_all_generator_expectations_are_registered() -> None:
    from scripts.generate_cli_sweep import expectations_for_case
    from scripts.gate_expectations import resolve_expectation

    sample = expectations_for_case(
        {
            "output_format": "both",
            "export_balanced_csv": True,
            "include_calculated": True,
            "compare_presets": True,
            "analyze_distortion": True,
            "secondary_metrics": ["txn_cnt"],
            "per_dimension_weights": True,
            "validate_input": False,
            "entity": "Target",
            "fraud_col": "fraud",
            "fraud_in_bps": True,
        },
        "rate",
        Path("outputs/sample.xlsx"),
    )
    for token in sample:
        assert resolve_expectation(token) is not None, token
        spec = resolve_expectation(token)
        assert spec is not None
        if spec.status == "enforced" and not spec.prefix_match:
            assert spec.token == token
