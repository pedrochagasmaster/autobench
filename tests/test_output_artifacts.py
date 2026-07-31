import logging
import subprocess
import sys
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from benchmark import EXIT_STRICT_NON_COMPLIANT, run_share_analysis
from core.analysis_run import build_run_request, execute_share_run
from core.contracts import PrivacyOutputDecision
from core.output_artifacts import _flatten_rate_results, write_outputs
from core.privacy_output_policy import _attest_privacy_output


def test_flatten_rate_results_names_rate_and_dimension() -> None:
    approval = pd.DataFrame({"Rate": [1.0]})
    fraud = pd.DataFrame({"Rate": [2.0]})
    flattened = _flatten_rate_results(
        {"approval": {"channel": approval}, "fraud": {"card_type": fraud}}
    )
    assert list(flattened) == ["approval_channel", "fraud_card_type"]
    assert flattened["approval_channel"] is approval
    assert flattened["fraud_card_type"] is fraud


def test_core_modules_import_without_benchmark() -> None:
    # Run in a fresh interpreter: importing core must not pull in `benchmark`.
    # A subprocess avoids mutating this process's sys.modules (in-process
    # pop/reimport corrupts patch targets for later tests).
    code = (
        "import sys\n"
        "import core.analysis_run\n"
        "import core.output_artifacts\n"
        "assert 'benchmark' not in sys.modules, 'core imports must not pull in benchmark'\n"
    )
    result = subprocess.run(
        [sys.executable, "-c", code],
        cwd=Path(__file__).resolve().parent.parent,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr


def _share_args(output: Path, df: pd.DataFrame, output_format: str = "both") -> SimpleNamespace:
    return SimpleNamespace(
        csv="",
        df=df,
        metric="txn_cnt",
        secondary_metrics=None,
        entity="Target",
        entity_col="issuer_name",
        output=str(output),
        dimensions=["card_type"],
        auto=False,
        time_col=None,
        config=None,
        preset=None,
        debug=True,
        log_level="INFO",
        per_dimension_weights=False,
        export_balanced_csv=False,
        validate_input=False,
        compare_presets=False,
        analyze_distortion=False,
        output_format=output_format,
        include_calculated=False,
        auto_subset_search=None,
        subset_search_max_tests=None,
        trigger_subset_on_slack=None,
        max_cap_slack=None,
        compliance_posture=None,
        acknowledge_accuracy_first=False,
    )


def _share_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "issuer_name": ["Target", "P1", "P2", "P3", "P4", "P5", "P6"],
            "card_type": ["A", "A", "A", "A", "A", "A", "A"],
            "txn_cnt": [100, 200, 180, 160, 140, 120, 110],
        }
    )


def _gate_demo_df() -> pd.DataFrame:
    return pd.read_csv(Path("tests/fixtures/gate_demo.csv"))


def test_output_format_both_writes_analysis_and_publication(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"
    args = _share_args(output, _share_df(), output_format="both")
    args.compliance_posture = "best_effort"

    result = run_share_analysis(args, logging.getLogger("test_output"))

    assert result == 0
    assert output.exists()
    assert (tmp_path / "share_publication.xlsx").exists()


def test_no_validate_input_cannot_emit_fully_compliant_audit(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"

    result = run_share_analysis(
        _share_args(output, _share_df(), output_format="analysis"),
        logging.getLogger("test_no_validate_verdict"),
    )

    # Strict default posture + unchecked input is not publishable, so the CLI
    # now signals it via the strict non-compliant exit code.
    assert result == EXIT_STRICT_NON_COMPLIANT
    audit_log = tmp_path / "share_audit.log"
    assert audit_log.exists()
    audit_text = audit_log.read_text(encoding="utf-8")
    assert "compliance_verdict: not_publishable_input" in audit_text
    assert "data_quality_checked: False" in audit_text


def test_debug_workbook_contains_diagnostic_sheets(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"

    args = _share_args(output, _share_df(), output_format="analysis")
    args.compliance_posture = "best_effort"
    result = run_share_analysis(
        args,
        logging.getLogger("test_sheets"),
    )

    assert result == 0
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Peer Weights" in workbook.sheetnames
        assert "Weight Methods" in workbook.sheetnames
        assert "Privacy Validation" in workbook.sheetnames
        assert "Rank Changes" in workbook.sheetnames
    finally:
        workbook.close()


def test_analysis_workbook_keeps_weight_methods_when_privacy_sheet_disabled(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"
    config_path = tmp_path / "no_privacy_sheet.yaml"
    config_path.write_text(
        '\n'.join(
            [
                'version: "3.0"',
                'compliance_posture: strict',
                'output:',
                '  include_debug_sheets: false',
                '  include_privacy_validation: false',
            ]
        ),
        encoding="utf-8",
    )
    args = _share_args(output, _gate_demo_df(), output_format="analysis")
    args.config = str(config_path)
    args.debug = False
    args.dimensions = ["card_type", "channel"]
    args.time_col = "year_month"

    result = run_share_analysis(args, logging.getLogger("test_privacy_sheet_disabled"))

    # Strict posture + unchecked input exits with the strict non-compliant code
    # while still writing the analysis workbook this test inspects.
    assert result == EXIT_STRICT_NON_COMPLIANT
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Weight Methods" in workbook.sheetnames
        assert "Privacy Validation" not in workbook.sheetnames
    finally:
        workbook.close()


def test_privacy_failed_subset_search_diagnostics_are_not_written_to_workbook(
    tmp_path: Path,
) -> None:
    output = tmp_path / "share.xlsx"
    config_path = tmp_path / "subset_search.yaml"
    config_path.write_text(
        '\n'.join(
            [
                'version: "3.0"',
                'compliance_posture: strict',
                'optimization:',
                '  bounds:',
                '    max_weight: 1.0',
                '    min_weight: 0.9',
                '  linear_programming:',
                '    tolerance: 50.0',
                '  subset_search:',
                '    enabled: false',
                'output:',
                '  include_debug_sheets: false',
                '  include_privacy_validation: false',
            ]
        ),
        encoding="utf-8",
    )
    args = _share_args(output, _gate_demo_df(), output_format="analysis")
    args.config = str(config_path)
    args.debug = False
    args.dimensions = ["card_type", "channel"]
    args.time_col = "year_month"

    result = run_share_analysis(args, logging.getLogger("test_subset_lists"))

    assert result == EXIT_STRICT_NON_COMPLIANT
    assert not output.exists()
    assert list(
        tmp_path.glob("autobench_NON_PUBLISHABLE_control3_*.json")
    )


def test_output_format_publication_only_writes_publication_workbook(tmp_path: Path) -> None:
    """§2.6: --output-format publication must actually create the publication workbook
    (`generate_publication_workbook` was dead code on `main`).
    """
    output = tmp_path / "share.xlsx"
    args = _share_args(output, _share_df(), output_format="publication")
    args.compliance_posture = "best_effort"

    result = run_share_analysis(
        args,
        logging.getLogger("test_pub_only"),
    )

    assert result == 0
    assert (tmp_path / "share_publication.xlsx").exists()


def test_metric_sheet_names_use_plain_dimension_names(tmp_path: Path) -> None:
    """Q3: sheets are named after the dimension (`card_type`) not `Metric_1_card_type`,
    so the workbook matches pre-refactor archives and the CSV validator's Dimension
    column join.
    """
    output = tmp_path / "share.xlsx"

    args = _share_args(output, _share_df(), output_format="analysis")
    args.compliance_posture = "best_effort"
    result = run_share_analysis(
        args,
        logging.getLogger("test_sheet_names"),
    )

    assert result == 0
    workbook = load_workbook(output, read_only=True)
    try:
        assert "card_type" in workbook.sheetnames
        assert not any(name.startswith("Metric_") for name in workbook.sheetnames)
    finally:
        workbook.close()


def test_publication_workbook_includes_compliance_evidence_sheets(tmp_path: Path) -> None:
    """Q9: the publication workbook contains stakeholder-facing privacy evidence
    (Peer Weights, Privacy Validation, Rank Changes) but excludes solver-internal
    diagnostics (Weight Methods, Subset Search, Structural Diagnostics).
    """
    output = tmp_path / "share.xlsx"
    args = _share_args(output, _share_df(), output_format="both")
    args.compliance_posture = "best_effort"

    result = run_share_analysis(
        args,
        logging.getLogger("test_pub_scope"),
    )

    assert result == 0
    pub_path = tmp_path / "share_publication.xlsx"
    assert pub_path.exists()
    workbook = load_workbook(pub_path, read_only=True)
    try:
        sheetnames = workbook.sheetnames
        # Allow-list (Q9): stakeholder evidence belongs in publication.
        assert "Peer Weights" in sheetnames
        assert "Privacy Validation" in sheetnames
        # Solver internals stay analysis-only.
        assert "Weight Methods" not in sheetnames
        assert "Subset Search" not in sheetnames
        assert "Structural Detail" not in sheetnames
        assert "Data Quality" not in sheetnames
    finally:
        workbook.close()


def _artifacts_ready_for_write(
    tmp_path: Path,
    *,
    output_format: str = "both",
) -> tuple[SimpleNamespace, object]:
    """Run share analysis without writing outputs; return request and artifacts."""
    logger = logging.getLogger("test_publication_gate")
    args = _share_args(tmp_path / "capture.xlsx", _share_df(), output_format=output_format)
    request = build_run_request("share", args)
    with patch("core.analysis_run.write_outputs", lambda _req, art, **_kw: art):
        artifacts = execute_share_run(request, logger)
    return request, artifacts


def _write_with_posture(
    tmp_path: Path,
    *,
    output_format: str,
    posture: str,
    violations: int,
    stem: str,
) -> tuple[SimpleNamespace, object, Path, Path]:
    request, artifacts = _artifacts_ready_for_write(tmp_path, output_format=output_format)
    output_base = tmp_path / f"{stem}.xlsx"
    pub_path = output_base.with_name(f"{output_base.stem}_publication.xlsx")
    artifacts.analysis_output_file = str(output_base)
    artifacts.publication_output = str(pub_path)
    if artifacts.metadata is None:
        artifacts.metadata = {}
    summary = dict(artifacts.compliance_summary or {})
    summary["posture"] = posture
    summary["violations"] = violations
    summary["compliance_verdict"] = (
        "blocked" if violations else "fully_compliant"
    )
    artifacts.compliance_summary = summary
    artifacts.metadata["compliance_summary"] = summary
    write_outputs(
        request,
        artifacts,
        logger=logging.getLogger("test_publication_gate"),
        privacy_output_decision=artifacts.privacy_output_decision,
        privacy_output_attestation=_attest_privacy_output(
            artifacts.privacy_rule_strategy_result,
            artifacts.privacy_output_decision,
        ),
    )
    return request, artifacts, output_base, pub_path


def test_strict_posture_with_violations_withholds_publication_workbook(tmp_path: Path) -> None:
    _, artifacts, output_base, pub_path = _write_with_posture(
        tmp_path,
        output_format="both",
        posture="strict",
        violations=1,
        stem="strict_block",
    )

    assert output_base.exists()
    assert not pub_path.exists()
    assert artifacts.publication_output is None
    assert artifacts.metadata.get("publication_withheld_reason") == "strict_posture_violations"


def test_best_effort_posture_with_violations_still_writes_publication(tmp_path: Path) -> None:
    _, artifacts, output_base, pub_path = _write_with_posture(
        tmp_path,
        output_format="both",
        posture="best_effort",
        violations=1,
        stem="best_effort_pub",
    )

    assert output_base.exists()
    assert pub_path.exists()
    assert artifacts.publication_output == str(pub_path)
    assert artifacts.privacy_output_decision is not None
    assert artifacts.privacy_output_decision.privacy_publication_authorized
    assert "publication_withheld_reason" not in (artifacts.metadata or {})


def test_output_writer_requires_explicit_privacy_decision(tmp_path: Path) -> None:
    request, artifacts = _artifacts_ready_for_write(
        tmp_path,
        output_format="both",
    )
    output = tmp_path / "missing_decision.xlsx"
    artifacts.analysis_output_file = str(output)
    artifacts.publication_output = str(
        tmp_path / "missing_decision_publication.xlsx"
    )

    with pytest.raises(TypeError):
        write_outputs(  # type: ignore[call-arg]
            request,
            artifacts,
            logger=logging.getLogger("test_missing_privacy_decision"),
        )

    assert not output.exists()
    assert not (tmp_path / "missing_decision_publication.xlsx").exists()


def test_output_writer_rejects_inconsistent_privacy_decision(
    tmp_path: Path,
) -> None:
    with pytest.raises(ValueError, match="hard-blocked"):
        PrivacyOutputDecision(
            privacy_publication_authorized=False,
            hard_privacy_block=False,
        )

    request, artifacts = _artifacts_ready_for_write(
        tmp_path,
        output_format="both",
    )
    output = tmp_path / "forged_decision.xlsx"
    publication = tmp_path / "forged_decision_publication.xlsx"
    artifacts.analysis_output_file = str(output)
    artifacts.publication_output = str(publication)
    forged = SimpleNamespace(
        privacy_publication_authorized=False,
        hard_privacy_block=False,
        withholding_reason=None,
    )

    write_outputs(
        request,
        artifacts,
        logger=logging.getLogger("test_forged_privacy_decision"),
        privacy_output_decision=forged,  # type: ignore[arg-type]
        privacy_output_attestation=None,
    )

    assert artifacts.analysis_output_file is None
    assert artifacts.publication_output is None
    assert not output.exists()
    assert not publication.exists()


def test_output_writer_rejects_forged_authorized_bundle_and_clears_stale_paths(
    tmp_path: Path,
) -> None:
    request, artifacts = _artifacts_ready_for_write(
        tmp_path,
        output_format="both",
    )
    stale_fields = {
        "analysis_output_file": tmp_path / "stale.xlsx",
        "publication_output": tmp_path / "stale_publication.xlsx",
        "csv_output": tmp_path / "stale.csv",
        "json_output": tmp_path / "stale.json",
        "audit_log_output": tmp_path / "stale_audit.log",
        "audit_package_output": tmp_path / "stale_audit.zip",
    }
    for field_name, path in stale_fields.items():
        setattr(artifacts, field_name, str(path))
    artifacts.report_paths = [
        str(stale_fields["analysis_output_file"]),
        str(stale_fields["publication_output"]),
    ]
    artifacts.privacy_rule_strategy_result = SimpleNamespace(
        strategy="select_by_peer_count",
        is_anonymized_aggregated_merchant_spend=False,
        status="numerically_compliant",
        numeric_rules_passed=True,
        mandatory_overlays_passed=True,
        publication_authorized_by_numeric_policy=True,
        display_rule="5/25",
        feasible_candidate_rules=("5/25",),
        authorizing_rules=("5/25",),
        candidate_attempt_evaluations=(),
        emitted_output_evaluations=(),
        mandatory_overlay_evaluations=(),
        rule_set_digest="a" * 64,
        policy_version="v5 (2026-06-03)",
        policy_source=(
            "docs/control-3-customer-merchant-performance-v5-20260603.md"
        ),
    )
    artifacts.privacy_output_decision = PrivacyOutputDecision(
        privacy_publication_authorized=False,
        hard_privacy_block=True,
        withholding_reason="control3_numeric_policy_blocked",
    )
    forged_authorization = PrivacyOutputDecision(
        privacy_publication_authorized=True,
        hard_privacy_block=False,
    )

    write_outputs(
        request,
        artifacts,
        logger=logging.getLogger("test_forged_authorized_bundle"),
        privacy_output_decision=forged_authorization,
        privacy_output_attestation=None,
    )

    for field_name in stale_fields:
        assert getattr(artifacts, field_name) is None
    assert artifacts.report_paths == []
    assert not any(path.exists() for path in stale_fields.values())


def test_strict_posture_without_violations_writes_publication(tmp_path: Path) -> None:
    _, artifacts, output_base, pub_path = _write_with_posture(
        tmp_path,
        output_format="both",
        posture="strict",
        violations=0,
        stem="strict_compliant",
    )

    assert output_base.exists()
    assert pub_path.exists()
    assert artifacts.publication_output == str(pub_path)
    assert "publication_withheld_reason" not in (artifacts.metadata or {})


def test_strict_posture_publication_only_with_violations_writes_no_workbook(tmp_path: Path) -> None:
    _, artifacts, output_base, pub_path = _write_with_posture(
        tmp_path,
        output_format="publication",
        posture="strict",
        violations=1,
        stem="strict_pub_only",
    )

    assert not output_base.exists()
    assert not pub_path.exists()
    assert artifacts.publication_output is None
    assert artifacts.metadata.get("publication_withheld_reason") == "strict_posture_violations"
