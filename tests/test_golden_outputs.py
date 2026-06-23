"""Golden output regression tests for business-visible report invariants."""

from __future__ import annotations

import subprocess
import sys
import zipfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
FIXTURE = ROOT / "tests" / "fixtures" / "gate_demo.csv"
PY = sys.executable


def _run_benchmark(args: list[str], cwd: Path, output_xlsx: Path) -> subprocess.CompletedProcess[str]:
    cmd = [
        PY,
        str(ROOT / "benchmark.py"),
        *args,
        "--csv",
        str(FIXTURE),
        "--entity",
        "Target",
        "--dimensions",
        "card_type",
        "channel",
        "--time-col",
        "year_month",
        "--preset",
        "balanced_default",
        "--output",
        str(output_xlsx),
    ]
    return subprocess.run(cmd, cwd=cwd, capture_output=True, text=True, check=False)


def _assert_core_workbook_invariants(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        assert "Summary" in wb.sheetnames
        assert "Weight Methods" in wb.sheetnames
        assert "Rank Changes" in wb.sheetnames
        reserved = {
            "Summary",
            "Metadata",
            "Weight Methods",
            "Rank Changes",
            "Peer Weights",
            "Privacy Validation",
            "Preset Comparison",
            "Impact Summary",
            "Impact Detail",
            "Subset Search",
            "Structural Summary",
            "Structural Detail",
        }
        dim_sheets = [name for name in wb.sheetnames if name not in reserved]
        assert dim_sheets, "expected at least one dimension sheet"
    finally:
        wb.close()


def _summary_metadata(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Summary"]
        metadata: dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).rstrip(":")
            if len(row) > 1 and row[1] is not None:
                metadata[key] = str(row[1])
        return metadata
    finally:
        wb.close()


def _assert_compliant_summary(path: Path) -> None:
    metadata = _summary_metadata(path)
    assert metadata["Compliance Posture"] == "best_effort"
    assert metadata["Compliance Verdict"] == "fully_compliant"
    assert metadata["Run Status"] == "compliant"
    assert metadata["Input Validation"] == "pass"


def _assert_peer_only_summary(path: Path) -> None:
    metadata = _summary_metadata(path)
    assert metadata["Entity"] == "PEER-ONLY"
    assert metadata["Compliance Posture"] == "best_effort"
    assert metadata["Compliance Verdict"] in {"fully_compliant", "violations_detected"}
    # best_effort posture reports violations as completed_with_warnings;
    # "non_compliant" only exists under the strict posture.
    assert metadata["Run Status"] in {"compliant", "completed_with_warnings"}
    assert metadata["Input Validation"] == "pass"


@pytest.mark.parametrize(
    "mode,extra_args,csv_suffix",
    [
        ("share", ["share", "--metric", "txn_cnt"], ""),
        (
            "rate",
            [
                "rate",
                "--total-col",
                "total",
                "--approved-col",
                "approved",
                "--fraud-col",
                "fraud",
                "--privacy-basis",
                "clearing_spend",
            ],
            "_balanced.csv",
        ),
    ],
)
def test_golden_cli_outputs(tmp_path: Path, mode: str, extra_args: list[str], csv_suffix: str) -> None:
    output_xlsx = tmp_path / f"golden_{mode}.xlsx"
    result = _run_benchmark(extra_args, tmp_path, output_xlsx)
    assert result.returncode == 0, result.stderr or result.stdout
    assert output_xlsx.exists()
    _assert_core_workbook_invariants(output_xlsx)
    _assert_compliant_summary(output_xlsx)

    if csv_suffix:
        export_cmd = [
            PY,
            str(ROOT / "benchmark.py"),
            *extra_args,
            "--csv",
            str(FIXTURE),
            "--entity",
            "Target",
            "--dimensions",
            "card_type",
            "channel",
            "--time-col",
            "year_month",
            "--preset",
            "balanced_default",
            "--output",
            str(tmp_path / f"golden_{mode}_csv.xlsx"),
            "--export-balanced-csv",
        ]
        export_result = subprocess.run(export_cmd, cwd=tmp_path, capture_output=True, text=True, check=False)
        assert export_result.returncode == 0, export_result.stderr or export_result.stdout
        exported = list(tmp_path.glob("*_balanced.csv"))
        assert exported, "expected balanced CSV export"
        csv_df = pd.read_csv(exported[0])
        assert {"Dimension", "Category"}.issubset(csv_df.columns)


def test_golden_peer_only_share(tmp_path: Path) -> None:
    output_xlsx = tmp_path / "golden_peer_only.xlsx"
    cmd = [
        PY,
        str(ROOT / "benchmark.py"),
        "share",
        "--csv",
        str(FIXTURE),
        "--metric",
        "txn_cnt",
        "--dimensions",
        "card_type",
        "channel",
        "--time-col",
        "year_month",
        "--preset",
        "balanced_default",
        "--output",
        str(output_xlsx),
    ]
    result = subprocess.run(cmd, cwd=tmp_path, capture_output=True, text=True, check=False)
    assert result.returncode == 0, result.stderr or result.stdout
    _assert_core_workbook_invariants(output_xlsx)
    _assert_peer_only_summary(output_xlsx)


def test_golden_publication_output(tmp_path: Path) -> None:
    output_xlsx = tmp_path / "golden_publication.xlsx"
    result = _run_benchmark(["share", "--metric", "txn_cnt", "--output-format", "publication"], tmp_path, output_xlsx)
    publication_path = tmp_path / "golden_publication_publication.xlsx"

    assert result.returncode == 0, result.stderr or result.stdout
    assert publication_path.exists()
    wb = load_workbook(publication_path, read_only=True, data_only=True)
    try:
        assert "Executive Summary" in wb.sheetnames
        assert any(name not in {"Summary", "Methodology"} for name in wb.sheetnames)
    finally:
        wb.close()


def test_audit_package_contains_expected_artifacts(tmp_path: Path) -> None:
    output_xlsx = tmp_path / "audit_package_share.xlsx"
    result = _run_benchmark(
        ["share", "--metric", "txn_cnt", "--export-balanced-csv", "--audit-package"],
        tmp_path,
        output_xlsx,
    )
    package_path = tmp_path / "audit_package_share_audit_package.zip"

    assert result.returncode == 0, result.stderr or result.stdout
    assert package_path.exists()
    with zipfile.ZipFile(package_path) as zf:
        names = set(zf.namelist())
        assert "audit_package_share.xlsx" in names
        assert "audit_package_share_balanced.csv" in names
        assert "audit_package_share_audit.log" in names
        assert "config_snapshot.json" in names
        assert "validation_summary.json" in names
