"""In-process integration tests for core.analysis_run orchestration.

Minimal required fields (AnalysisRunRequest defaults cover the rest):

Share run:
  - csv (or pre-loaded df)
  - entity (target name in data)
  - metric (column present in data)
  - dimensions (list, or auto=True with auto-detect enabled in config)
  - output (workbook path)
  - compliance_posture when preset is set (see note below)

Rate run (mode defaults to "share"; must set mode="rate"):
  - csv (or df)
  - entity
  - total_col, and at least one of approved_col / fraud_col (columns in data)
  - dimensions
  - output
  - export_balanced_csv=True when asserting balanced CSV output
  - control3_overrides privacy_basis=clearing_spend when fraud_col is set

When preset is set, also pass compliance_posture explicitly (matches TUI; avoids
ConfigManager material-override guard on per_dimension_weights default).
"""

from __future__ import annotations

import json
import logging
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

import benchmark
from core.analysis_run import execute_rate_run, execute_share_run
from core.analysis_run import RunAborted, RunBlocked
from core.contracts import (
    AnalysisRunRequest,
    PrivacyEvaluationStatus,
    PrivacyRuleStrategy,
    PrivacySweepStatus,
)
from core.dimensional_analyzer import DimensionalAnalyzer
from utils.logger import finalize_deferred_logging, setup_deferred_logging

FIXTURE = Path(__file__).parent / "fixtures" / "gate_demo.csv"
SHARE_DIMENSIONS = ["card_type", "channel"]


def _single_category_df(shares: list[float]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"issuer_name": f"P{index}", "segment": "all", "amount": share}
            for index, share in enumerate(shares, start=1)
        ]
    )


def test_share_run_end_to_end(tmp_path: Path) -> None:
    out = tmp_path / "share_it.xlsx"
    request = AnalysisRunRequest(
        csv=str(FIXTURE),
        entity="Target",
        metric="txn_cnt",
        dimensions=SHARE_DIMENSIONS,
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        output=str(out),
    )
    artifacts = execute_share_run(request, logging.getLogger("test"))

    assert out.exists()
    wb = load_workbook(out)
    sheet_names = set(wb.sheetnames)
    assert "Summary" in sheet_names
    assert "Weight Methods" in sheet_names
    assert "Rank Changes" in sheet_names
    for dim in SHARE_DIMENSIONS:
        assert dim in sheet_names

    assert artifacts.compliance_summary is not None
    assert artifacts.compliance_summary["compliance_verdict"] == "fully_compliant"
    assert artifacts.privacy_rule_strategy_result is not None
    assert (
        artifacts.privacy_rule_strategy_result.strategy
        == PrivacyRuleStrategy.SELECT_BY_PEER_COUNT
    )

    expected = {out.name, f"{out.stem}_audit.log"}
    assert {p.name for p in tmp_path.iterdir()} == expected


def test_share_run_sweep_optimizes_each_applicable_rule(tmp_path: Path) -> None:
    out = tmp_path / "share_sweep.xlsx"
    request = AnalysisRunRequest(
        csv=str(FIXTURE),
        entity="Target",
        metric="txn_cnt",
        dimensions=SHARE_DIMENSIONS,
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        output=str(out),
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )
    artifacts = execute_share_run(request, logging.getLogger("test"))

    sweep = artifacts.privacy_rule_strategy_result
    assert sweep is not None
    assert sweep.strategy == PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE
    assert len(sweep.rule_set_digest) == 64
    assert tuple(e.rule_name for e in sweep.candidate_attempt_evaluations) == (
        "5/25", "6/30", "7/35", "10/40", "4/35"
    )
    assert sweep.display_rule == artifacts.analyzer.privacy_rule_name
    assert set(sweep.authorizing_rules).issubset({"5/25", "6/30", "7/35"})
    assert artifacts.compliance_summary is not None
    assert artifacts.compliance_summary["compliance_verdict"] == "fully_compliant"


def test_sweep_authorizers_are_recomputed_against_emitted_candidate(tmp_path: Path) -> None:
    request = AnalysisRunRequest(
        df=_single_category_df([34, 12, 12, 8, 7, 7, 6, 5, 5, 4]),
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / "divergent.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))
    sweep = artifacts.privacy_rule_strategy_result

    assert sweep is not None
    assert sweep.display_rule == "10/40"
    assert "5/25" in sweep.feasible_candidate_rules
    assert "5/25" not in sweep.authorizing_rules
    assert "10/40" in sweep.authorizing_rules
    emitted = {
        evaluation.rule_name: evaluation.status.value
        for evaluation in sweep.emitted_output_evaluations
    }
    assert emitted["5/25"] == "failed"
    assert emitted["10/40"] == "passed"
    assert {
        evaluation.rule_name
        for evaluation in sweep.emitted_output_evaluations
        if evaluation.status.value != "not_applicable"
    } == {"5/25", "6/30", "7/35", "10/40"}
    assert artifacts.analyzer.privacy_rule_name == sweep.display_rule
    assert artifacts.metadata is not None
    assert artifacts.metadata["privacy_rule_strategy"]["display_rule"] == sweep.display_rule


def test_sweep_selects_only_feasible_merchant_4_35_candidate(tmp_path: Path) -> None:
    request = AnalysisRunRequest(
        df=_single_category_df([34, 22, 22, 22]),
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / "merchant.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        is_anonymized_aggregated_merchant_spend=True,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))
    sweep = artifacts.privacy_rule_strategy_result

    assert sweep is not None
    assert sweep.feasible_candidate_rules == ("4/35",)
    assert sweep.authorizing_rules == ("4/35",)
    assert sweep.display_rule == "4/35"


def test_merchant_4_35_context_reaches_normal_input_validation(tmp_path: Path) -> None:
    request = AnalysisRunRequest(
        df=_single_category_df([34, 22, 22, 22]),
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / "merchant_validated.xlsx"),
        compliance_posture="strict",
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        is_anonymized_aggregated_merchant_spend=True,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))

    assert artifacts.privacy_rule_strategy_result is not None
    assert artifacts.privacy_rule_strategy_result.authorizing_rules == ("4/35",)


@pytest.mark.parametrize(
    ("output_format", "export_csv", "report_format"),
    [
        ("analysis", False, "xlsx"),
        ("both", False, "xlsx"),
        ("analysis", True, "xlsx"),
        ("analysis", False, "json"),
    ],
)
def test_sole_4_35_blocks_analysis_bearing_artifacts(
    tmp_path: Path,
    output_format: str,
    export_csv: bool,
    report_format: str,
) -> None:
    artifacts = execute_share_run(
        AnalysisRunRequest(
            df=_single_category_df([34, 22, 22, 22]),
            csv="",
            metric="amount",
            dimensions=["segment"],
            output=str(tmp_path / "merchant_blocked.xlsx"),
            output_format=output_format,
            report_format=report_format,
            export_balanced_csv=export_csv,
            debug=True,
            compliance_posture="best_effort",
            validate_input=False,
            privacy_rule_strategy=(
                PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE
            ),
            is_anonymized_aggregated_merchant_spend=True,
        ),
        logging.getLogger("test"),
    )

    assert artifacts.privacy_sink_authorized is False
    assert artifacts.privacy_log_authorized is False
    assert not list(tmp_path.glob("*.xlsx"))
    assert not list(tmp_path.glob("*_balanced.csv"))
    assert not list(tmp_path.glob("merchant_blocked.json"))


def test_sole_4_35_publication_is_anonymized_aggregate_only(
    tmp_path: Path,
) -> None:
    output = tmp_path / "merchant_publication.xlsx"
    merchant_df = pd.concat(
        [
            pd.DataFrame(
                [
                    {
                        "issuer_name": "Target",
                        "segment": "all",
                        "amount": 100,
                    }
                ]
            ),
            _single_category_df([34, 22, 22, 22]),
        ],
        ignore_index=True,
    )
    artifacts = execute_share_run(
        AnalysisRunRequest(
            df=merchant_df,
            csv="",
            entity="Target",
            metric="amount",
            dimensions=["segment"],
            output=str(output),
            output_format="publication",
            debug=True,
            compliance_posture="best_effort",
            validate_input=False,
            privacy_rule_strategy=(
                PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE
            ),
            is_anonymized_aggregated_merchant_spend=True,
        ),
        logging.getLogger("test"),
    )

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.privacy_log_authorized is False
    assert artifacts.analysis_output_file is None
    assert artifacts.publication_output is not None
    publication = Path(artifacts.publication_output)
    assert publication.exists()
    workbook = load_workbook(publication, read_only=True)
    try:
        assert "Peer Weights" not in workbook.sheetnames
        assert "Privacy Validation" not in workbook.sheetnames
        text = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
    finally:
        workbook.close()
    for forbidden in ("P1", "P2", "P3", "P4", "BIC", "Target"):
        assert forbidden not in text
    assert artifacts.weights_df is None
    assert artifacts.privacy_validation_df is None
    assert artifacts.analyzer is None
    assert artifacts.report_model is None
    for frame in artifacts.results.values():
        if isinstance(frame, pd.DataFrame):
            assert not any(
                token in str(column).casefold()
                for column in frame.columns
                for token in ("target", "bic", "original")
            )


def test_default_strategy_preserves_legacy_yaml_merchant_mode(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "merchant_mode.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "version": "3.0",
                "compliance_posture": "best_effort",
                "analysis": {"merchant_mode": True},
            }
        ),
        encoding="utf-8",
    )
    artifacts = execute_share_run(
        AnalysisRunRequest(
            df=_single_category_df([34, 22, 22, 22]),
            csv="",
            metric="amount",
            dimensions=["segment"],
            config=str(config_path),
            output=str(tmp_path / "legacy_merchant.xlsx"),
            output_format="publication",
            compliance_posture="best_effort",
            validate_input=False,
        ),
        logging.getLogger("test"),
    )

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.privacy_rule_strategy_result is not None
    assert (
        artifacts.privacy_rule_strategy_result
        .is_anonymized_aggregated_merchant_spend
        is True
    )
    assert artifacts.privacy_rule_strategy_result.authorizing_rules == (
        "4/35",
    )


def test_default_strategy_honors_explicit_merchant_spend_fact(
    tmp_path: Path,
) -> None:
    artifacts = execute_share_run(
        AnalysisRunRequest(
            df=_single_category_df([34, 22, 22, 22]),
            csv="",
            metric="amount",
            dimensions=["segment"],
            output=str(tmp_path / "explicit_merchant.xlsx"),
            output_format="publication",
            compliance_posture="best_effort",
            validate_input=False,
            is_anonymized_aggregated_merchant_spend=True,
        ),
        logging.getLogger("test"),
    )

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.privacy_rule_strategy_result is not None
    assert artifacts.privacy_rule_strategy_result.authorizing_rules == (
        "4/35",
    )


def test_sweep_cannot_authorize_when_every_candidate_attempt_raises(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_fit(*_args, **_kwargs):
        raise ValueError("forced optimizer failure")

    monkeypatch.setattr(DimensionalAnalyzer, "fit_privacy_weights", fail_fit)
    request = AnalysisRunRequest(
        df=_single_category_df([25, 25, 20, 15, 15]),
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / "all_attempts_fail.xlsx"),
        compliance_posture="best_effort",
        validate_input=False,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))
    sweep = artifacts.privacy_rule_strategy_result

    assert sweep is not None
    assert not sweep.feasible_candidate_rules
    assert not sweep.numeric_rules_passed
    assert not sweep.authorizing_rules
    assert sweep.status == PrivacySweepStatus.NUMERICALLY_NONCOMPLIANT
    assert all(
        evaluation.status.value != "passed"
        for evaluation in sweep.emitted_output_evaluations
    )
    assert artifacts.analysis_output_file is None
    assert artifacts.publication_output is None
    assert artifacts.csv_output is None
    assert artifacts.json_output is None
    assert artifacts.report_paths == []
    assert not list(tmp_path.glob("*.xlsx"))
    assert not list(tmp_path.glob("*.csv"))


def test_citi_overlay_blocks_emitted_candidate_that_passes_base_rule(tmp_path: Path) -> None:
    primary = [20, 20, 10, 10, 10, 6, 6, 6, 6, 6]
    secondary = [30, 20, 10, 10, 10, 5, 5, 4, 3, 3]
    df = _single_category_df(primary)
    df["secondary"] = secondary
    request = AnalysisRunRequest(
        df=df,
        csv="",
        metric="amount",
        secondary_metrics=["secondary"],
        dimensions=["segment"],
        output=str(tmp_path / "citi_block.xlsx"),
        compliance_posture="strict",
        output_format="both",
        validate_input=False,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        citibank_entity_name="P1",
        citi_competitor_receives_output=True,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))
    sweep = artifacts.privacy_rule_strategy_result

    assert sweep is not None
    assert sweep.numeric_rules_passed
    assert not sweep.mandatory_overlays_passed
    assert not sweep.authorizing_rules
    assert sweep.status == PrivacySweepStatus.BLOCKED_BY_MANDATORY_OVERLAY
    assert sweep.mandatory_overlay_evaluations[0].status.value == "failed"
    assert (
        sweep.mandatory_overlay_evaluations[0].failure_reasons[0].code
        == "citibank_concentration_exceeded"
    )
    assert artifacts.compliance_summary is not None
    assert artifacts.compliance_summary["compliance_verdict"] != "fully_compliant"
    assert artifacts.analysis_output_file is None
    assert artifacts.publication_output is None
    assert artifacts.audit_log_output is not None
    audit_payload = json.loads(
        Path(artifacts.audit_log_output).read_text(encoding="utf-8")
    )
    assert audit_payload["artifact_type"] == (
        "non_publishable_control3_privacy_audit"
    )
    assert audit_payload["publishable"] is False
    assert audit_payload["withholding_reason"] == (
        "control3_mandatory_overlay_blocked"
    )
    assert audit_payload["authorizing_rules"] == []
    assert (
        benchmark._resolve_exit_code(
            "strict",
            artifacts.compliance_summary["compliance_verdict"],
        )
        == benchmark.EXIT_STRICT_NON_COMPLIANT
    )


def test_default_strategy_still_blocks_failed_citi_overlay(tmp_path: Path) -> None:
    primary = [20, 20, 10, 10, 10, 6, 6, 6, 6, 6]
    df = _single_category_df(primary)
    df["secondary"] = [30, 20, 10, 10, 10, 5, 5, 4, 3, 3]
    request = AnalysisRunRequest(
        df=df,
        csv="",
        metric="amount",
        secondary_metrics=["secondary"],
        dimensions=["segment"],
        output=str(tmp_path / "default_citi_block.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        citibank_entity_name="P1",
        citi_competitor_receives_output=True,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))

    assert artifacts.privacy_rule_strategy_result is not None
    assert not artifacts.privacy_rule_strategy_result.mandatory_overlays_passed
    assert artifacts.compliance_summary is not None
    assert artifacts.compliance_summary["compliance_verdict"] != "fully_compliant"


def test_cli_sweep_exit_audit_and_publication_agree_when_citi_blocks(
    tmp_path: Path,
) -> None:
    df = _single_category_df([20, 20, 10, 10, 10, 6, 6, 6, 6, 6])
    df["secondary"] = [30, 20, 10, 10, 10, 5, 5, 4, 3, 3]
    csv_path = tmp_path / "citi_cli.csv"
    output = tmp_path / "citi_cli.xlsx"
    df.to_csv(csv_path, index=False)
    args = benchmark.create_parser().parse_args(
        [
            "share",
            "--csv",
            str(csv_path),
            "--metric",
            "amount",
            "--secondary-metrics",
            "secondary",
            "--dimensions",
            "segment",
            "--compliance-posture",
            "strict",
            "--output",
            str(output),
            "--output-format",
            "both",
            "--privacy-rule-sweep",
            "--citibank-entity-name",
            "P1",
            "--citi-competitor-receives-output",
        ]
    )

    exit_code = benchmark.run_share_analysis(args, logging.getLogger("test"))

    assert exit_code == benchmark.EXIT_STRICT_NON_COMPLIANT
    assert not output.exists()
    assert not list(tmp_path.glob("*publication*.xlsx"))
    audit_path = next(
        tmp_path.glob("autobench_NON_PUBLISHABLE_control3_*.json")
    )
    audit_payload = json.loads(audit_path.read_text(encoding="utf-8"))
    assert audit_payload["withholding_reason"] == (
        "control3_mandatory_overlay_blocked"
    )
    assert audit_payload["authorizing_rules"] == []


@pytest.mark.parametrize("citi_name", ["missing", "Citibank"])
def test_integrated_citi_identity_fails_closed(
    tmp_path: Path,
    citi_name: str | None,
) -> None:
    names = ["Citibank", "CITIBANK", "P3", "P4", "P5"]
    df = pd.DataFrame(
        [
            {"issuer_name": name, "segment": "all", "amount": 20}
            for name in names
        ]
    )
    request = AnalysisRunRequest(
        df=df,
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / "citi_invalid.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        citibank_entity_name=citi_name,
        citi_competitor_receives_output=True,
    )

    with pytest.raises(RunBlocked):
        execute_share_run(request, logging.getLogger("test"))


@pytest.mark.parametrize(
    "strategy",
    [
        PrivacyRuleStrategy.SELECT_BY_PEER_COUNT,
        PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    ],
)
def test_integrated_competitor_recipient_without_citi_is_not_applicable(
    tmp_path: Path,
    strategy: PrivacyRuleStrategy,
) -> None:
    request = AnalysisRunRequest(
        df=_single_category_df([25, 25, 20, 15, 15]),
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / f"citi_absent_{strategy.value}.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        privacy_rule_strategy=strategy,
        citibank_entity_name=None,
        citi_competitor_receives_output=True,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.privacy_rule_strategy_result is not None
    overlay = artifacts.privacy_rule_strategy_result.mandatory_overlay_evaluations[0]
    assert overlay.status == PrivacyEvaluationStatus.NOT_APPLICABLE


def test_integrated_observational_citi_identity_does_not_trigger_overlay(
    tmp_path: Path,
) -> None:
    request = AnalysisRunRequest(
        df=_single_category_df([25, 25, 20, 15, 15]),
        csv="",
        metric="amount",
        dimensions=["segment"],
        output=str(tmp_path / "citi_observational.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        citibank_entity_name="P1",
        citi_competitor_receives_output=False,
    )

    artifacts = execute_share_run(request, logging.getLogger("test"))

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.privacy_rule_strategy_result is not None
    overlay = artifacts.privacy_rule_strategy_result.mandatory_overlay_evaluations[0]
    assert overlay.status == PrivacyEvaluationStatus.NOT_APPLICABLE


def test_rate_run_end_to_end(tmp_path: Path) -> None:
    out = tmp_path / "rate_it.xlsx"
    request = AnalysisRunRequest(
        mode="rate",
        csv=str(FIXTURE),
        entity="Target",
        total_col="total",
        approved_col="approved",
        fraud_col="fraud",
        privacy_concentration_col="total",
        dimensions=SHARE_DIMENSIONS,
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        control3_overrides={"privacy_basis": "clearing_spend"},
        output=str(out),
        export_balanced_csv=True,
    )
    artifacts = execute_rate_run(request, logging.getLogger("test"))

    assert out.exists()
    csv_path = tmp_path / f"{out.stem}_balanced.csv"
    assert csv_path.exists()
    assert artifacts.csv_output == str(csv_path)

    df = pd.read_csv(csv_path)
    assert "Dimension" in df.columns
    assert "Category" in df.columns
    assert len(df) >= 1

    expected = {out.name, csv_path.name, f"{out.stem}_audit.log"}
    assert {p.name for p in tmp_path.iterdir()} == expected


def test_python_rate_sweep_runs_through_shared_executor(tmp_path: Path) -> None:
    request = AnalysisRunRequest(
        mode="rate",
        csv=str(FIXTURE),
        entity="Target",
        total_col="total",
        approved_col="approved",
        dimensions=SHARE_DIMENSIONS,
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        output=str(tmp_path / "rate_sweep.xlsx"),
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )

    artifacts = execute_rate_run(request, logging.getLogger("test"))

    assert artifacts.privacy_rule_strategy_result is not None
    assert (
        artifacts.privacy_rule_strategy_result.strategy
        == PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE
    )


def test_fraud_rate_sweep_requires_explicit_concentration_column(tmp_path: Path) -> None:
    request = AnalysisRunRequest(
        mode="rate",
        csv=str(FIXTURE),
        entity="Target",
        total_col="total",
        fraud_col="fraud",
        dimensions=SHARE_DIMENSIONS,
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        control3_overrides={"privacy_basis": "clearing_spend"},
        output=str(tmp_path / "fraud_sweep.xlsx"),
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )

    with pytest.raises(RunBlocked):
        execute_rate_run(request, logging.getLogger("test"))


def test_fraud_only_sweep_uses_clearing_spend_not_total_distribution(
    tmp_path: Path,
) -> None:
    df = pd.DataFrame(
        [
            {
                "issuer_name": f"P{index}",
                "segment": "all",
                "total": total,
                "fraud": 1,
                "clearing_spend": clearing,
            }
            for index, (total, clearing) in enumerate(
                zip([80, 5, 5, 5, 5], [20, 20, 20, 20, 20]),
                start=1,
            )
        ]
    )
    request = AnalysisRunRequest(
        mode="rate",
        df=df,
        csv="",
        total_col="total",
        fraud_col="fraud",
        privacy_concentration_col="clearing_spend",
        dimensions=["segment"],
        output=str(tmp_path / "fraud_basis.xlsx"),
        compliance_posture="strict",
        validate_input=False,
        control3_overrides={"privacy_basis": "clearing_spend"},
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )

    artifacts = execute_rate_run(request, logging.getLogger("test"))

    assert artifacts.privacy_rule_strategy_result is not None
    assert artifacts.privacy_rule_strategy_result.authorizing_rules == ("5/25",)


def test_approval_and_fraud_sweep_governs_total_and_clearing_spend(
    tmp_path: Path,
) -> None:
    df = pd.DataFrame(
        [
            {
                "issuer_name": f"P{index}",
                "segment": "all",
                "total": total,
                "approved": total * 0.8,
                "fraud": 1,
                "clearing_spend": clearing,
            }
            for index, (total, clearing) in enumerate(
                zip([80, 5, 5, 5, 5], [20, 20, 20, 20, 20]),
                start=1,
            )
        ]
    )
    request = AnalysisRunRequest(
        mode="rate",
        df=df,
        csv="",
        total_col="total",
        approved_col="approved",
        fraud_col="fraud",
        privacy_concentration_col="clearing_spend",
        dimensions=["segment"],
        output=str(tmp_path / "approval_fraud_bases.xlsx"),
        output_format="both",
        export_balanced_csv=True,
        audit_package=True,
        compliance_posture="best_effort",
        validate_input=False,
        control3_overrides={"privacy_basis": "clearing_spend"},
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
    )

    artifacts = execute_rate_run(request, logging.getLogger("test"))

    assert artifacts.privacy_rule_strategy_result is not None
    assert not artifacts.privacy_rule_strategy_result.numeric_rules_passed
    assert not artifacts.privacy_rule_strategy_result.authorizing_rules
    assert artifacts.analysis_output_file is None
    assert artifacts.report_paths == []
    assert not list(tmp_path.glob("*.xlsx"))
    assert not list(tmp_path.glob("*_balanced.csv"))
    assert not list(tmp_path.glob("*.zip"))


@pytest.mark.parametrize(
    ("strategy", "clearing_b"),
    [
        (PrivacyRuleStrategy.SELECT_BY_PEER_COUNT, [25, 25, 25, 25, 0]),
        (PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE, [25, 25, 25, 25, 0]),
        (PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE, [0, 0, 0, 0, 0]),
    ],
)
def test_fraud_output_omits_category_with_underpopulated_clearing_basis(
    tmp_path: Path,
    strategy: PrivacyRuleStrategy,
    clearing_b: list[int],
) -> None:
    rows: list[dict[str, object]] = []
    for category, clearing_values in (
        ("SAFE_A", [20, 20, 20, 20, 20]),
        ("SECRET_B", clearing_b),
    ):
        for index, clearing in enumerate(clearing_values, start=1):
            rows.append(
                {
                    "issuer_name": f"P{index}",
                    "segment": category,
                    "total": 20,
                    "fraud": 9999 if category == "SECRET_B" else 1,
                    "clearing_spend": clearing,
                }
            )
    output = tmp_path / f"fraud_suppressed_{strategy.value}.xlsx"
    artifacts = execute_rate_run(
        AnalysisRunRequest(
            mode="rate",
            df=pd.DataFrame(rows),
            csv="",
            total_col="total",
            fraud_col="fraud",
            privacy_concentration_col="clearing_spend",
            dimensions=["segment"],
            output=str(output),
            report_format="json",
            export_balanced_csv=True,
            compliance_posture="best_effort",
            validate_input=False,
            control3_overrides={"privacy_basis": "clearing_spend"},
            privacy_rule_strategy=strategy,
        ),
        logging.getLogger("test"),
    )

    fraud_results = artifacts.results["fraud"]["segment"]
    assert "SAFE_A" in fraud_results["Category"].values
    assert "SECRET_B" not in fraud_results["Category"].values
    assert artifacts.csv_output is not None
    csv_text = Path(artifacts.csv_output).read_text(encoding="utf-8")
    assert "SECRET_B" not in csv_text
    assert "9999" not in csv_text
    assert artifacts.json_output is not None
    json_text = Path(artifacts.json_output).read_text(encoding="utf-8")
    assert "SECRET_B" not in json_text
    assert "9999" not in json_text
    assert artifacts.metadata["export_validation"]["checked"] is True
    assert artifacts.metadata["export_validation"]["passed"] is True
    workbook = load_workbook(output, read_only=True)
    try:
        workbook_text = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
    finally:
        workbook.close()
    assert "SECRET_B" not in workbook_text
    assert "9999" not in workbook_text


def test_suppressed_rate_group_disjoint_peers_are_absent_from_every_sink(
    tmp_path: Path,
) -> None:
    rows: list[dict[str, object]] = []
    for category, peer_prefix, clearing_values in (
        ("SAFE_A", "P", [20, 20, 20, 20, 20]),
        ("SECRET_B", "Q", [25, 25, 25, 25]),
    ):
        for index, clearing in enumerate(clearing_values, start=1):
            rows.append(
                {
                    "issuer_name": f"{peer_prefix}{index}",
                    "segment": category,
                    "total": 20,
                    "fraud": 9999 if category == "SECRET_B" else 1,
                    "clearing_spend": clearing,
                    "year_month": (
                        "SUPPRESSED_PERIOD_SENTINEL"
                        if category == "SECRET_B"
                        else "2024-01"
                    ),
                }
            )
    output = tmp_path / "disjoint_peers.xlsx"
    log_path = tmp_path / "authorized_run.log"
    logger = setup_deferred_logging(
        "DEBUG",
        str(log_path),
    )
    artifacts = execute_rate_run(
        AnalysisRunRequest(
            mode="rate",
            df=pd.DataFrame(rows),
            csv="",
            total_col="total",
            fraud_col="fraud",
            privacy_concentration_col="clearing_spend",
            dimensions=["segment"],
            time_col="year_month",
            output=str(output),
            output_format="both",
            report_format="json",
            export_balanced_csv=True,
            audit_package=True,
            debug=True,
            analyze_distortion=True,
            compliance_posture="best_effort",
            validate_input=False,
            control3_overrides={"privacy_basis": "clearing_spend"},
            privacy_rule_strategy=(
                PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE
            ),
        ),
        logger,
    )
    finalize_deferred_logging(
        logger,
        privacy_authorized=artifacts.privacy_log_authorized,
    )

    forbidden = {
        "SECRET_B",
        "9999",
        "Q1",
        "Q2",
        "Q3",
        "Q4",
        "SUPPRESSED_PERIOD_SENTINEL",
    }

    def assert_workbook_safe(source: Path | BytesIO) -> None:
        workbook = load_workbook(source, read_only=True)
        try:
            text = " ".join(
                str(cell.value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
        finally:
            workbook.close()
        assert not any(value in text for value in forbidden)

    for workbook_path in tmp_path.glob("*.xlsx"):
        assert_workbook_safe(workbook_path)
    for text_path in [
        *tmp_path.glob("*.json"),
        *tmp_path.glob("*.csv"),
        *tmp_path.glob("*.log"),
        *tmp_path.glob("*.txt"),
    ]:
        text = text_path.read_text(encoding="utf-8")
        assert not any(value in text for value in forbidden)
    for package_path in tmp_path.glob("*.zip"):
        with ZipFile(package_path) as package:
            for member in package.namelist():
                payload = package.read(member)
                if member.lower().endswith(".xlsx"):
                    assert_workbook_safe(BytesIO(payload))
                elif member.lower().endswith(
                    (".json", ".csv", ".log", ".txt")
                ):
                    text = payload.decode("utf-8")
                    assert not any(value in text for value in forbidden)


@pytest.mark.parametrize("invalid_value", [-1.0, float("nan")])
def test_invalid_privacy_basis_blocks_even_without_input_validation(
    tmp_path: Path,
    invalid_value: float,
) -> None:
    df = _single_category_df([20, 20, 20, 20, 20]).rename(
        columns={"amount": "total"}
    )
    df["fraud"] = 1.0
    df["clearing_spend"] = [20, 20, 20, 20, invalid_value]

    with pytest.raises(RunBlocked):
        execute_rate_run(
            AnalysisRunRequest(
                mode="rate",
                df=df,
                csv="",
                total_col="total",
                fraud_col="fraud",
                privacy_concentration_col="clearing_spend",
                dimensions=["segment"],
                output=str(tmp_path / "invalid_basis.xlsx"),
                compliance_posture="best_effort",
                validate_input=False,
                control3_overrides={"privacy_basis": "clearing_spend"},
            ),
            logging.getLogger("test"),
        )


@pytest.mark.parametrize(
    "invalid_strategy",
    ["select_by_peer_count", "garbage"],
)
def test_python_api_rejects_non_enum_privacy_strategy(
    tmp_path: Path,
    invalid_strategy: str,
) -> None:
    with pytest.raises(RunAborted, match="PrivacyRuleStrategy"):
        execute_share_run(
            AnalysisRunRequest(
                df=_single_category_df([20, 20, 20, 20, 20]),
                csv="",
                metric="amount",
                dimensions=["segment"],
                output=str(tmp_path / "invalid_strategy.xlsx"),
                privacy_rule_strategy=invalid_strategy,  # type: ignore[arg-type]
            ),
            logging.getLogger("test"),
        )


@pytest.mark.parametrize("invalid_scope", ["false", 0, 1, None, object()])
def test_python_api_rejects_non_bool_merchant_scope(
    tmp_path: Path,
    invalid_scope: object,
) -> None:
    with pytest.raises(RunAborted, match="explicit bool"):
        execute_share_run(
            AnalysisRunRequest(
                df=_single_category_df([25, 25, 25, 25]),
                csv="",
                metric="amount",
                dimensions=["segment"],
                output=str(tmp_path / "invalid_scope.xlsx"),
                is_anonymized_aggregated_merchant_spend=invalid_scope,  # type: ignore[arg-type]
            ),
            logging.getLogger("test"),
        )


@pytest.mark.parametrize("mode", ["share", "rate"])
def test_cli_sweep_executes_normal_analysis(
    tmp_path: Path,
    mode: str,
) -> None:
    output = tmp_path / f"cli_{mode}.xlsx"
    common = [
        mode,
        "--csv",
        str(FIXTURE),
        "--entity",
        "Target",
        "--dimensions",
        *SHARE_DIMENSIONS,
        "--preset",
        "balanced_default",
        "--compliance-posture",
        "strict",
        "--output",
        str(output),
        "--privacy-rule-sweep",
        "--time-col",
        "year_month",
    ]
    args = benchmark.create_parser().parse_args(
        common
        + (
            ["--metric", "txn_cnt"]
            if mode == "share"
            else ["--total-col", "total", "--approved-col", "approved"]
        )
    )

    exit_code = (
        benchmark.run_share_analysis(args, logging.getLogger("test"))
        if mode == "share"
        else benchmark.run_rate_analysis(args, logging.getLogger("test"))
    )

    assert exit_code == benchmark.EXIT_OK
    assert output.exists()
