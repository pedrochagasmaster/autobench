"""Leak, denial, and complete-output parity tests for Verified Safe Coverage outputs."""

from __future__ import annotations

import io
import json
import logging
import zipfile
from pathlib import Path
from typing import Any, Iterable, List, Optional, Set
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.analysis_run import execute_share_run
from core.contracts import (
    AnalysisRunRequest,
    PrivacyReleaseMode,
    PrivacyRuleStrategy,
)
from core.privacy_coverage_verifier import (
    VERIFIER_RESULT_FAILED,
    VERIFIER_RESULT_PASSED,
    VerificationOutcome,
)
from core.privacy_output_policy import (
    CONTROL3_SAFE_COVERAGE_EMPTY,
    CONTROL3_SAFE_COVERAGE_VERIFIER_FAILED,
)
from tests.fixtures.safe_coverage_fixture import (
    build_safe_coverage_getnet_shaped_df,
    write_safe_coverage_bounds_config,
)


SUPPRESSED_MARKER = "SUPP_CAT_MARKER_e7c1b2a9"


def _msc_request(
    tmp_path: Path,
    df: pd.DataFrame,
    *,
    privacy_release_mode: PrivacyReleaseMode = PrivacyReleaseMode.VERIFIED_SAFE_COVERAGE,
    output_name: str = "out.xlsx",
    export_balanced_csv: bool = True,
    audit_package: bool = True,
    report_format: Optional[str] = "json",
    output_format: str = "both",
    output: Optional[str] = None,
) -> AnalysisRunRequest:
    config_path = write_safe_coverage_bounds_config(tmp_path / "bounds.yaml")
    return AnalysisRunRequest(
        mode="share",
        df=df,
        metric="transaction_amount",
        secondary_metrics=["transaction_count", "merchant_count"],
        dimensions=["region", "sector"],
        time_col="quarter",
        entity_col="issuer_name",
        output=str(tmp_path / output_name) if output is None else output,
        output_format=output_format,
        report_format=report_format,
        export_balanced_csv=export_balanced_csv,
        audit_package=audit_package,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        privacy_release_mode=privacy_release_mode,
        validate_input=False,
        compliance_posture="best_effort",
        config=str(config_path),
        preset=None,
    )


def _seed_suppressed_marker(df: pd.DataFrame) -> pd.DataFrame:
    marked = df.copy()
    marked.loc[marked["sector"] == "SectorX", "sector"] = SUPPRESSED_MARKER
    return marked


def _collect_artifact_paths(artifacts: Any) -> List[Path]:
    paths: List[Path] = []
    for attr in (
        "analysis_output_file",
        "publication_output",
        "csv_output",
        "json_output",
        "audit_package_output",
        "audit_log_output",
        "coverage_certificate_output",
    ):
        value = getattr(artifacts, attr, None)
        if value:
            path = Path(value)
            if path.is_file():
                paths.append(path)
    return paths


def _xlsx_text_blobs(path: Path) -> List[bytes]:
    blobs: List[bytes] = [path.read_bytes()]
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            blobs.append(archive.read(name))
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            blobs.append(str(cell).encode("utf-8", errors="ignore"))
        finally:
            workbook.close()
    return blobs


def _zip_member_blobs(path: Path) -> List[bytes]:
    blobs: List[bytes] = [path.read_bytes()]
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            blobs.append(name.encode("utf-8", errors="ignore"))
            blobs.append(archive.read(name))
    return blobs


def _artifact_blobs(path: Path) -> List[bytes]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _xlsx_text_blobs(path)
    if suffix == ".zip":
        return _zip_member_blobs(path)
    return [path.read_bytes()]


def _assert_marker_absent(
    artifacts: Any,
    marker: str,
    *,
    extra_texts: Iterable[str] = (),
) -> None:
    needle = marker.encode("utf-8")
    for path in _collect_artifact_paths(artifacts):
        for blob in _artifact_blobs(path):
            assert needle not in blob, f"marker leaked in {path}"
    if artifacts.coverage_certificate is not None:
        cert_text = json.dumps(
            {
                "visible": list(
                    artifacts.coverage_certificate.visible_publication_unit_keys
                ),
                "rules": dict(artifacts.coverage_certificate.authorizing_rules),
                "digest": artifacts.coverage_certificate.certificate_digest,
            },
            sort_keys=True,
        )
        assert marker not in cert_text
    metadata = artifacts.metadata or {}
    metadata_text = json.dumps(metadata, sort_keys=True, default=str)
    assert marker not in metadata_text
    for text in extra_texts:
        assert marker not in text


def _result_categories(artifacts: Any) -> Set[str]:
    categories: Set[str] = set()
    results = artifacts.results or {}
    for frame in results.values():
        if isinstance(frame, pd.DataFrame) and "Category" in frame.columns:
            categories.update(str(value) for value in frame["Category"].tolist())
    return categories


def test_verified_safe_coverage_fixture_publishes_suppressed_view(
    tmp_path: Path,
) -> None:
    df = _seed_suppressed_marker(build_safe_coverage_getnet_shaped_df())
    log_stream = io.StringIO()
    logger = logging.getLogger("test_msc_outputs")
    logger.handlers.clear()
    logger.addHandler(logging.StreamHandler(log_stream))
    logger.setLevel(logging.INFO)

    artifacts = execute_share_run(_msc_request(tmp_path, df), logger)

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.safe_coverage_result is not None
    result = artifacts.safe_coverage_result
    assert result.verifier_result == VERIFIER_RESULT_PASSED
    assert result.search_state == "search_complete"
    assert result.candidate_vectors_evaluated > 0
    assert len(result.candidate_universe) == 9
    assert len(result.release_set) == 4
    assert len(result.suppression_set) == 5
    assert artifacts.coverage_certificate is not None
    assert artifacts.coverage_certificate_output is not None
    assert Path(artifacts.coverage_certificate_output).is_file()
    assert artifacts.analysis_output_file is not None
    assert artifacts.csv_output is not None
    assert artifacts.json_output is not None
    assert artifacts.audit_package_output is not None

    categories = _result_categories(artifacts)
    assert categories == {"North", "SectorY"}
    assert SUPPRESSED_MARKER not in categories
    assert "SectorZ" not in categories
    assert "South" not in categories

    metadata = artifacts.metadata or {}
    assert metadata.get("artifact_name") == "Suppressed Publication View"
    assert metadata.get("privacy_suppressed_missing_units") is True
    assert any(
        "privacy-suppressed" in str(warning).lower()
        for warning in metadata.get("run_warnings", [])
    )
    assert SUPPRESSED_MARKER not in json.dumps(metadata, default=str)

    _assert_marker_absent(
        artifacts,
        SUPPRESSED_MARKER,
        extra_texts=[log_stream.getvalue()],
    )


def test_complete_output_regression_parity_when_all_units_pass(
    tmp_path: Path,
) -> None:
    df = build_safe_coverage_getnet_shaped_df()
    df = df[df["sector"] == "SectorY"].reset_index(drop=True)

    # Near-neutral bounds keep the tiny all-safe universe stable.
    def _request(
        folder: Path,
        *,
        mode: PrivacyReleaseMode,
        output_name: str,
    ) -> AnalysisRunRequest:
        config_path = write_safe_coverage_bounds_config(
            folder / "bounds.yaml",
            min_weight=0.999,
            max_weight=1.001,
        )
        return AnalysisRunRequest(
            mode="share",
            df=df,
            metric="transaction_amount",
            secondary_metrics=["transaction_count", "merchant_count"],
            dimensions=["region", "sector"],
            time_col="quarter",
            entity_col="issuer_name",
            output=str(folder / output_name),
            output_format="analysis",
            report_format=None,
            export_balanced_csv=False,
            audit_package=False,
            privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
            privacy_release_mode=mode,
            validate_input=False,
            compliance_posture="best_effort",
            config=str(config_path),
            preset=None,
        )

    complete = execute_share_run(
        _request(
            tmp_path / "complete",
            mode=PrivacyReleaseMode.COMPLETE_OUTPUT,
            output_name="complete.xlsx",
        ),
        logging.getLogger("test_complete_parity"),
    )
    maximize = execute_share_run(
        _request(
            tmp_path / "maximize",
            mode=PrivacyReleaseMode.VERIFIED_SAFE_COVERAGE,
            output_name="maximize.xlsx",
        ),
        logging.getLogger("test_msc_parity"),
    )

    assert complete.privacy_sink_authorized is True
    assert maximize.privacy_sink_authorized is True
    assert _result_categories(complete) == _result_categories(maximize)
    assert maximize.safe_coverage_result is not None
    assert set(maximize.safe_coverage_result.release_set) == {
        unit.internal_key
        for unit in maximize.safe_coverage_result.candidate_universe
    }
    # Ignore new-mode evidence when comparing analytical category coverage.
    assert complete.coverage_certificate is None
    assert maximize.coverage_certificate is not None


def test_verifier_failure_writes_denial_only(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    df = _seed_suppressed_marker(build_safe_coverage_getnet_shaped_df())

    def _fail_verify(*_args: Any, **_kwargs: Any) -> VerificationOutcome:
        return VerificationOutcome(
            passed=False,
            failures=(),
            computed_release_mask_digest="0" * 64,
            computed_candidate_universe_digest="0" * 64,
            computed_rule_set_digest="0" * 64,
            computed_artifact_hashes={},
        )

    monkeypatch.setattr(
        "core.analysis_run.verify_safe_coverage_result",
        _fail_verify,
    )
    artifacts = execute_share_run(
        _msc_request(tmp_path, df),
        logging.getLogger("test_verifier_fail"),
    )

    assert artifacts.privacy_sink_authorized is False
    assert artifacts.analysis_output_file is None
    assert artifacts.publication_output is None
    assert artifacts.csv_output is None
    assert artifacts.json_output is None
    assert artifacts.audit_package_output is None
    assert artifacts.coverage_certificate is None
    assert artifacts.coverage_certificate_output is None
    assert artifacts.safe_coverage_result is not None
    assert (
        artifacts.safe_coverage_result.verifier_result == VERIFIER_RESULT_FAILED
    )
    assert (
        artifacts.privacy_output_decision is not None
        and artifacts.privacy_output_decision.withholding_reason
        == CONTROL3_SAFE_COVERAGE_VERIFIER_FAILED
    )
    assert artifacts.audit_log_output is not None
    denial = Path(artifacts.audit_log_output)
    assert denial.is_file()
    assert SUPPRESSED_MARKER.encode("utf-8") not in denial.read_bytes()
    assert not list(tmp_path.glob("*.xlsx"))
    assert not list(tmp_path.glob("*_balanced.csv"))
    assert not list(tmp_path.glob("*_coverage_certificate.json"))


def test_empty_release_set_writes_denial_only(tmp_path: Path) -> None:
    df = build_safe_coverage_getnet_shaped_df()
    df = df[df["sector"].isin(["SectorX", "SectorZ"])].reset_index(drop=True)
    df = _seed_suppressed_marker(df)

    config_path = write_safe_coverage_bounds_config(
        tmp_path / "empty.yaml",
        min_weight=1.0,
        max_weight=1.01,
    )
    request = AnalysisRunRequest(
        mode="share",
        df=df,
        metric="transaction_amount",
        secondary_metrics=["transaction_count", "merchant_count"],
        dimensions=["region", "sector"],
        time_col="quarter",
        entity_col="issuer_name",
        output=str(tmp_path / "empty.xlsx"),
        output_format="both",
        report_format="json",
        export_balanced_csv=True,
        audit_package=True,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        privacy_release_mode=PrivacyReleaseMode.VERIFIED_SAFE_COVERAGE,
        validate_input=False,
        compliance_posture="best_effort",
        config=str(config_path),
        preset=None,
    )
    artifacts = execute_share_run(
        request,
        logging.getLogger("test_empty_release"),
    )

    assert artifacts.privacy_sink_authorized is False
    assert artifacts.safe_coverage_result is not None
    assert artifacts.safe_coverage_result.release_set == ()
    assert (
        artifacts.privacy_output_decision is not None
        and artifacts.privacy_output_decision.withholding_reason
        == CONTROL3_SAFE_COVERAGE_EMPTY
    )
    assert artifacts.analysis_output_file is None
    assert artifacts.csv_output is None
    assert artifacts.json_output is None
    assert artifacts.audit_package_output is None
    assert artifacts.coverage_certificate is None
    assert artifacts.audit_log_output is not None
    assert (
        SUPPRESSED_MARKER.encode("utf-8")
        not in Path(artifacts.audit_log_output).read_bytes()
    )


def test_python_interface_without_output_skips_certificate_disk_write(
    tmp_path: Path,
) -> None:
    df = build_safe_coverage_getnet_shaped_df()
    config_path = write_safe_coverage_bounds_config(tmp_path / "bounds.yaml")
    request = AnalysisRunRequest(
        mode="share",
        df=df,
        metric="transaction_amount",
        secondary_metrics=["transaction_count", "merchant_count"],
        dimensions=["region", "sector"],
        time_col="quarter",
        entity_col="issuer_name",
        output=None,
        output_format="analysis",
        report_format=None,
        export_balanced_csv=False,
        audit_package=False,
        privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        privacy_release_mode=PrivacyReleaseMode.VERIFIED_SAFE_COVERAGE,
        validate_input=False,
        compliance_posture="best_effort",
        config=str(config_path),
        preset=None,
    )
    # Avoid writing default timestamped workbooks into the repo cwd.
    with patch(
        "core.analysis_run._share_output_filename",
        return_value=str(tmp_path / "implicit.xlsx"),
    ):
        artifacts = execute_share_run(
            request,
            logging.getLogger("test_no_disk_cert"),
        )

    assert artifacts.privacy_sink_authorized is True
    assert artifacts.safe_coverage_result is not None
    assert artifacts.coverage_certificate is not None
    assert artifacts.coverage_certificate_output is None
    assert not list(tmp_path.glob("*_coverage_certificate.json"))


def test_complete_output_unchanged_for_partial_fixture(tmp_path: Path) -> None:
    """complete-output must still hard-block when any governed unit fails."""
    df = build_safe_coverage_getnet_shaped_df()
    artifacts = execute_share_run(
        _msc_request(
            tmp_path,
            df,
            privacy_release_mode=PrivacyReleaseMode.COMPLETE_OUTPUT,
            output_name="complete_block.xlsx",
        ),
        logging.getLogger("test_complete_block"),
    )
    assert artifacts.privacy_sink_authorized is False
    assert artifacts.safe_coverage_result is None
    assert artifacts.coverage_certificate is None
    assert artifacts.analysis_output_file is None
    assert artifacts.csv_output is None
