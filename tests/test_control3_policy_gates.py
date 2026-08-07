from __future__ import annotations

import argparse
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
import pytest
from openpyxl import load_workbook

from benchmark import create_parser, run_share_analysis
from core.analysis_run import build_run_request, enforce_compliance_preconditions
from core.contracts import AnalysisRunRequest
from core.control3_policy import Control3PolicyInput, evaluate_control3_policy
from utils.config_manager import ConfigManager
from utils.validators import ConfigValidator


REMOVED_BUSINESS_FLAGS = {
    "--contains-digital-wallet-metrics",
    "--digital-wallet-review-approved",
    "--contains-top-merchant-output",
    "--dual-entity-axis",
    "--dual-entity-axis-review-approved",
    "--recurring-deliverable",
    "--last-privacy-recheck-date",
    "--peer-group-altered",
    "--privacy-basis",
}


def _share_parser():
    parser = create_parser()
    subparsers = next(
        action for action in parser._actions if isinstance(action, argparse._SubParsersAction)
    )
    return subparsers.choices["share"]


def test_cli_defaults_to_compliance_strict() -> None:
    args = create_parser().parse_args(["share", "--csv", "input.csv", "--metric", "txn_cnt"])

    assert args.preset == "compliance_strict"


def test_python_request_defaults_to_compliance_strict() -> None:
    assert AnalysisRunRequest().preset == "compliance_strict"
    assert AnalysisRunRequest(preset=None).preset == "compliance_strict"


def test_cli_does_not_expose_upstream_business_decisions() -> None:
    exposed = {
        option
        for action in _share_parser()._actions
        for option in action.option_strings
    }

    assert REMOVED_BUSINESS_FLAGS.isdisjoint(exposed)


def test_fraud_metrics_derive_clearing_spend_basis() -> None:
    policy = evaluate_control3_policy(
        Control3PolicyInput(
            analysis_mode="rate",
            rate_types=["fraud"],
        )
    )

    assert policy.allowed
    assert policy.requirements["fraud_concentration_basis"] == "derived_from_total_col"
    assert policy.details["fraud_concentration_basis"] == "clearing_spend"


def test_approval_metrics_do_not_derive_clearing_spend_basis() -> None:
    policy = evaluate_control3_policy(
        Control3PolicyInput(
            analysis_mode="rate",
            rate_types=["approval"],
        )
    )

    assert policy.allowed
    assert "fraud_concentration_basis" not in policy.requirements
    assert "fraud_concentration_basis" not in policy.details


@pytest.mark.parametrize(
    "removed_key",
    [
        "contains_digital_wallet_metrics",
        "digital_wallet_review_approved",
        "contains_top_merchant_output",
        "dual_entity_axis",
        "dual_entity_axis_review_approved",
        "recurring_deliverable",
        "last_privacy_recheck_date",
        "peer_group_altered",
    ],
)
def test_removed_control3_section_is_not_valid_configuration(removed_key: str) -> None:
    errors = ConfigValidator.validate(
        {
            "version": "3.0",
            "compliance_posture": "strict",
            "control3": {removed_key: True},
        }
    )

    assert "Unknown configuration fields: control3" in errors


def test_control3_config_rejects_overloaded_privacy_review_approval() -> None:
    errors = ConfigValidator.validate(
        {
            "version": "3.0",
            "compliance_posture": "strict",
            "control3": {"privacy_review_approved": True},
        }
    )

    assert "Unknown configuration fields: control3" in errors


def test_configuration_template_has_no_privacy_basis_setting() -> None:
    template = Path("config/template.yaml").read_text(encoding="utf-8")

    assert "privacy_basis" not in template
    assert "control3:" not in template


def test_compliance_preconditions_allow_fraud_without_public_basis_input() -> None:
    request = build_run_request(
        "rate",
        SimpleNamespace(
            acknowledge_accuracy_first=False,
            approved_col=None,
            fraud_col="fraud",
        ),
    )
    config = ConfigManager()

    result = enforce_compliance_preconditions(config, request)

    assert result["control3_policy"]["allowed"] is True
    assert (
        result["control3_policy"]["details"]["fraud_concentration_basis"]
        == "clearing_spend"
    )


def test_publication_peer_evidence_redacts_peer_composition(tmp_path: Path) -> None:
    output = tmp_path / "share.xlsx"
    df = pd.DataFrame(
        {
            "issuer_name": ["Target", "P1", "P2", "P3", "P4", "P5"],
            "card_type": ["A", "A", "A", "A", "A", "A"],
            "txn_cnt": [100, 100, 100, 100, 100, 100],
        }
    )
    args = SimpleNamespace(
        csv="",
        df=df,
        metric="txn_cnt",
        secondary_metrics=None,
        entity="Target",
        entity_col="issuer_name",
        output=str(output),
        dimensions=["card_type"],
        auto=False,
        time_col=None,
        config=None,
        preset=None,
        debug=True,
        log_level="INFO",
        per_dimension_weights=False,
        export_balanced_csv=False,
        validate_input=False,
        compare_presets=False,
        analyze_distortion=False,
        output_format="both",
        include_calculated=False,
        auto_subset_search=None,
        subset_search_max_tests=None,
        trigger_subset_on_slack=None,
        max_cap_slack=None,
        compliance_posture="best_effort",
        acknowledge_accuracy_first=False,
    )

    assert run_share_analysis(args, __import__("logging").getLogger("test_control3_publication")) == 0

    workbook = load_workbook(tmp_path / "share_publication.xlsx", read_only=True)
    try:
        assert "Peer Weights" in workbook.sheetnames
        values = [
            str(cell.value)
            for row in workbook["Peer Weights"].iter_rows()
            for cell in row
            if cell.value is not None
        ]
        assert "P1" not in values
        assert "P2" not in values
        assert any("Control 3.3" in value for value in values)
    finally:
        workbook.close()
