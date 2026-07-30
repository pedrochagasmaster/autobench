"""Tests for under-populated and structurally infeasible category suppression (F1)."""

from __future__ import annotations

import logging
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.analysis_run import execute_share_run
from core.category_suppression import (
    apply_suppression_to_results,
    compute_suppressed_categories,
    filter_suppressed_rows,
    is_category_suppressed,
)
from core.contracts import AnalysisRunRequest, PrivacyRuleStrategy

FIXTURE = Path(__file__).parent / "fixtures" / "gate_demo.csv"
SHARE_DIMENSIONS = ["card_type", "channel"]


def _build_peer_df(
    *,
    include_prepaid_exclusive: bool = False,
    include_duo_category: bool = False,
    time_aware: bool = False,
) -> pd.DataFrame:
    entities = ["Target", "P1", "P2", "P3", "P4", "P5", "P6"]
    rows: list[dict[str, object]] = []
    months = ["2024-01", "2024-02"] if time_aware else ["2024-01"]

    for month in months:
        for entity in entities:
            rows.append(
                {
                    "issuer_name": entity,
                    "year_month": month,
                    "card_type": "CREDIT",
                    "channel": "Online",
                    "txn_cnt": 100,
                    "total": 1000,
                    "approved": 900,
                    "fraud": 10,
                }
            )

    if include_prepaid_exclusive:
        rows.append(
            {
                "issuer_name": "P1",
                "year_month": "2024-01",
                "card_type": "PREPAID",
                "channel": "Online",
                "txn_cnt": 250,
                "total": 2500,
                "approved": 2200,
                "fraud": 25,
            }
        )

    if include_duo_category:
        for entity in ("P1", "P2"):
            rows.append(
                {
                    "issuer_name": entity,
                    "year_month": "2024-01",
                    "card_type": "DUO",
                    "channel": "Online",
                    "txn_cnt": 150,
                    "total": 1500,
                    "approved": 1350,
                    "fraud": 15,
                }
            )

    if time_aware:
        rows.append(
            {
                "issuer_name": "P1",
                "year_month": "2024-02",
                "card_type": "SPARSE",
                "channel": "Online",
                "txn_cnt": 80,
                "total": 800,
                "approved": 720,
                "fraud": 8,
            }
        )

    return pd.DataFrame(rows)


class TestComputeSuppressedCategories:
    def test_exclusive_category_suppressed(self) -> None:
        df = _build_peer_df(include_prepaid_exclusive=True)
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
        )
        prepaid = [record for record in suppressed if record["category"] == "PREPAID"]
        assert len(prepaid) == 1
        assert prepaid[0]["reason"] == "below_min_entities"
        assert prepaid[0]["participants"] == 1

    def test_duo_category_suppressed(self) -> None:
        df = _build_peer_df(include_duo_category=True)
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
        )
        duo = [record for record in suppressed if record["category"] == "DUO"]
        assert len(duo) == 1
        assert duo[0]["participants"] == 2

    def test_well_populated_category_untouched(self) -> None:
        df = _build_peer_df()
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
        )
        assert not any(record["category"] == "CREDIT" for record in suppressed)

    def test_target_excluded_from_participant_count(self) -> None:
        df = _build_peer_df(include_prepaid_exclusive=True)
        rows = df.to_dict("records")
        rows.append(
            {
                "issuer_name": "Target",
                "year_month": "2024-01",
                "card_type": "PREPAID",
                "channel": "Online",
                "txn_cnt": 500,
                "total": 5000,
                "approved": 4500,
                "fraud": 50,
            }
        )
        df = pd.DataFrame(rows)
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
        )
        prepaid = [record for record in suppressed if record["category"] == "PREPAID"]
        assert len(prepaid) == 1
        assert prepaid[0]["participants"] == 1

    def test_time_aware_grouping(self) -> None:
        df = _build_peer_df(time_aware=True)
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
            time_col="year_month",
        )
        sparse = [record for record in suppressed if record["category"] == "SPARSE"]
        assert len(sparse) == 2
        assert any(record["time_period"] == "2024-02" for record in sparse)
        assert any(record["time_period"] is None for record in sparse)

    def test_structural_infeasible_deduped_when_below_min(self) -> None:
        df = _build_peer_df(include_prepaid_exclusive=True)
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
            structural_infeasible=[("card_type", "PREPAID")],
        )
        prepaid = [record for record in suppressed if record["category"] == "PREPAID"]
        assert len(prepaid) == 1
        assert prepaid[0]["reason"] == "below_min_entities"

    def test_structural_infeasible_without_under_population(self) -> None:
        df = _build_peer_df()
        suppressed = compute_suppressed_categories(
            df,
            entity_col="issuer_name",
            target_entity="Target",
            dimensions=["card_type"],
            metric_col="txn_cnt",
            min_entities=6,
            structural_infeasible=[("card_type", "CREDIT")],
        )
        credit = [record for record in suppressed if record["category"] == "CREDIT"]
        assert len(credit) == 1
        assert credit[0]["reason"] == "structurally_infeasible"


class TestFilterSuppressedRows:
    def test_filter_share_results(self) -> None:
        results_df = pd.DataFrame(
            {
                "Category": ["CREDIT", "PREPAID"],
                "Balanced_txn_cnt": [100.0, 250.0],
            }
        )
        suppressed = [
            {
                "dimension": "card_type",
                "category": "PREPAID",
                "time_period": None,
                "participants": 1,
                "reason": "below_min_entities",
            }
        ]
        filtered = filter_suppressed_rows(results_df, suppressed, "card_type")
        assert list(filtered["Category"]) == ["CREDIT"]

    def test_apply_suppression_rate_shape(self) -> None:
        results = {
            "approval": {
                "card_type": pd.DataFrame({"Category": ["CREDIT", "PREPAID"], "Rate": [90.0, 100.0]}),
            },
            "fraud": {
                "card_type": pd.DataFrame({"Category": ["CREDIT", "PREPAID"], "Rate": [1.0, 2.0]}),
            },
        }
        suppressed = [
            {
                "dimension": "card_type",
                "category": "PREPAID",
                "time_period": None,
                "participants": 1,
                "reason": "below_min_entities",
            }
        ]
        filtered = apply_suppression_to_results(results, suppressed, is_rate=True)
        assert "approval" in filtered
        assert list(filtered["approval"]["card_type"]["Category"]) == ["CREDIT"]
        assert list(filtered["fraud"]["card_type"]["Category"]) == ["CREDIT"]


class TestIsCategorySuppressed:
    def test_category_level_suppression_applies_to_all_times(self) -> None:
        suppressed = [
            {
                "dimension": "card_type",
                "category": "PREPAID",
                "time_period": None,
                "participants": 1,
                "reason": "structurally_infeasible",
            }
        ]
        assert is_category_suppressed(suppressed, "card_type", "PREPAID", "2024-01")
        assert is_category_suppressed(suppressed, "card_type", "PREPAID", "2024-02")
        assert not is_category_suppressed(suppressed, "card_type", "CREDIT", "2024-01")


def test_share_run_suppresses_exclusive_category(tmp_path: Path) -> None:
    csv_path = tmp_path / "exclusive.csv"
    _build_peer_df(include_prepaid_exclusive=True).to_csv(csv_path, index=False)
    out = tmp_path / "share_exclusive.xlsx"

    request = AnalysisRunRequest(
        csv=str(csv_path),
        entity="Target",
        metric="txn_cnt",
        dimensions=["card_type"],
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        output=str(out),
        export_balanced_csv=True,
    )
    artifacts = execute_share_run(request, logging.getLogger("test"))

    card_type_df = artifacts.results["card_type"]
    assert "PREPAID" not in card_type_df["Category"].values

    csv_output = pd.read_csv(artifacts.csv_output)
    assert "PREPAID" not in csv_output["Category"].values

    suppressed = artifacts.metadata.get("suppressed_categories", [])
    assert suppressed
    assert all("category" not in record for record in suppressed)
    assert not any(
        "PREPAID" in warning
        for warning in artifacts.metadata.get("run_warnings", [])
    )

    privacy_df = artifacts.privacy_validation_df
    assert privacy_df is not None
    assert "PREPAID" not in privacy_df["Category"].values


def test_suppressed_group_is_absent_from_every_analysis_workbook_sheet(
    tmp_path: Path,
) -> None:
    safe_values = [20, 20, 10, 8, 7, 7, 7, 7, 7, 7]
    df = pd.DataFrame(
        [
            {
                "issuer_name": f"P{index}",
                "year_month": "2024-01",
                "card_type": "SAFE_CATEGORY",
                "channel": "Online",
                "txn_cnt": value,
            }
            for index, value in enumerate(safe_values, start=1)
        ]
        + [
            {
                "issuer_name": "ONLY_SECRET_PEER",
                "year_month": "2024-01",
                "card_type": "SECRET_CATEGORY",
                "channel": "Online",
                "txn_cnt": 1,
            }
        ]
    )
    output = tmp_path / "suppressed_diagnostics.xlsx"
    artifacts = execute_share_run(
        AnalysisRunRequest(
            df=df,
            csv="",
            metric="txn_cnt",
            dimensions=["card_type"],
            time_col="year_month",
            output=str(output),
            compliance_posture="best_effort",
            validate_input=False,
            debug=True,
            privacy_rule_strategy=PrivacyRuleStrategy.SWEEP_ANY_APPLICABLE,
        ),
        logging.getLogger("test"),
    )

    assert artifacts.privacy_sink_authorized
    workbook = load_workbook(output, read_only=True)
    try:
        text = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
    finally:
        workbook.close()
    assert "SECRET_CATEGORY" not in text
    assert "ONLY_SECRET_PEER" not in text


def test_gate_demo_produces_no_suppressions(tmp_path: Path) -> None:
    out = tmp_path / "gate_share.xlsx"
    request = AnalysisRunRequest(
        csv=str(FIXTURE),
        entity="Target",
        metric="txn_cnt",
        dimensions=SHARE_DIMENSIONS,
        time_col="year_month",
        preset="balanced_default",
        compliance_posture="strict",
        output=str(out),
    )
    artifacts = execute_share_run(request, logging.getLogger("test"))
    assert artifacts.metadata.get("suppressed_categories", []) == []


def test_secondary_metric_below_rule_minimum_omits_only_unsafe_metric(
    tmp_path: Path,
) -> None:
    df = pd.DataFrame(
        {
            "issuer_name": ["P1", "P2", "P3", "P4", "P5"],
            "segment": ["all"] * 5,
            "amount": [20, 20, 20, 20, 20],
            "secondary": [7777, 7777, 7777, 7777, 0],
        }
    )
    artifacts = execute_share_run(
        AnalysisRunRequest(
            df=df,
            csv="",
            metric="amount",
            secondary_metrics=["secondary"],
            dimensions=["segment"],
            output=str(tmp_path / "secondary_suppressed.xlsx"),
            report_format="json",
            export_balanced_csv=True,
            compliance_posture="best_effort",
            validate_input=False,
        ),
        logging.getLogger("test"),
    )

    assert artifacts.privacy_output_decision is not None
    assert artifacts.privacy_output_decision.privacy_publication_authorized
    assert not artifacts.results["segment"].empty
    assert artifacts.secondary_results_df is not None
    assert "secondary" not in artifacts.secondary_results_df.columns
    suppressed = artifacts.metadata.get(
        "suppressed_metric_categories",
        [],
    )
    assert any(
        record["metric"] == "secondary"
        for record in suppressed
    )

    assert artifacts.csv_output is not None
    csv_frame = pd.read_csv(artifacts.csv_output)
    assert "Balanced_amount" in csv_frame.columns
    assert "Balanced_secondary" not in csv_frame.columns

    workbook = load_workbook(
        tmp_path / "secondary_suppressed.xlsx",
        read_only=True,
    )
    try:
        workbook_text = " ".join(
            str(cell.value)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.value is not None
        )
    finally:
        workbook.close()
    assert "7777" not in workbook_text

    assert artifacts.json_output is not None
    json_payload = json.loads(
        Path(artifacts.json_output).read_text(encoding="utf-8")
    )
    assert "7777" not in json.dumps(json_payload)
