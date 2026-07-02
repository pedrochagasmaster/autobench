"""Tests for representativeness warnings, Summary disclosure, and validation fixes."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from core.data_loader import DataLoader
from core.representativeness import MULTIPLIER_RATIO_WARN, compute_representativeness
from core.report_generator import ReportGenerator


def _summary_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb["Summary"].iter_rows(values_only=True))
    finally:
        wb.close()


def _summary_as_dict(path: Path) -> dict[str, str]:
    rows = _summary_rows(path)
    metadata: dict[str, str] = {}
    for row in rows:
        if not row or row[0] is None:
            continue
        key = str(row[0]).rstrip(":")
        if len(row) > 1 and row[1] is not None:
            metadata[key] = str(row[1])
    return metadata


class TestComputeRepresentativeness:
    def test_ratio_warning_fires_above_threshold(self) -> None:
        metadata = {
            "method_breakdown_df": pd.DataFrame(
                {"Peer": ["P1", "P2"], "Multiplier": [1.0, 50.0]}
            ),
            "min_weight": 0.01,
        }
        result = compute_representativeness(metadata)
        assert result["multiplier_ratio"] == pytest.approx(50.0)
        assert any("50x" in warning for warning in result["warnings"])

    def test_ratio_warning_does_not_fire_at_2x(self) -> None:
        metadata = {
            "method_breakdown_df": pd.DataFrame(
                {"Peer": ["P1", "P2"], "Multiplier": [1.0, 2.0]}
            ),
            "min_weight": 0.01,
        }
        result = compute_representativeness(metadata)
        assert result["multiplier_ratio"] == pytest.approx(2.0)
        assert not any("max/min" in warning for warning in result["warnings"])

    def test_missing_metadata_keys_are_safe(self) -> None:
        result = compute_representativeness({})
        assert result["warnings"] == []
        assert result["multiplier_ratio"] == 0.0
        assert result["mean_abs_impact_pp"] is None

    def test_mean_impact_warning(self) -> None:
        metadata = {
            "method_breakdown_df": pd.DataFrame({"Peer": ["P1"], "Multiplier": [1.0]}),
            "impact_summary": {"mean_abs_impact_pp": 54.0},
            "min_weight": 0.01,
        }
        result = compute_representativeness(metadata)
        assert any("54.0pp" in warning for warning in result["warnings"])

    def test_peers_at_min_floor_warning(self) -> None:
        metadata = {
            "method_breakdown_df": pd.DataFrame(
                {"Peer": ["P1", "P2"], "Multiplier": [0.01, 1.0]}
            ),
            "min_weight": 0.01,
        }
        result = compute_representativeness(metadata)
        assert result["peers_at_min_floor"] == 1
        assert any("minimum weight floor" in warning for warning in result["warnings"])


class TestValidateRateInputPeerOutliers:
    @pytest.fixture
    def data_loader(self) -> DataLoader:
        return DataLoader(MagicMock())

    def test_peer_outlier_rate_warning_for_15pct_vs_06pct_median(self, data_loader: DataLoader) -> None:
        df = pd.DataFrame(
            {
                "issuer_name": ["Outlier", "P2", "P3", "P4", "P5", "P6"],
                "total": [1000, 1000, 1000, 1000, 1000, 1000],
                "fraud": [150, 6, 6, 6, 6, 6],
                "dimension": ["X"] * 6,
            }
        )
        issues = data_loader.validate_rate_input(
            df=df,
            total_col="total",
            numerator_cols={"fraud": "fraud"},
            entity_col="issuer_name",
            dimensions=["dimension"],
        )
        warning = next(issue for issue in issues if issue.category == "peer_outlier_rate")
        assert "Outlier" in warning.message
        assert "15.00%" in warning.message
        assert "0.60%" in warning.message

    def test_uniform_rates_do_not_emit_peer_outlier_warning(self, data_loader: DataLoader) -> None:
        df = pd.DataFrame(
            {
                "issuer_name": [f"P{i}" for i in range(1, 7)],
                "total": [1000] * 6,
                "fraud": [6] * 6,
                "dimension": ["X"] * 6,
            }
        )
        issues = data_loader.validate_rate_input(
            df=df,
            total_col="total",
            numerator_cols={"fraud": "fraud"},
            entity_col="issuer_name",
            dimensions=["dimension"],
        )
        assert not any(issue.category == "peer_outlier_rate" for issue in issues)

    def test_outlier_band_message_uses_real_band(self, data_loader: DataLoader) -> None:
        df = pd.DataFrame(
            {
                "issuer_name": ["A", "B", "C", "D", "E", "F"],
                "total": [100, 100, 100, 100, 100, 100],
                "approved": [200, 95, 90, 80, 70, 60],
                "dimension": ["X"] * 6,
            }
        )
        issues = data_loader.validate_rate_input(
            df=df,
            total_col="total",
            numerator_cols={"custom_rate": "approved"},
            entity_col="issuer_name",
            dimensions=["dimension"],
            thresholds={"max_rate_deviation": 50.0},
        )
        warning = next(issue for issue in issues if issue.category == "outlier_rates")
        assert "outside 0-150% plausible band" in warning.message


class TestSummarySheetDisclosure:
    def _write_summary_workbook(self, tmp_path: Path, metadata: dict, results: dict, analysis_type: str) -> Path:
        output = tmp_path / "summary.xlsx"
        generator = ReportGenerator(MagicMock())
        generator._font_class = Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        generator._write_summary_sheet(
            ws,
            results=results,
            analysis_type=analysis_type,
            metadata=metadata,
            report_model=None,
        )
        wb.save(output)
        return output

    def test_summary_renders_warnings_and_suppressed_categories(self, tmp_path: Path) -> None:
        metadata = {
            "entity": "Target",
            "compliance_posture": "best_effort",
            "compliance_verdict": "fully_compliant",
            "acknowledgement_state": "not_required",
            "run_status": "compliant",
            "validation_warnings": 2,
            "compliance_summary": {
                "data_quality_publishable": True,
                "data_quality_checked": True,
                "validation_warnings": 2,
                "strict_final_validation": {"primary_cap_fail_rows": 0, "secondary_rule_fail_categories": 0, "relaxed_rows": 0, "total_violations": 0},
            },
            "run_warnings": [
                "Peer weight ratio max/min is 50x — balanced averages may not represent the raw peer market",
                "Suppressed card_type/PREPAID: 1 participant(s) < rule minimum 6",
            ],
            "suppressed_categories": [
                {
                    "dimension": "card_type",
                    "category": "PREPAID",
                    "time_period": None,
                    "participants": 1,
                    "reason": "below_min_entities",
                }
            ],
        }
        output = self._write_summary_workbook(
            tmp_path,
            metadata,
            results={"card_type": pd.DataFrame({"Category": ["A"]})},
            analysis_type="share",
        )
        summary = _summary_as_dict(output)
        assert summary["Input Validation"] == "pass (2 warnings)"
        assert summary["Suppressed Categories"] == "1"
        assert summary["Warnings"] == "2"
        rows = _summary_rows(output)
        rendered = "\n".join(str(cell) for row in rows for cell in row if cell is not None)
        assert "card_type/PREPAID" in rendered
        assert "below_min_entities" in rendered
        assert "50x" in rendered

    def test_summary_renders_without_optional_metadata_keys(self, tmp_path: Path) -> None:
        metadata = {
            "entity": "Target",
            "compliance_posture": "best_effort",
            "compliance_verdict": "fully_compliant",
            "acknowledgement_state": "not_required",
            "run_status": "compliant",
            "compliance_summary": {
                "data_quality_publishable": True,
                "data_quality_checked": True,
                "strict_final_validation": {"primary_cap_fail_rows": 0, "secondary_rule_fail_categories": 0, "relaxed_rows": 0, "total_violations": 0},
            },
        }
        output = self._write_summary_workbook(
            tmp_path,
            metadata,
            results={"card_type": pd.DataFrame({"Category": ["A"]})},
            analysis_type="share",
        )
        summary = _summary_as_dict(output)
        assert summary["Input Validation"] == "pass"
        assert "Suppressed Categories" not in summary
        assert "Warnings" not in summary

    def test_summary_shows_zero_bic_caveat_for_rate_results(self, tmp_path: Path) -> None:
        metadata = {
            "entity": "Target",
            "compliance_posture": "best_effort",
            "compliance_verdict": "fully_compliant",
            "acknowledgement_state": "not_required",
            "run_status": "compliant",
            "compliance_summary": {
                "data_quality_publishable": True,
                "data_quality_checked": True,
                "strict_final_validation": {"primary_cap_fail_rows": 0, "secondary_rule_fail_categories": 0, "relaxed_rows": 0, "total_violations": 0},
            },
        }
        results = {
            "fraud": {
                "card_type": pd.DataFrame(
                    {
                        "Category": ["A", "B"],
                        "BIC (%)": [0.0, 1.5],
                    }
                )
            }
        }
        output = self._write_summary_workbook(tmp_path, metadata, results, analysis_type="rate")
        rows = _summary_rows(output)
        rendered = "\n".join(str(cell) for row in rows for cell in row if cell is not None)
        assert "Best-in-class of 0.0%" in rendered

    def test_multiplier_ratio_threshold_constant(self) -> None:
        assert MULTIPLIER_RATIO_WARN == 20.0
