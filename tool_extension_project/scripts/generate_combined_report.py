"""
Generate combined market share report with distortion analysis
Creates Excel file with multiple sheets showing:
- Balanced market shares
- Raw vs Balanced comparison
- Distortion analysis
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from pathlib import Path


def load_data(csv_path: Path) -> pd.DataFrame:
    return pd.read_csv(csv_path)


def map_tier(tier):
    """Map tier to combined categories."""
    if tier in ['Gold', 'Standard']:
        return 'Gold + Standard'
    else:
        return tier


def calculate_all_shares(raw_df: pd.DataFrame, balanced_df: pd.DataFrame, 
                         entity: str = "NU PAGAMENTOS SA", time_col: str = "ano") -> pd.DataFrame:
    """Calculate raw shares, balanced shares, and distortion."""
    
    # Get raw shares
    nubank_data = raw_df[raw_df['issuer_name'] == entity].copy()
    raw_totals = raw_df.groupby(['ano', 'function_variant']).agg({
        'volume_brl': 'sum',
        'txn_count': 'sum'
    }).reset_index()
    
    results = []
    
    for _, bal_row in balanced_df.iterrows():
        category = bal_row['Category']
        year = bal_row[time_col]
        
        # Get Nubank's raw volumes
        nubank = nubank_data[
            (nubank_data['function_variant'] == category) &
            (nubank_data[time_col] == year)
        ]
        
        nubank_vol = nubank['volume_brl'].sum()
        nubank_cnt = nubank['txn_count'].sum()
        
        # Get raw totals
        raw_total_row = raw_totals[
            (raw_totals['function_variant'] == category) &
            (raw_totals[time_col] == year)
        ]
        
        raw_total_vol = raw_total_row['volume_brl'].sum()
        raw_total_cnt = raw_total_row['txn_count'].sum()
        
        # Balanced totals
        peer_vol_balanced = bal_row['volume_brl']
        peer_cnt_balanced = bal_row['txn_count']
        balanced_total_vol = nubank_vol + peer_vol_balanced
        balanced_total_cnt = nubank_cnt + peer_cnt_balanced
        
        # Calculate shares
        raw_vol_share = (nubank_vol / raw_total_vol * 100) if raw_total_vol > 0 else 0
        raw_cnt_share = (nubank_cnt / raw_total_cnt * 100) if raw_total_cnt > 0 else 0
        
        bal_vol_share = (nubank_vol / balanced_total_vol * 100) if balanced_total_vol > 0 else 0
        bal_cnt_share = (nubank_cnt / balanced_total_cnt * 100) if balanced_total_cnt > 0 else 0
        
        # Parse function and tier
        parts = category.split(' / ')
        function = parts[0] if len(parts) > 0 else ''
        tier = parts[1] if len(parts) > 1 else ''
        
        results.append({
            'Year': int(year),
            'Function': function,
            'Tier': tier,
            'Tier_Mapped': map_tier(tier),
            'Volume_Share_Raw': raw_vol_share,
            'Volume_Share_Balanced': bal_vol_share,
            'Volume_Distortion_pp': bal_vol_share - raw_vol_share,
            'Count_Share_Raw': raw_cnt_share,
            'Count_Share_Balanced': bal_cnt_share,
            'Count_Distortion_pp': bal_cnt_share - raw_cnt_share
        })
    
    return pd.DataFrame(results)


def create_pivot(df: pd.DataFrame, metric: str) -> pd.DataFrame:
    """Create pivot table for a specific metric."""
    # Aggregate by mapped tier
    agg_df = df.groupby(['Year', 'Function', 'Tier_Mapped']).agg({
        metric: 'sum'
    }).reset_index()
    
    pivot = agg_df.pivot_table(
        index=['Function', 'Tier_Mapped'],
        columns='Year',
        values=metric,
        aggfunc='first'
    )
    
    pivot = pivot.round(0).astype(int)
    pivot = pivot.sort_index(axis=1)
    
    # Sort index
    function_order = ['Credit', 'Debit']
    tier_order = ['Gold + Standard', 'Platinum', 'Black']
    
    pivot = pivot.reindex(
        [(f, t) for f in function_order for t in tier_order if (f, t) in pivot.index],
        fill_value=0
    )
    
    return pivot


def format_sheet(ws, pivot_df: pd.DataFrame, title: str, subtitle: str, show_percent: bool = True):
    """Format a worksheet with a pivot table."""
    
    # Title
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(pivot_df.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Subtitle
    ws.cell(row=2, column=1, value=subtitle)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(pivot_df.columns))
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Headers
    header_row = 4
    ws.cell(row=header_row, column=1, value="Function")
    ws.cell(row=header_row, column=2, value="Tier")
    
    for col_idx, year in enumerate(pivot_df.columns, start=3):
        ws.cell(row=header_row, column=col_idx, value=int(year))
    
    # Style headers
    num_cols = 2 + len(pivot_df.columns)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.border = Border(bottom=Side(style='medium'))
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    data_start_row = header_row + 1
    prev_function = None
    
    for row_idx, ((function, tier), row_data) in enumerate(pivot_df.iterrows(), start=data_start_row):
        # Function column (only show if different from previous)
        if function != prev_function:
            ws.cell(row=row_idx, column=1, value=function)
            prev_function = function
        
        # Tier column
        ws.cell(row=row_idx, column=2, value=tier)
        
        # Year data
        for col_idx, value in enumerate(row_data, start=3):
            cell = ws.cell(row=row_idx, column=col_idx)
            if value == 0 or pd.isna(value):
                cell.value = "-"
                cell.alignment = Alignment(horizontal='center')
            else:
                if show_percent:
                    cell.value = f"{int(value)}%"
                else:
                    # For distortion, show with sign
                    sign = "+" if value > 0 else ""
                    cell.value = f"{sign}{int(value)} pp"
                cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    year_cols = ['C', 'D', 'E', 'F', 'G', 'H']
    for col in year_cols[:len(pivot_df.columns)]:
        ws.column_dimensions[col].width = 10


def main():
    script_dir = Path(__file__).parent
    raw_csv = script_dir / "data" / "e176097_tpv_nubank_extended_std.csv"
    balanced_csv = script_dir / "benchmark_share_NU_PAGAMENTOS_SA_20260109_115832_balanced.csv"
    output_file = script_dir / "nubank_market_share_complete_report.xlsx"
    
    print("Loading data...")
    raw_df = load_data(raw_csv)
    balanced_df = pd.read_csv(balanced_csv)
    
    print("Calculating all shares and distortion...")
    all_shares = calculate_all_shares(raw_df, balanced_df)
    
    print("Creating pivot tables...")
    vol_balanced_pivot = create_pivot(all_shares, 'Volume_Share_Balanced')
    cnt_balanced_pivot = create_pivot(all_shares, 'Count_Share_Balanced')
    vol_distortion_pivot = create_pivot(all_shares, 'Volume_Distortion_pp')
    cnt_distortion_pivot = create_pivot(all_shares, 'Count_Distortion_pp')
    
    print("Generating Excel workbook...")
    wb = openpyxl.Workbook()
    
    # Sheet 1: Volume Market Share (Balanced)
    ws1 = wb.active
    ws1.title = "Volume Share"
    format_sheet(ws1, vol_balanced_pivot, 
                "Nubank – Market Share Estimates (volume)",
                "Share by Product within Mastercard Portfolio (2020-2025) - Balanced Data",
                show_percent=True)
    
    # Sheet 2: Count Market Share (Balanced)
    ws2 = wb.create_sheet("Count Share")
    format_sheet(ws2, cnt_balanced_pivot,
                "Nubank – Market Share Estimates (# of Transactions)",
                "Share by Product within Mastercard Portfolio (2020-2025) - Balanced Data",
                show_percent=True)
    
    # Sheet 3: Volume Distortion
    ws3 = wb.create_sheet("Volume Distortion")
    format_sheet(ws3, vol_distortion_pivot,
                "Market Share Distortion - Volume",
                "Difference between Balanced and Raw Shares (Balanced - Raw)",
                show_percent=False)
    
    # Sheet 4: Count Distortion
    ws4 = wb.create_sheet("Count Distortion")
    format_sheet(ws4, cnt_distortion_pivot,
                "Market Share Distortion - Transaction Count",
                "Difference between Balanced and Raw Shares (Balanced - Raw)",
                show_percent=False)
    
    # Save
    wb.save(output_file)
    print(f"\nComplete report saved to: {output_file}")
    print("\nSheets created:")
    print("  1. Volume Share - Balanced market shares by volume")
    print("  2. Count Share - Balanced market shares by transaction count")
    print("  3. Volume Distortion - Impact of balancing on volume shares")
    print("  4. Count Distortion - Impact of balancing on count shares")


if __name__ == "__main__":
    main()
