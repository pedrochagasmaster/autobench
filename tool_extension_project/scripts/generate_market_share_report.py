"""
Generate Nubank market share report for Excel export
Creates two tables: Volume market share and Transaction count market share
Aggregated by year (2023-2025) and tier (Gold+Standard, Platinum, Black)
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from pathlib import Path


def load_data(csv_path: Path) -> pd.DataFrame:
    """Load the extended data CSV."""
    return pd.read_csv(csv_path)


def calculate_market_shares_from_balanced(
    raw_df: pd.DataFrame, 
    balanced_df: pd.DataFrame,
    entity: str = "NU PAGAMENTOS SA",
    time_col: str = "ano"
) -> pd.DataFrame:
    """Calculate Nubank's market share using balanced peer volumes."""
    
    results = []
    
    # Get Nubank volumes from raw data (entity is not balanced)
    nubank_data = raw_df[raw_df['issuer_name'] == entity].copy()
    
    for _, bal_row in balanced_df.iterrows():
        category = bal_row['Category']
        year = bal_row[time_col]
        
        # Get Nubank's raw volume for this category and year
        nubank = nubank_data[
            (nubank_data['function_variant'] == category) &
            (nubank_data[time_col] == year)
        ]
        
        nubank_vol = nubank['volume_brl'].sum()
        nubank_cnt = nubank['txn_count'].sum()
        
        # Balanced peer volumes
        peer_vol_balanced = bal_row['volume_brl']
        peer_cnt_balanced = bal_row['txn_count']
        
        # Total = Nubank + Balanced Peers
        total_vol = nubank_vol + peer_vol_balanced
        total_cnt = nubank_cnt + peer_cnt_balanced
        
        vol_share = (nubank_vol / total_vol * 100) if total_vol > 0 else 0
        cnt_share = (nubank_cnt / total_cnt * 100) if total_cnt > 0 else 0
        
        # Parse function and tier
        parts = category.split(' / ')
        function = parts[0] if len(parts) > 0 else ''
        tier = parts[1] if len(parts) > 1 else ''
        
        results.append({
            'Year': int(year),
            'Function': function,
            'Tier': tier,
            'Volume_Share_Pct': vol_share,
            'Count_Share_Pct': cnt_share
        })
    
    return pd.DataFrame(results)


def create_pivot_table(df: pd.DataFrame, metric: str) -> pd.DataFrame:
    """Create pivot table for a specific metric."""
    
    # Map tiers to combined categories
    def map_tier(tier):
        if tier in ['Gold', 'Standard']:
            return 'Gold + Standard'
        else:
            return tier
    
    df = df.copy()
    df['Tier_Mapped'] = df['Tier'].apply(map_tier)
    
    # Pivot
    pivot = df.pivot_table(
        index=['Function', 'Tier_Mapped'],
        columns='Year',
        values=metric,
        aggfunc='sum'
    )
    
    # Round to integers (percentages)
    pivot = pivot.round(0).astype(int)
    
    # Sort columns by year
    pivot = pivot.sort_index(axis=1)
    
    # Sort index
    function_order = ['Credit', 'Debit']
    tier_order = ['Gold + Standard', 'Platinum', 'Black']
    
    pivot = pivot.reindex(
        [(f, t) for f in function_order for t in tier_order if (f, t) in pivot.index],
        fill_value=0
    )
    
    return pivot


def format_table_in_excel(ws, start_row: int, pivot_df: pd.DataFrame, title: str):
    """Format a pivot table in Excel worksheet."""
    
    # Title
    ws.cell(row=start_row, column=1, value=title)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    title_cell = ws.cell(row=start_row, column=1)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Subtitle
    subtitle_row = start_row + 1
    ws.cell(row=subtitle_row, column=1, value="Share by Product within Mastercard Portfolio (2020-2025)")
    ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=8)
    subtitle_cell = ws.cell(row=subtitle_row, column=1)
    subtitle_cell.font = Font(size=10)
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Header row
    header_row = start_row + 3
    ws.cell(row=header_row, column=1, value="Function")
    ws.cell(row=header_row, column=2, value="Tier")
    
    for col_idx, year in enumerate(pivot_df.columns, start=3):
        ws.cell(row=header_row, column=col_idx, value=int(year))
    
    # Style header
    num_cols = 2 + len(pivot_df.columns)  # Function + Tier + year columns
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='medium'))
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    data_start_row = header_row + 1
    prev_function = None
    
    for row_idx, ((function, tier), row_data) in enumerate(pivot_df.iterrows(), start=data_start_row):
        # Function column (only show if different from previous)
        if function != prev_function:
            ws.cell(row=row_idx, column=1, value=function)
            prev_function = function
        
        # Tier column
        ws.cell(row=row_idx, column=2, value=tier)
        
        # Year data
        for col_idx, value in enumerate(row_data, start=3):
            cell = ws.cell(row=row_idx, column=col_idx)
            if value == 0 or pd.isna(value):
                cell.value = "-"
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = f"{int(value)}%"
                cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    # Set width for all year columns
    year_cols = ['C', 'D', 'E', 'F', 'G', 'H']
    for col in year_cols[:len(pivot_df.columns)]:
        ws.column_dimensions[col].width = 10


def main():
    script_dir = Path(__file__).parent
    raw_csv_path = script_dir / "data" / "e176097_tpv_nubank_extended_std.csv"
    balanced_csv_path = script_dir / "benchmark_share_NU_PAGAMENTOS_SA_20260109_115832_balanced.csv"
    output_path = script_dir / "nubank_market_share_report.xlsx"
    
    print("Loading data...")
    raw_df = load_data(raw_csv_path)
    balanced_df = pd.read_csv(balanced_csv_path)
    
    print("Calculating market shares from balanced data...")
    shares_df = calculate_market_shares_from_balanced(raw_df, balanced_df)
    
    print("Creating pivot tables...")
    volume_pivot = create_pivot_table(shares_df, 'Volume_Share_Pct')
    count_pivot = create_pivot_table(shares_df, 'Count_Share_Pct')
    
    print("Generating Excel file...")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Share Report"
    
    # Add volume table
    format_table_in_excel(ws, start_row=1, pivot_df=volume_pivot, 
                         title="Nubank – Market Share Estimates (volume)")
    
    # Add count table (leave space)
    format_table_in_excel(ws, start_row=15, pivot_df=count_pivot,
                         title="Nubank – Market Share Estimates (# of Transactions)")
    
    # Save
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")
    
    # Print preview
    print("\n" + "="*60)
    print("VOLUME MARKET SHARE (BALANCED)")
    print("="*60)
    print(volume_pivot)
    
    print("\n" + "="*60)
    print("TRANSACTION COUNT MARKET SHARE (BALANCED)")
    print("="*60)
    print(count_pivot)


if __name__ == "__main__":
    main()
