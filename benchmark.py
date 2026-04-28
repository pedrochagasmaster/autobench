#!/usr/bin/env python
"""
Privacy-Compliant Benchmarking Tool

A dimensional analysis tool for comparing entity performance against peer groups
while maintaining Mastercard privacy compliance (Control 3.2).

Supports:
- Share Analysis: Transaction count/amount distribution across dimensions
- Rate Analysis: Approval rates and fraud rates across dimensions

Version: 2.0
"""

import argparse
import sys
import json
import logging
import gc
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, Tuple

import pandas as pd

# Import core modules
from core.dimensional_analyzer import DimensionalAnalyzer
from core.data_loader import ValidationSeverity
from core.report_generator import ReportGenerator
from core.privacy_validator import PrivacyValidator
from core.analysis_run import (
    build_report_paths,
    build_run_config,
    prepare_run_data,
    resolve_dimensions,
    resolve_output_settings,
    resolve_target_entity,
    validate_analysis_input,
    write_audit_log,
)
from utils.config_manager import ConfigManager
from utils.logger import setup_logging

# Best preset marker for comparison tables
BEST_PRESET_MARKER = '*'


def _resolve_consistency_mode(
    opt_config: Dict[str, Any],
    logger: logging.Logger,
) -> Tuple[bool, str]:
    """Translate config consistency mode into a bool for the analyzer."""
    consistency_mode = opt_config.get('constraints', {}).get('consistency_mode', 'global')
    mode_normalized = str(consistency_mode).strip().lower().replace('_', '-')
    if mode_normalized in ('per-dimension', 'perdimension', 'per dimension'):
        return False, consistency_mode
    if mode_normalized in ('global', 'consistent', 'global-weights'):
        return True, consistency_mode

    logger.warning(f"Unknown consistency_mode '{consistency_mode}', defaulting to global weights")
    return True, consistency_mode


def _build_dimensional_analyzer(
    *,
    target_entity: Optional[str],
    entity_col: str,
    analysis_config: Dict[str, Any],
    opt_config: Dict[str, Any],
    time_col: Optional[str],
    debug_mode: bool,
    bic_percentile: float,
    logger: logging.Logger,
    consistent_weights: Optional[bool] = None,
) -> Tuple[DimensionalAnalyzer, Dict[str, Any]]:
    """Build a DimensionalAnalyzer from merged config."""
    dyn_constraints = opt_config.get('constraints', {}).get('dynamic_constraints', {})
    if consistent_weights is None:
        consistent_weights, consistency_mode = _resolve_consistency_mode(opt_config, logger)
    else:
        consistency_mode = opt_config.get('constraints', {}).get('consistency_mode', 'global')

    rank_penalty_weight = opt_config['linear_programming'].get('rank_penalty_weight', 1.0)
    volume_preservation = opt_config['constraints']['volume_preservation']
    rank_preservation_strength = volume_preservation * float(rank_penalty_weight)
    lambda_penalty = opt_config['linear_programming'].get('lambda_penalty')
    bayesian_max_iterations = opt_config.get('bayesian', {}).get('max_iterations', 500)
    bayesian_learning_rate = opt_config.get('bayesian', {}).get('learning_rate', 0.01)
    violation_penalty_weight = opt_config.get('bayesian', {}).get('violation_penalty_weight', 1000.0)
    enforce_single_weight_set = bool(
        opt_config.get('constraints', {}).get('enforce_single_weight_set', False)
    )

    analyzer = DimensionalAnalyzer(
        target_entity=target_entity,
        entity_column=entity_col,
        bic_percentile=bic_percentile,
        debug_mode=debug_mode,
        consistent_weights=consistent_weights,
        merchant_mode=analysis_config.get('merchant_mode', False),
        rank_constraint_mode=opt_config['linear_programming'].get('rank_constraints', {}).get('mode', 'all'),
        rank_constraint_k=opt_config['linear_programming'].get('rank_constraints', {}).get('neighbor_k', 1),
        max_iterations=opt_config['linear_programming']['max_iterations'],
        tolerance=opt_config['linear_programming']['tolerance'],
        max_weight=opt_config['bounds']['max_weight'],
        min_weight=opt_config['bounds']['min_weight'],
        volume_preservation_strength=rank_preservation_strength,
        prefer_slacks_first=opt_config['subset_search'].get('prefer_slacks_first', False),
        auto_subset_search=opt_config['subset_search'].get('enabled', True),
        subset_search_max_tests=opt_config['subset_search'].get('max_attempts', 200),
        greedy_subset_search=(opt_config['subset_search'].get('strategy', 'greedy') == 'greedy'),
        trigger_subset_on_slack=opt_config['subset_search'].get('trigger_on_slack', True),
        max_cap_slack=opt_config['subset_search'].get('max_slack_threshold', 0.0),
        time_column=time_col,
        volume_weighted_penalties=opt_config['linear_programming'].get('volume_weighted_penalties', False),
        volume_weighting_exponent=opt_config['linear_programming'].get('volume_weighting_exponent', 1.0),
        lambda_penalty=lambda_penalty,
        enforce_additional_constraints=opt_config.get('constraints', {}).get('enforce_additional_constraints', True),
        dynamic_constraints_enabled=dyn_constraints.get('enabled', True),
        min_peer_count_for_constraints=dyn_constraints.get('min_peer_count', 4),
        min_effective_peer_count=dyn_constraints.get('min_effective_peer_count', 3.0),
        min_category_volume_share=dyn_constraints.get('min_category_volume_share', 0.001),
        min_overall_volume_share=dyn_constraints.get('min_overall_volume_share', 0.0005),
        min_representativeness=dyn_constraints.get('min_representativeness', 0.1),
        dynamic_threshold_scale_floor=dyn_constraints.get('threshold_scale_floor', 0.6),
        dynamic_count_scale_floor=dyn_constraints.get('count_scale_floor', 0.5),
        representativeness_penalty_floor=dyn_constraints.get('penalty_floor', 0.25),
        representativeness_penalty_power=dyn_constraints.get('penalty_power', 1.0),
        bayesian_max_iterations=bayesian_max_iterations,
        bayesian_learning_rate=bayesian_learning_rate,
        violation_penalty_weight=violation_penalty_weight,
        enforce_single_weight_set=enforce_single_weight_set,
    )
    return analyzer, {
        'consistent_weights': consistent_weights,
        'consistency_mode': consistency_mode,
        'rank_penalty_weight': rank_penalty_weight,
        'rank_preservation_strength': rank_preservation_strength,
        'lambda_penalty': lambda_penalty,
        'bayesian_max_iterations': bayesian_max_iterations,
        'bayesian_learning_rate': bayesian_learning_rate,
        'violation_penalty_weight': violation_penalty_weight,
        'enforce_single_weight_set': enforce_single_weight_set,
        'dynamic_constraints_config': dyn_constraints,
    }


def get_presets_help() -> str:
    """Generate help text for available presets."""
    try:
        from utils.preset_manager import PresetManager
        preset_mgr = PresetManager()
        presets = preset_mgr.list_presets()
        if not presets:
            return ""
        
        help_text = "\nAVAILABLE PRESETS:\n"
        for name in presets:
            desc = preset_mgr.get_preset_description(name) or "No description"
            help_text += f"  {name:20s}: {desc}\n"
        return help_text
    except Exception:
        return ""


def create_parser() -> argparse.ArgumentParser:
    """Create and configure the argument parser."""
    
    parser = argparse.ArgumentParser(
        prog='benchmark',
        description='Privacy-Compliant Dimensional Benchmarking Tool v3.0',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
EXAMPLES:
  # Share analysis with preset
  python benchmark.py share --csv data.csv --entity "BANCO SANTANDER" --metric txn_cnt --preset standard

  # Share analysis with custom config
  python benchmark.py share --csv data.csv --entity "BANCO SANTANDER" --metric txn_cnt --config my_config.yaml

  # Rate analysis (approval rates)
  python benchmark.py rate --csv data.csv --entity "BANCO SANTANDER" \\
    --total-col txn_cnt --approved-col app_cnt --preset conservative

  # List available presets
  python benchmark.py config list

  # Show preset details
  python benchmark.py config show conservative

  # Generate config template
  python benchmark.py config generate my_config.yaml

  # Validate config file
  python benchmark.py config validate my_config.yaml

{get_presets_help()}
        """
    )
    
    # Add version flag
    parser.add_argument('--version', action='store_true',
                       help='Show version information')
    
    # Create subparsers for different commands
    subparsers = parser.add_subparsers(dest='command', help='Command to run')
    
    # Get available preset choices
    def get_preset_choices():
        try:
            from utils.preset_manager import PresetManager
            preset_mgr = PresetManager()
            return preset_mgr.list_presets()
        except Exception:
            return []
    
    preset_choices = get_preset_choices()
    
    # ========================================================================
    # SHARE ANALYSIS COMMAND
    # ========================================================================
    share_parser = subparsers.add_parser(
        'share',
        help='Share-based dimensional analysis',
        description='Analyze how an entity\'s volume is distributed across dimensions'
    )
    
    # Required arguments
    share_parser.add_argument('--csv', required=True, 
                             help='Path to CSV input file')
    share_parser.add_argument('--metric', required=True,
                             help='Metric column name to analyze (e.g., txn_cnt, tpv, transaction_count, transaction_amount)')
    
    # Secondary metrics (can specify multiple)
    share_parser.add_argument('--secondary-metrics', nargs='+',
                             help='Secondary metric columns to analyze using weights derived from the primary metric (space-separated list)')
    
    # Optional - Essential
    share_parser.add_argument('--entity',
                             help='Name of the entity to benchmark (omit for peer-only analysis)')
    share_parser.add_argument('--entity-col', default='issuer_name',
                             help='Entity identifier column name (default: issuer_name)')
    share_parser.add_argument('--output', '-o',
                             help='Output file path (default: auto-generated)')
    
    # Dimension selection
    dim_group = share_parser.add_mutually_exclusive_group()
    dim_group.add_argument('--dimensions', nargs='+',
                          help='Specific dimensions to analyze (e.g., flag_domestic cp_cnp)')
    # NOTE: default=None with store_true allows distinguishing "not provided" from
    # "explicitly False", enabling clean CLI-to-config override logic
    dim_group.add_argument('--auto', action='store_true', default=None,
                          help='Auto-detect all available dimensions')
    
    # Time awareness
    share_parser.add_argument('--time-col', 
                             help='Time column name for time-aware consistency (e.g., ano_mes, year_month)')
    
    # Configuration
    share_parser.add_argument('--config',
                             help='Configuration file (YAML)')
    share_parser.add_argument('--preset', choices=preset_choices,
                             help='Preset configuration name')
    
    # Debug/Logging
    share_parser.add_argument('--debug', action='store_true', default=None,
                             help='Enable debug mode (includes unweighted averages and weight details)')
    share_parser.add_argument('--log-level',
                             choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                             help='Logging level (default: INFO)')
    share_parser.add_argument('--per-dimension-weights', action='store_true', default=None,
                             help='Optimize each dimension independently (disables global weighting mode)')
    share_parser.add_argument('--export-balanced-csv', action='store_true',
                             help='Export balanced shares and volumes to CSV (without weights or original values)')
    
    # Enhanced Analysis Options
    share_parser.add_argument('--compare-presets', action='store_true', default=None,
                             help='Compare all presets and report impact for each')
    share_parser.add_argument('--analyze-impact', action='store_true', default=None,
                             help='Include impact details and summary sheets in output')
    share_parser.add_argument('--analyze-distortion', action='store_true', default=None,
                             help='Alias for --analyze-impact (deprecated)')
    share_parser.add_argument('--validate-input', action='store_true', default=None,
                             dest='validate_input',
                             help='Enable input data validation before analysis (default: enabled)')
    share_parser.add_argument('--no-validate-input', action='store_false', dest='validate_input',
                             help='Disable input data validation')
    share_parser.add_argument('--output-format', choices=['analysis', 'publication', 'both'],
                             default=None,
                             help='Output format: analysis (default), publication, or both')
    share_parser.add_argument('--publication-format', action='store_const', const='publication',
                             dest='output_format',
                             help='Convenience alias for --output-format=publication')
    share_parser.add_argument('--include-calculated', action='store_true', default=None,
                             help='Include calculated metrics (raw/balanced share, impact) in balanced CSV export')

    # Advanced Optimization
    share_parser.add_argument('--auto-subset-search', action='store_true', default=None,
                             help='Automatically search for largest feasible global dimension subset')
    share_parser.add_argument('--subset-search-max-tests', type=int,
                             help='Maximum attempts during subset search')
    share_parser.add_argument('--trigger-subset-on-slack', action='store_true', default=None,
                             help='Trigger subset search if LP uses slack')
    share_parser.add_argument('--max-cap-slack', type=float,
                             help='Slack sum threshold to trigger subset search')

    # ========================================================================
    # RATE ANALYSIS COMMAND
    # ========================================================================
    rate_parser = subparsers.add_parser(
        'rate',
        help='Rate-based dimensional analysis',
        description='Analyze approval rates or fraud rates across dimensions'
    )
    
    # Required arguments
    rate_parser.add_argument('--csv', required=True,
                            help='Path to CSV input file')
    rate_parser.add_argument('--total-col', required=True,
                            help='Total transactions column (e.g., txn_cnt)')
    
    # Secondary metrics (can specify multiple)
    rate_parser.add_argument('--secondary-metrics', nargs='+',
                            help='Secondary metric columns (e.g., txn_count) to analyze using weights derived from the total column (space-separated list)')
    
    # Rate type selection (both can be specified for simultaneous analysis)
    rate_parser.add_argument('--approved-col',
                            help='Approved transactions column (for approval rate)')
    rate_parser.add_argument('--fraud-col',
                            help='Fraud transactions column (for fraud rate)')
    
    # Optional - Essential
    rate_parser.add_argument('--entity',
                            help='Name of the entity to benchmark (omit for peer-only analysis)')
    rate_parser.add_argument('--entity-col', default='issuer_name',
                            help='Entity identifier column name (default: issuer_name)')
    rate_parser.add_argument('--output', '-o',
                            help='Output file path (default: auto-generated)')
    
    # Dimension selection
    rate_dim_group = rate_parser.add_mutually_exclusive_group()
    rate_dim_group.add_argument('--dimensions', nargs='+',
                               help='Specific dimensions to analyze')
    rate_dim_group.add_argument('--auto', action='store_true', default=None,
                               help='Auto-detect all dimensions')
    
    # Time awareness
    rate_parser.add_argument('--time-col', 
                            help='Time column name for time-aware consistency (e.g., ano_mes, year_month)')
    
    # Configuration
    rate_parser.add_argument('--config',
                            help='Configuration file (YAML)')
    rate_parser.add_argument('--preset', choices=preset_choices,
                            help='Preset configuration name')
    
    # Debug/Logging
    rate_parser.add_argument('--debug', action='store_true', default=None,
                            help='Enable debug mode (includes unweighted averages and weight details)')
    rate_parser.add_argument('--log-level',
                            choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'],
                            help='Logging level (default: INFO)')
    rate_parser.add_argument('--per-dimension-weights', action='store_true', default=None,
                            help='Optimize each dimension independently (disables global weighting mode)')
    rate_parser.add_argument('--export-balanced-csv', action='store_true',
                            help='Export balanced shares and volumes to CSV (without weights or original values)')
    
    # Enhanced Analysis Options
    rate_parser.add_argument('--compare-presets', action='store_true', default=None,
                            help='Compare all presets and report impact for each')
    rate_parser.add_argument('--analyze-impact', action='store_true', default=None,
                            help='Include impact details and summary sheets in output')
    rate_parser.add_argument('--analyze-distortion', action='store_true', default=None,
                            help='Alias for --analyze-impact (deprecated)')
    rate_parser.add_argument('--validate-input', action='store_true', default=None,
                            dest='validate_input',
                            help='Enable input data validation before analysis (default: enabled)')
    rate_parser.add_argument('--no-validate-input', action='store_false', dest='validate_input',
                            help='Disable input data validation')
    rate_parser.add_argument('--output-format', choices=['analysis', 'publication', 'both'],
                            default=None,
                            help='Output format: analysis (default), publication, or both')
    rate_parser.add_argument('--publication-format', action='store_const', const='publication',
                            dest='output_format',
                            help='Convenience alias for --output-format=publication')
    rate_parser.add_argument('--include-calculated', action='store_true', default=None,
                            help='Include calculated metrics (rate impact) in balanced CSV export')
    rate_parser.add_argument('--fraud-in-bps', action='store_true', default=None,
                            dest='fraud_in_bps',
                            help='Convert fraud rates to basis points in publication format (default: enabled)')
    rate_parser.add_argument('--no-fraud-in-bps', action='store_false', dest='fraud_in_bps',
                            help='Keep fraud rates as percentages in publication format')
    
    # Advanced Optimization
    rate_parser.add_argument('--auto-subset-search', action='store_true', default=None,
                            help='Automatically search for largest feasible global dimension subset')
    rate_parser.add_argument('--subset-search-max-tests', type=int,
                            help='Maximum attempts during subset search')
    rate_parser.add_argument('--trigger-subset-on-slack', action='store_true', default=None,
                            help='Trigger subset search if LP uses slack')
    rate_parser.add_argument('--max-cap-slack', type=float,
                            help='Slack sum threshold to trigger subset search')

    # ========================================================================
    # CONFIG MANAGEMENT COMMAND
    # ========================================================================
    config_parser = subparsers.add_parser(
        'config',
        help='Manage configurations and presets'
    )
    
    config_subparsers = config_parser.add_subparsers(dest='config_command',
                                                     help='Config management command')
    
    # List presets
    config_subparsers.add_parser('list',
                                 help='List available presets')
    
    # Show preset
    show_parser = config_subparsers.add_parser('show',
                                               help='Show preset details')
    show_parser.add_argument('preset_name',
                            help='Preset name')
    
    # Validate config
    validate_parser = config_subparsers.add_parser('validate',
                                                   help='Validate config file')
    validate_parser.add_argument('config_file',
                                help='Config file path')
    
    # Generate template
    generate_parser = config_subparsers.add_parser('generate',
                                                   help='Generate config template')
    generate_parser.add_argument('output_file',
                                help='Output file path')
    
    return parser


def handle_config_command(args: argparse.Namespace) -> int:
    """Handle config subcommands.
    
    Args:
        args: Parsed command-line arguments
        
    Returns:
        Exit code (0 for success, 1 for failure)
    """
    import shutil
    from utils.preset_manager import PresetManager
    from utils.validators import validate_config_file
    
    preset_mgr = PresetManager()
    
    if args.config_command == 'list':
        print(preset_mgr.format_preset_list())
        return 0
    
    elif args.config_command == 'show':
        print(preset_mgr.format_preset_detail(args.preset_name))
        return 0
    
    elif args.config_command == 'validate':
        is_valid, errors = validate_config_file(Path(args.config_file))
        if is_valid:
            print(f"[OK] Configuration file is valid: {args.config_file}")
            return 0
        else:
            print(f"[FAIL] Configuration validation failed:")
            for error in errors:
                print(f"  {error}")
            return 1
    
    elif args.config_command == 'generate':
        template_path = Path(__file__).parent / 'config' / 'template.yaml'
        output_path = Path(args.output_file)
        
        if not template_path.exists():
            print(f"[FAIL] Template file not found: {template_path}")
            print(f"  Please ensure the config/template.yaml file exists.")
            return 1
        
        if output_path.exists():
            print(f"[FAIL] File already exists: {output_path}")
            print(f"  Please choose a different filename or delete the existing file.")
            return 1
        
        try:
            shutil.copy(template_path, output_path)
            print(f"[OK] Configuration template created: {output_path}")
            print(f"  Edit this file to customize your analysis settings.")
            print(f"  Validate with: benchmark config validate {output_path}")
            return 0
        except Exception as e:
            print(f"[FAIL] Failed to create config file: {e}")
            return 1
    
    else:
        print("Usage: benchmark config {list|show|validate|generate}")
        print("  list                    List available presets")
        print("  show PRESET             Show preset details")
        print("  validate PATH           Validate a config file")
        print("  generate PATH           Generate a template config file")
        return 1


def print_version() -> None:
    """Print version information."""
    import sys
    import platform
    
    print("Privacy-Compliant Peer Benchmark Tool")
    print(f"Version: 3.0.0")
    print(f"Python: {sys.version.split()[0]}")
    print(f"Platform: {platform.system()} {platform.release()}")


def list_presets() -> None:
    """Display available presets (deprecated - use 'benchmark config list')."""
    print("Note: 'benchmark presets' is deprecated. Use 'benchmark config list' instead.\n")
    from utils.preset_manager import PresetManager
    preset_mgr = PresetManager()
    print(preset_mgr.format_preset_list())


def run_preset_comparison(
    df: pd.DataFrame,
    metric_col: str,
    entity_col: str,
    dimensions: list,
    target_entity: Optional[str],
    time_col: Optional[str],
    analysis_type: str,
    logger: logging.Logger,
    total_col: Optional[str] = None,
    numerator_cols: Optional[Dict[str, str]] = None
) -> pd.DataFrame:
    """
    Run analysis for all standard presets and compare impact.
    
    Parameters:
    -----------
    df : pd.DataFrame
        Input data
    metric_col : str
        Metric column for share analysis
    entity_col : str
        Entity identifier column
    dimensions : list
        Dimensions to analyze
    target_entity : Optional[str]
        Target entity (None for peer-only)
    time_col : Optional[str]
        Time column
    analysis_type : str
        'share' or 'rate'
    logger : logging.Logger
        Logger instance
    total_col : Optional[str]
        For rate analysis: total transactions column
    numerator_cols : Optional[Dict[str, str]]
        For rate analysis: numerator columns
        
    Returns:
    --------
    pd.DataFrame
        Comparison table with preset, mean/min/max/std impact
    """
    if not dimensions:
        logger.warning("No dimensions provided for preset comparison. Skipping.")
        return pd.DataFrame()

    from utils.preset_manager import PresetManager
    
    preset_mgr = PresetManager()
    presets = preset_mgr.list_presets()
    presets_to_test = sorted(presets)
    variants = []
    for preset_name in presets_to_test:
        variants.append((preset_name, True, preset_name))
        variants.append((preset_name, False, f"{preset_name}+perdim"))
    
    logger.info(f"Comparing {len(variants)} presets: {', '.join([v[2] for v in variants])}")
    
    comparison_results = []
    
    for preset_name, consistent_weights, display_name in variants:
        mode_label = "global" if consistent_weights else "per-dimension"
        logger.info(f"  Testing preset: {display_name} ({mode_label})...")
        
        try:
            # Load preset config
            config = ConfigManager(preset=preset_name)
            opt_config = config.config['optimization']
            analysis_config = config.config['analysis']

            analyzer, _ = _build_dimensional_analyzer(
                target_entity=target_entity,
                entity_col=entity_col,
                analysis_config=analysis_config,
                opt_config=opt_config,
                time_col=time_col,
                debug_mode=False,
                bic_percentile=analysis_config.get('best_in_class_percentile', 0.85),
                logger=logger,
                consistent_weights=consistent_weights,
            )
            
            # Calculate global weights
            if consistent_weights:
                if analysis_type == 'share':
                    analyzer.calculate_global_privacy_weights(df, metric_col, dimensions)
                else:
                    analyzer.calculate_global_privacy_weights(df, total_col, dimensions)
            
            # Calculate impact
            if analysis_type == 'share' and target_entity:
                impact_df = analyzer.calculate_share_impact(df, metric_col, dimensions, target_entity)
                if not impact_df.empty:
                    stats = {
                        'Preset': display_name,
                        'Weight_Mode': mode_label,
                        'Mean_Abs_Impact_PP': round(impact_df['Impact_PP'].abs().mean(), 4),
                        'Max_Abs_Impact_PP': round(impact_df['Impact_PP'].abs().max(), 4),
                        'Min_Impact_PP': round(impact_df['Impact_PP'].min(), 4),
                        'Max_Impact_PP': round(impact_df['Impact_PP'].max(), 4),
                        'Std_Impact_PP': round(impact_df['Impact_PP'].std(), 4) if len(impact_df) > 1 else 0.0,
                        'Categories_Analyzed': len(impact_df),
                        'Dimensions_In_Global_LP': len(getattr(analyzer, 'global_dimensions_used', [])),
                        'LP_Method': getattr(analyzer, 'last_lp_stats', {}).get('method', 'N/A'),
                        'Max_Slack_%': round(getattr(analyzer, 'last_lp_stats', {}).get('max_slack', 0.0), 4),
                    }
                    comparison_results.append(stats)
                    logger.info(f"    Mean abs impact: {stats['Mean_Abs_Impact_PP']:.4f} pp")
                else:
                    logger.warning(f"    No impact data for preset {display_name}")
            elif analysis_type == 'rate' and numerator_cols:
                # For rate analysis, calculate impact
                impact_df = analyzer.calculate_rate_impact(df, total_col, numerator_cols, dimensions)
                if not impact_df.empty:
                    # Get first rate's impact column
                    rate_cols = [c for c in impact_df.columns if c.endswith('_Impact_PP')]
                    if rate_cols:
                        mean_abs_values = []
                        max_abs_values = []
                        stats = {
                            'Preset': display_name,
                            'Weight_Mode': mode_label,
                            'Categories_Analyzed': len(impact_df),
                            'Dimensions_In_Global_LP': len(getattr(analyzer, 'global_dimensions_used', [])),
                            'LP_Method': getattr(analyzer, 'last_lp_stats', {}).get('method', 'N/A'),
                            'Max_Slack_%': round(getattr(analyzer, 'last_lp_stats', {}).get('max_slack', 0.0), 4),
                        }
                        for effect_col in rate_cols:
                            rate_name = effect_col.replace('_Impact_PP', '')
                            mean_abs = round(impact_df[effect_col].abs().mean(), 4)
                            max_abs = round(impact_df[effect_col].abs().max(), 4)
                            stats[f'{rate_name}_Mean_Abs_Impact_PP'] = mean_abs
                            stats[f'{rate_name}_Max_Abs_Impact_PP'] = max_abs
                            stats[f'{rate_name}_Min_Impact_PP'] = round(impact_df[effect_col].min(), 4)
                            stats[f'{rate_name}_Max_Impact_PP'] = round(impact_df[effect_col].max(), 4)
                            stats[f'{rate_name}_Std_Impact_PP'] = round(impact_df[effect_col].std(), 4) if len(impact_df) > 1 else 0.0
                            mean_abs_values.append(mean_abs)
                            max_abs_values.append(max_abs)
                        if mean_abs_values:
                            stats['Mean_Abs_Impact_PP'] = round(sum(mean_abs_values) / len(mean_abs_values), 4)
                            stats['Max_Abs_Impact_PP'] = round(max(max_abs_values), 4)
                        comparison_results.append(stats)
                        if 'Mean_Abs_Impact_PP' in stats:
                            logger.info(f"    Mean abs impact: {stats['Mean_Abs_Impact_PP']:.4f} pp")
            else:
                # Peer-only mode - use LP stats instead
                stats = {
                    'Preset': display_name,
                    'Weight_Mode': mode_label,
                    'Dimensions_In_Global_LP': len(getattr(analyzer, 'global_dimensions_used', [])),
                    'LP_Method': getattr(analyzer, 'last_lp_stats', {}).get('method', 'N/A'),
                    'Max_Slack_%': round(getattr(analyzer, 'last_lp_stats', {}).get('max_slack', 0.0), 4),
                    'Sum_Slack_%': round(getattr(analyzer, 'last_lp_stats', {}).get('sum_slack', 0.0), 4),
                }
                comparison_results.append(stats)
                logger.info(f"    LP stats collected for preset {display_name}")
                    
        except Exception as e:
            logger.error(f"    Failed to test preset {display_name}: {e}")
            comparison_results.append({
                'Preset': display_name,
                'Weight_Mode': mode_label,
                'Error': str(e)
            })
    
    comparison_df = pd.DataFrame(comparison_results)
    
    # Mark best preset (lowest mean absolute impact)
    if not comparison_df.empty:
        impact_col = None
        for candidate in ('Mean_Abs_Impact_PP', 'Mean_Abs_Effect_PP', 'Mean_Abs_Distortion_PP'):
            if candidate in comparison_df.columns:
                impact_col = candidate
                break
        if impact_col:
            valid_values = comparison_df[impact_col].dropna()
            if not valid_values.empty:
                min_idx = comparison_df[impact_col].idxmin()
                comparison_df['Best'] = ''
                comparison_df.loc[min_idx, 'Best'] = BEST_PRESET_MARKER
                best_preset = comparison_df.loc[min_idx, 'Preset']
                logger.info(f"\nBest preset (lowest mean abs impact): {best_preset}")
            else:
                logger.warning("No valid impact data to determine best preset")
    
    return comparison_df


def run_share_analysis(args: argparse.Namespace, logger: logging.Logger) -> int:
    """Execute share-based dimensional analysis."""
    logger.info("Starting share-based dimensional analysis")
    
    try:
        # Create CLI overrides dictionary
        config = build_run_config(args)
        output_settings = resolve_output_settings(config)
        include_preset_comparison = output_settings['include_preset_comparison']
        include_impact_summary = output_settings['include_impact_summary']
        include_calculated_metrics = output_settings['include_calculated_metrics']
        include_privacy_validation = output_settings['include_privacy_validation']
        include_audit_log = output_settings['include_audit_log']
        output_format = output_settings['output_format']
        fraud_in_bps = output_settings['fraud_in_bps']
        
        # Load data
        try:
            data_loader, df, entity_col, time_col = prepare_run_data(
                args,
                config,
                logger,
                preferred_entity_col=config.get('input', 'entity_col'),
            )
        except ValueError as exc:
            logger.error(str(exc))
            return 1
        
        # Validate metric column exists (use exact name as specified by user)
        metric_col = args.metric
        if metric_col not in df.columns:
            logger.error(f"Metric column '{metric_col}' not found in data")
            logger.error(f"Available columns: {', '.join(df.columns.tolist())}")
            return 1
        
        # ========================================
        # Input Data Validation
        # ========================================
        validation_issues, should_abort = validate_analysis_input(
            df=df,
            config=config,
            data_loader=data_loader,
            analysis_type='share',
            metric_col=metric_col,
            entity_col=entity_col,
            dimensions=args.dimensions,
            time_col=time_col,
            target_entity=args.entity,
        )
        if should_abort:
            return 1
        
        resolved_entity = resolve_target_entity(df, entity_col, args.entity, logger)
        if args.entity and resolved_entity is None:
            return 1

        logger.info(f"Using entity column: {entity_col}")
        logger.info(f"Analyzing metric: {metric_col}")
        
        # Check for peer-only mode
        if resolved_entity is None:
            logger.info("Running in PEER-ONLY mode (no target entity specified)")
        else:
            logger.info(f"Target entity: {resolved_entity}")
        
        # Get configuration values
        opt_config = config.config['optimization']
        analysis_config = config.config['analysis']

        # Log configuration source
        if getattr(args, 'preset', None):
            logger.info(f"Using preset: {args.preset}")
        if getattr(args, 'config', None):
            logger.info(f"Using config file: {args.config}")

        # Get unique entities and counts for metadata
        unique_entities = df[entity_col].nunique()
        total_records = len(df)

        # Initialize analyzer with config values
        debug_mode = config.get('output', 'include_debug_sheets', default=False)
        analyzer, analyzer_settings = _build_dimensional_analyzer(
            target_entity=resolved_entity,
            entity_col=entity_col,
            analysis_config=analysis_config,
            opt_config=opt_config,
            time_col=time_col,
            debug_mode=debug_mode,
            bic_percentile=analysis_config.get('best_in_class_percentile', 0.85),
            logger=logger,
        )
        consistent_weights = analyzer_settings['consistent_weights']
        consistency_mode = analyzer_settings['consistency_mode']
        rank_penalty_weight = analyzer_settings['rank_penalty_weight']
        rank_preservation_strength = analyzer_settings['rank_preservation_strength']
        lambda_penalty = analyzer_settings['lambda_penalty']
        bayesian_max_iterations = analyzer_settings['bayesian_max_iterations']
        bayesian_learning_rate = analyzer_settings['bayesian_learning_rate']
        violation_penalty_weight = analyzer_settings['violation_penalty_weight']
        enforce_single_weight_set = analyzer_settings['enforce_single_weight_set']
        dyn_constraints = analyzer_settings['dynamic_constraints_config']
        
        # Determine dimensions
        dimensions = resolve_dimensions(args, config, data_loader, df, logger)
        if dimensions is None:
            return 1

        # Calculate global weights if consistent_weights mode is enabled (default)
        if consistent_weights:
            analyzer.calculate_global_privacy_weights(df, metric_col, dimensions)
            structural_summary = analyzer.get_structural_infeasibility_summary()
            if structural_summary.get('has_structural_infeasibility'):
                logger.warning(
                    "Structural infeasibility detected (dimensions=%s, categories=%s, worst_margin=%0.4fpp)",
                    structural_summary.get('infeasible_dimensions'),
                    structural_summary.get('infeasible_categories'),
                    structural_summary.get('worst_margin_pp'),
                )

        # Run analysis
        results = {}
        for dim in dimensions:
            try:
                result_df = analyzer.analyze_dimension_share(
                    df=df,
                    dimension_column=dim,
                    metric_col=metric_col
                )
                results[dim] = result_df
            except Exception as e:
                logger.error(f"Error analyzing dimension {dim}: {e}")
                continue

        if not results:
            logger.error("No analysis results generated")
            return 1

        # Run secondary analysis if requested
        secondary_results_df = None
        secondary_metrics = getattr(args, 'secondary_metrics', None)
        if secondary_metrics:
            logger.info(f"\nCalculating balanced metrics for {len(secondary_metrics)} secondary metric(s): {', '.join(secondary_metrics)}")
            logger.info(f"Using weights calculated from primary metric: {metric_col}")
            
            secondary_results_df = get_balanced_metrics_df(
                df=df,
                analyzer=analyzer,
                dimensions=dimensions,
                metric_col=metric_col,
                secondary_metrics=secondary_metrics
            )
            logger.info(f"Secondary analysis complete. Generated {len(secondary_results_df)} rows of balanced data.")

        # Collect metadata for report
        peer_count = unique_entities if resolved_entity is None else max(unique_entities - 1, 0)
        metadata = {
            'entity': resolved_entity if resolved_entity else 'PEER-ONLY',
            'analysis_type': 'share',
            'metric': metric_col,
            'secondary_metrics': secondary_metrics,
            'entity_column': entity_col,
            'total_records': total_records,
            'unique_entities': unique_entities,
            'peer_count': peer_count,
            'bic_percentile': config.get('analysis', 'best_in_class_percentile'),
            'dimensions_analyzed': len(results),
            'dimension_names': list(results.keys()),
            'preset': getattr(args, 'preset', None),
            'debug_mode': debug_mode,
            'consistent_weights': consistent_weights,
            'merchant_mode': config.get('analysis', 'merchant_mode', default=False),
            'include_debug_sheets': debug_mode,
            'include_privacy_validation': include_privacy_validation,
            'include_impact_summary': include_impact_summary,
            'include_preset_comparison': include_preset_comparison,
            'include_calculated_metrics': include_calculated_metrics,
            'output_format': output_format,
            'timestamp': datetime.now(),
            # New: capture all input parameters
            'input_csv': getattr(args, 'csv', None),
            'log_level': getattr(args, 'log_level', None),
            'dimensions_mode': 'manual' if bool(getattr(args, 'dimensions', None)) else ('auto' if getattr(args, 'auto', False) else 'manual'),
            'dimensions_requested': getattr(args, 'dimensions', None),
            'entity_col_arg': getattr(args, 'entity_col', None),
            'consistency_mode': consistency_mode,
            'enforce_single_weight_set': enforce_single_weight_set,
            'max_iterations': opt_config.get('linear_programming', {}).get('max_iterations'),
            'tolerance_pp': opt_config.get('linear_programming', {}).get('tolerance'),
            'max_weight': opt_config.get('bounds', {}).get('max_weight'),
            'min_weight': opt_config.get('bounds', {}).get('min_weight'),
            'volume_preservation_strength': opt_config.get('constraints', {}).get('volume_preservation'),
            'rank_penalty_weight': opt_config.get('linear_programming', {}).get('rank_penalty_weight'),
            'rank_preservation_strength': getattr(analyzer, 'rank_preservation_strength', None),
            'prefer_slacks_first': opt_config.get('subset_search', {}).get('prefer_slacks_first'),
            'auto_subset_search': opt_config.get('subset_search', {}).get('enabled'),
            'subset_search_max_tests': opt_config.get('subset_search', {}).get('max_attempts'),
            'trigger_subset_on_slack': opt_config.get('subset_search', {}).get('trigger_on_slack'),
            'max_cap_slack': opt_config.get('subset_search', {}).get('max_slack_threshold'),
            'analyzer_ref': analyzer,
            'last_lp_stats': getattr(analyzer, 'last_lp_stats', {}),
            'slack_subset_triggered': getattr(analyzer, 'slack_subset_triggered', False),
            'structural_infeasibility_summary': analyzer.get_structural_infeasibility_summary(),
            # Extended optimization parameters
            'lambda_penalty': opt_config.get('linear_programming', {}).get('lambda_penalty'),
            'volume_weighted_penalties': opt_config.get('linear_programming', {}).get('volume_weighted_penalties'),
            'volume_weighting_exponent': opt_config.get('linear_programming', {}).get('volume_weighting_exponent'),
            'subset_search_enabled': opt_config.get('subset_search', {}).get('enabled'),
            'subset_search_strategy': opt_config.get('subset_search', {}).get('strategy'),
            'subset_search_max_tests': opt_config.get('subset_search', {}).get('max_attempts'),
            'subset_search_trigger_on_slack': opt_config.get('subset_search', {}).get('trigger_on_slack'),
            'subset_search_max_slack_threshold': opt_config.get('subset_search', {}).get('max_slack_threshold'),
            'bayesian_max_iterations': opt_config.get('bayesian', {}).get('max_iterations'),
            'bayesian_learning_rate': opt_config.get('bayesian', {}).get('learning_rate'),
            'violation_penalty_weight': opt_config.get('bayesian', {}).get('violation_penalty_weight'),
            'impact_thresholds': {
                'high_pp': config.get(
                    'output',
                    'impact_thresholds',
                    'high_pp',
                    default=config.get('output', 'distortion_thresholds', 'high_distortion_pp', default=1.0)
                ),
                'low_pp': config.get(
                    'output',
                    'impact_thresholds',
                    'low_pp',
                    default=config.get('output', 'distortion_thresholds', 'low_distortion_pp', default=0.25)
                ),
            },
        }
        metadata['privacy_rule'] = getattr(analyzer, 'privacy_rule_name', None)
        metadata['additional_constraints_enforced'] = getattr(analyzer, 'enforce_additional_constraints', False)
        metadata['additional_constraint_violations_count'] = len(getattr(analyzer, 'additional_constraint_violations', []) or [])
        metadata['dynamic_constraints_enabled'] = getattr(analyzer, 'dynamic_constraints_enabled', False)
        metadata['dynamic_constraints_stats'] = getattr(analyzer, 'dynamic_constraint_stats', {})
        metadata['dynamic_constraints_config'] = opt_config.get('constraints', {}).get('dynamic_constraints', {})
        metadata['privacy_rule'] = getattr(analyzer, 'privacy_rule_name', None)
        metadata['additional_constraints_enforced'] = getattr(analyzer, 'enforce_additional_constraints', False)
        metadata['additional_constraint_violations_count'] = len(getattr(analyzer, 'additional_constraint_violations', []) or [])
        metadata['dynamic_constraints_enabled'] = getattr(analyzer, 'dynamic_constraints_enabled', False)
        metadata['dynamic_constraints_stats'] = getattr(analyzer, 'dynamic_constraint_stats', {})
        metadata['dynamic_constraints_config'] = dyn_constraints
        
        # Calculate rank preservation strength for metadata (new for v2.0)
        if consistent_weights:
            metadata['global_dimensions_used'] = getattr(analyzer, 'global_dimensions_used', [])
            metadata['removed_dimensions'] = getattr(analyzer, 'removed_dimensions', [])
            metadata['per_dimension_weighted'] = list(getattr(analyzer, 'per_dimension_weights', {}).keys())
            # capture subset search attempts if any
            metadata['subset_search_results'] = getattr(analyzer, 'subset_search_results', [])

        # Get weights data if debug mode
        weights_df = None
        privacy_validation_df = None
        export_csv = getattr(args, 'export_balanced_csv', False)
        
        if debug_mode:
            weights_df = analyzer.get_weights_dataframe()
            if not weights_df.empty:
                logger.info(f"Captured weights data: {len(weights_df)} weight entries")
        
        # Build privacy validation dataframe if debug mode OR CSV export is requested
        if (include_privacy_validation or debug_mode or export_csv) and consistent_weights:
            privacy_validation_df = analyzer.build_privacy_validation_dataframe(df, metric_col, dimensions)
            if not privacy_validation_df.empty:
                logger.info(f"Built privacy validation data: {len(privacy_validation_df)} validation entries")
                if 'Structural_Infeasible_Category' in privacy_validation_df.columns:
                    structural_rows = int((privacy_validation_df['Structural_Infeasible_Category'] == 'Yes').sum())
                    structural_categories = int(
                        privacy_validation_df.loc[
                            privacy_validation_df['Structural_Infeasible_Category'] == 'Yes',
                            ['Dimension', 'Category', 'Time_Period']
                        ].drop_duplicates().shape[0]
                    )
                    metadata['structural_infeasible_validation_rows'] = structural_rows
                    metadata['structural_infeasible_validation_categories'] = structural_categories
        
        # Build method breakdown tab (final tab)
        method_breakdown_df = None
        if consistent_weights:
            rows = []
            dims_all = list(dimensions)
            used_dims = set(getattr(analyzer, 'global_dimensions_used', []))
            removed_dims = set(getattr(analyzer, 'removed_dimensions', []))
            per_dim_dict: Dict[str, Dict[str, float]] = getattr(analyzer, 'per_dimension_weights', {})
            weight_methods: Dict[str, str] = getattr(analyzer, 'weight_methods', {})
            global_w = getattr(analyzer, 'global_weights', {})
            peers = set(global_w.keys())
            for d, wmap in per_dim_dict.items():
                peers.update(wmap.keys())
            for dim in dims_all:
                # Use weight_methods if available, otherwise determine from context
                if dim in weight_methods:
                    method = weight_methods[dim]
                elif dim in per_dim_dict:
                    method = 'Per-Dimension-LP'
                elif dim in used_dims:
                    method = 'Global-LP'
                elif dim in removed_dims:
                    method = 'Global weights (dropped in LP)'
                else:
                    method = 'Global weights'
                # rows per peer
                peer_list = sorted(peers)
                for p in peer_list:
                    if dim in per_dim_dict and p in per_dim_dict[dim]:
                        mult = float(per_dim_dict[dim][p])
                    else:
                        mult = float(global_w.get(p, {}).get('multiplier', 1.0))
                    global_weight_pct = global_w.get(p, {}).get('weight', None)
                    rows.append({
                        'Dimension': dim,
                        'Method': method,
                        'Peer': p,
                        'Multiplier': round(mult, 6),
                        'Global_Weight_%': round(global_weight_pct, 4) if isinstance(global_weight_pct, (int, float)) else None
                    })
            if rows:
                method_breakdown_df = pd.DataFrame(rows)
        
        # ===================================
        # Preset Comparison (Phase 2 feature)
        # ===================================
        preset_comparison_df = None
        if include_preset_comparison:
            logger.info("\n=== Running Preset Comparison ===")
            preset_comparison_df = run_preset_comparison(
                df=df,
                metric_col=metric_col,
                entity_col=entity_col,
                dimensions=dimensions,
                target_entity=resolved_entity,
                time_col=time_col,
                analysis_type='share',
                logger=logger
            )
            if preset_comparison_df is not None and not preset_comparison_df.empty:
                logger.info(f"Preset comparison complete. Analyzed {len(preset_comparison_df)} presets.")
                metadata['preset_comparison'] = preset_comparison_df.to_dict('records')
            else:
                logger.warning("Preset comparison returned no results.")
        
        # ===================================
        # Impact Analysis (Phase 2 feature)
        # ===================================
        impact_df = None
        impact_summary_df = None
        if include_impact_summary and resolved_entity:
            logger.info("\n=== Computing Impact Analysis ===")
            try:
                # Calculate impact for all categories
                impact_df = analyzer.calculate_share_impact(df, metric_col, dimensions, resolved_entity)
                
                if impact_df is not None and not impact_df.empty:
                    logger.info(f"Impact analysis complete. Analyzed {len(impact_df)} categories.")
                    
                    # Calculate summary statistics
                    impact_summary = {
                        'mean_impact_pp': round(impact_df['Impact_PP'].mean(), 4),
                        'mean_abs_impact_pp': round(impact_df['Impact_PP'].abs().mean(), 4),
                        'std_impact_pp': round(impact_df['Impact_PP'].std(), 4) if len(impact_df) > 1 else 0.0,
                        'min_impact_pp': round(impact_df['Impact_PP'].min(), 4),
                        'max_impact_pp': round(impact_df['Impact_PP'].max(), 4),
                        'categories_analyzed': len(impact_df),
                    }
                    metadata['impact_summary'] = impact_summary
                    
                    # Create summary by dimension
                    if 'Dimension' in impact_df.columns:
                        dimension_summary = []
                        for dim in impact_df['Dimension'].unique():
                            dim_data = impact_df[impact_df['Dimension'] == dim]
                            dimension_summary.append({
                                'Dimension': dim,
                                'Mean_Abs_Impact_PP': round(dim_data['Impact_PP'].abs().mean(), 4),
                                'Max_Abs_Impact_PP': round(dim_data['Impact_PP'].abs().max(), 4),
                                'Categories': len(dim_data),
                            })
                        impact_summary_df = pd.DataFrame(dimension_summary)
                        logger.info("Impact summary by dimension:")
                        for row in dimension_summary:
                            logger.info(f"  {row['Dimension']}: Mean Abs Impact = {row['Mean_Abs_Impact_PP']:.4f} pp, Max = {row['Max_Abs_Impact_PP']:.4f} pp")
                    
                    metadata['impact_details'] = impact_df.to_dict('records')
                else:
                    logger.warning("Impact analysis returned no results.")
            except Exception as e:
                logger.error(f"Failed to compute impact analysis: {e}")
        elif include_impact_summary and not resolved_entity:
            logger.warning("Impact analysis requires a target entity. Use --entity to specify.")
        
        # Generate output
        entity_name = resolved_entity.replace(' ', '_') if resolved_entity else 'PEER_ONLY'
        analysis_output_file = args.output or f"benchmark_share_{entity_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        publication_output = None
        if output_format in ('analysis', 'both'):
            generate_excel_report(
                results,
                analysis_output_file,
                resolved_entity or 'PEER-ONLY',
                'share',
                logger,
                metadata,
                weights_df,
                method_breakdown_df,
                privacy_validation_df,
                secondary_results=secondary_results_df,
                preset_comparison_df=preset_comparison_df,
                impact_df=impact_df,
                impact_summary_df=impact_summary_df,
                validation_issues=validation_issues
            )
        
        if output_format in ('publication', 'both'):
            pub_path = Path(analysis_output_file)
            publication_output = str(pub_path.with_name(f"{pub_path.stem}_publication{pub_path.suffix}"))
            report_generator = ReportGenerator(config)
            report_generator.generate_publication_workbook(
                results=results,
                output_file=publication_output,
                analysis_type='share',
                metadata=metadata
            )
        
        # Export balanced CSV if requested
        csv_output = None
        if getattr(args, 'export_balanced_csv', False):
            export_balanced_csv(
                results, analysis_output_file, logger, 
                analysis_type='share',
                df=df,
                analyzer=analyzer,
                dimensions=dimensions,
                metric_col=metric_col,
                secondary_metrics=secondary_metrics,
                include_calculated=include_calculated_metrics
            )
            csv_output = analysis_output_file.rsplit('.', 1)[0] + '_balanced.csv'

        # Build report paths list for audit/summary
        report_paths = build_report_paths(output_format, analysis_output_file, publication_output)

        # Audit log (dedicated)
        if include_audit_log:
            write_audit_log(
                config,
                analysis_output_file=analysis_output_file,
                metadata=metadata,
                report_paths=report_paths,
                dimensions_analyzed=len(results),
                csv_output=csv_output,
                impact_df=impact_df,
                privacy_validation_df=privacy_validation_df,
                validation_issues=validation_issues,
            )
        
        print(f"\n{'='*80}")
        print("SHARE ANALYSIS COMPLETE")
        print(f"{'='*80}")
        print(f"Entity: {resolved_entity if resolved_entity else 'PEER-ONLY MODE'}")
        print(f"Metric: {metric_col}")
        print(f"Dimensions Analyzed: {len(results)}")
        print(f"Report: {', '.join(report_paths)}")
        print(f"{'='*80}\n")
        
        return 0
        
    except Exception as e:
        logger.error(f"Analysis failed: {e}", exc_info=True)
        return 1


def run_rate_analysis(args: argparse.Namespace, logger: logging.Logger) -> int:
    """Execute rate-based dimensional analysis."""
    logger.info("Starting rate-based dimensional analysis")
    
    try:
        # Validate that at least one rate type is specified
        if not args.approved_col and not args.fraud_col:
            logger.error("At least one of --approved-col or --fraud-col must be specified")
            return 1
        
        config = build_run_config(
            args,
            extra_overrides={'fraud_in_bps': getattr(args, 'fraud_in_bps', None)},
        )
        output_settings = resolve_output_settings(config)
        include_preset_comparison = output_settings['include_preset_comparison']
        include_impact_summary = output_settings['include_impact_summary']
        include_calculated_metrics = output_settings['include_calculated_metrics']
        include_privacy_validation = output_settings['include_privacy_validation']
        include_audit_log = output_settings['include_audit_log']
        output_format = output_settings['output_format']
        fraud_in_bps = output_settings['fraud_in_bps']
        try:
            data_loader, df, entity_col, time_col = prepare_run_data(
                args,
                config,
                logger,
                preferred_entity_col=args.entity_col,
            )
        except ValueError as exc:
            logger.error(str(exc))
            return 1
        
        # Validate columns exist (use exact names as specified by user)
        total_col = args.total_col
        if total_col not in df.columns:
            logger.error(f"Total column '{total_col}' not found in data")
            logger.error(f"Available columns: {', '.join(df.columns.tolist())}")
            return 1
        
        rate_types = []
        numerator_cols = {}
        bic_percentiles = {}
        
        # Get BIC percentiles from config
        default_bic = config.get('analysis', 'best_in_class_percentile', default=0.85)
        fraud_bic = config.get('analysis', 'fraud_percentile', default=0.15)
        
        if args.approved_col:
            if args.approved_col not in df.columns:
                logger.error(f"Approved column '{args.approved_col}' not found in data")
                logger.error(f"Available columns: {', '.join(df.columns.tolist())}")
                return 1
            rate_types.append('approval')
            numerator_cols['approval'] = args.approved_col
            bic_percentiles['approval'] = default_bic  # Higher is better
            logger.info(f"Approval rate calculation: {args.approved_col} / {total_col}")
        
        if args.fraud_col:
            if args.fraud_col not in df.columns:
                logger.error(f"Fraud column '{args.fraud_col}' not found in data")
                logger.error(f"Available columns: {', '.join(df.columns.tolist())}")
                return 1
            rate_types.append('fraud')
            numerator_cols['fraud'] = args.fraud_col
            bic_percentiles['fraud'] = fraud_bic  # Lower is better for fraud
            logger.info(f"Fraud rate calculation: {args.fraud_col} / {total_col}")
        
        # ========================================
        # Input Data Validation
        # ========================================
        numerator_cols = {}
        if hasattr(args, 'approved_col') and args.approved_col:
            numerator_cols['approval'] = args.approved_col
        if hasattr(args, 'fraud_col') and args.fraud_col:
            numerator_cols['fraud'] = args.fraud_col

        validation_issues, should_abort = validate_analysis_input(
            df=df,
            config=config,
            data_loader=data_loader,
            analysis_type='rate',
            total_col=total_col,
            numerator_cols=numerator_cols,
            entity_col=entity_col,
            dimensions=args.dimensions,
            time_col=time_col,
            target_entity=args.entity,
        )
        if should_abort:
            return 1
        
        resolved_entity = resolve_target_entity(df, entity_col, args.entity, logger)
        if args.entity and resolved_entity is None:
            return 1

        logger.info(f"Using entity column: {entity_col}")
        logger.info(f"Analyzing rate types: {', '.join(rate_types)}")
        
        # Get unique entities and counts for metadata
        unique_entities = df[entity_col].nunique()
        total_records = len(df)
        
        # Determine dimensions
        debug_mode = config.get('output', 'include_debug_sheets', default=False)
        dimensions = resolve_dimensions(args, config, data_loader, df, logger)
        if dimensions is None:
            return 1
        
        # Get configuration values
        opt_config = config.config['optimization']
        analysis_config = config.config['analysis']

        # Initialize analyzer ONCE with first rate type's BIC (we'll override per-dimension)
        # The key insight: weights are based on total_col, not the numerator
        first_rate_type = rate_types[0]
        analyzer, analyzer_settings = _build_dimensional_analyzer(
            target_entity=resolved_entity,
            entity_col=entity_col,
            analysis_config=analysis_config,
            opt_config=opt_config,
            time_col=time_col,
            debug_mode=debug_mode,
            bic_percentile=bic_percentiles[first_rate_type],
            logger=logger,
        )
        consistent_weights = analyzer_settings['consistent_weights']
        consistency_mode = analyzer_settings['consistency_mode']
        rank_penalty_weight = analyzer_settings['rank_penalty_weight']
        rank_preservation_strength = analyzer_settings['rank_preservation_strength']
        lambda_penalty = analyzer_settings['lambda_penalty']
        bayesian_max_iterations = analyzer_settings['bayesian_max_iterations']
        bayesian_learning_rate = analyzer_settings['bayesian_learning_rate']
        violation_penalty_weight = analyzer_settings['violation_penalty_weight']
        enforce_single_weight_set = analyzer_settings['enforce_single_weight_set']
        dyn_constraints = analyzer_settings['dynamic_constraints_config']
        
        # Calculate global weights ONCE based on total_col
        # These weights apply to both rate types since they share the same denominator
        if consistent_weights:
            logger.info(f"\nCalculating global privacy-constrained weights based on {total_col}")
            analyzer.calculate_global_privacy_weights(df, total_col, dimensions)
            structural_summary = analyzer.get_structural_infeasibility_summary()
            if structural_summary.get('has_structural_infeasibility'):
                logger.warning(
                    "Structural infeasibility detected (dimensions=%s, categories=%s, worst_margin=%0.4fpp)",
                    structural_summary.get('infeasible_dimensions'),
                    structural_summary.get('infeasible_categories'),
                    structural_summary.get('worst_margin_pp'),
                )
            logger.info("Global weights will be used for all rate types")
        
        # Store results for each rate type
        all_results = {}
        
        # Analyze each rate type using the SAME weights
        for rate_type in rate_types:
            logger.info(f"{'='*60}")
            logger.info(f"Analyzing {rate_type.upper()} RATE")
            logger.info(f"{'='*60}")
            
            numerator_col = numerator_cols[rate_type]
            bic_percentile = bic_percentiles[rate_type]
            
            # Update BIC percentile for this rate type (used for BIC calculation, not weights)
            analyzer.bic_percentile = bic_percentile
            logger.info(f"Using {bic_percentile*100}th percentile for BIC")
            
            # Run analysis for this rate type
            results = {}
            for dim in dimensions:
                try:
                    result_df = analyzer.analyze_dimension_rate(
                        df=df,
                        dimension_column=dim,
                        total_col=total_col,
                        numerator_col=numerator_col
                    )
                    results[dim] = result_df
                except Exception as e:
                    logger.error(f"Error analyzing dimension {dim}: {e}")
                    continue
            
            if not results:
                logger.warning(f"No analysis results generated for {rate_type} rate")
                continue
            
            # Store results
            all_results[rate_type] = results
        
        if not all_results:
            logger.error("No analysis results generated for any rate type")
            return 1
        
        # Run secondary analysis if requested (Share analysis on the secondary metrics using Rate weights)
        secondary_results_df = None
        secondary_metrics = getattr(args, 'secondary_metrics', None)
        if secondary_metrics:
            logger.info(f"\nCalculating balanced metrics for {len(secondary_metrics)} secondary metric(s): {', '.join(secondary_metrics)}")
            logger.info(f"Using weights calculated from total column: {total_col}")
            
            secondary_results_df = get_balanced_metrics_df(
                df=df,
                analyzer=analyzer,
                dimensions=dimensions,
                total_col=total_col,
                secondary_metrics=secondary_metrics
            )
            logger.info(f"Secondary analysis complete. Generated {len(secondary_results_df)} rows of balanced data.")

        # Get optimization config for metadata
        opt_config = config.config.get('optimization', {})

        # Collect common metadata
        peer_count = unique_entities if resolved_entity is None else max(unique_entities - 1, 0)
        metadata = {
            'entity': resolved_entity or 'PEER-ONLY',
            'analysis_type': 'multi_rate' if len(all_results) > 1 else f'{rate_types[0]}_rate',
            'rate_types': rate_types,
            'secondary_metrics': secondary_metrics,
            'approved_col': getattr(args, 'approved_col', None),
            'fraud_col': getattr(args, 'fraud_col', None),
            'total_col': total_col,
            'entity_column': entity_col,
            'total_records': total_records,
            'unique_entities': unique_entities,
            'peer_count': peer_count,
            'dimensions_analyzed': len(dimensions),
            'dimension_names': dimensions,
            'preset': getattr(args, 'preset', None),
            'debug_mode': debug_mode,
            'consistent_weights': consistent_weights,
            'merchant_mode': config.get('analysis', 'merchant_mode', default=False),
            'include_debug_sheets': debug_mode,
            'include_privacy_validation': include_privacy_validation,
            'include_impact_summary': include_impact_summary,
            'include_preset_comparison': include_preset_comparison,
            'include_calculated_metrics': include_calculated_metrics,
            'output_format': output_format,
            'timestamp': datetime.now(),
            'input_csv': getattr(args, 'csv', None),
            'log_level': getattr(args, 'log_level', None),
            'dimensions_mode': 'manual' if bool(getattr(args, 'dimensions', None)) else ('auto' if getattr(args, 'auto', False) else 'manual'),
            'dimensions_requested': getattr(args, 'dimensions', None),
            'entity_col_arg': getattr(args, 'entity_col', None),
            'consistency_mode': consistency_mode,
            'enforce_single_weight_set': enforce_single_weight_set,
            'max_iterations': opt_config.get('linear_programming', {}).get('max_iterations'),
            'tolerance_pp': opt_config.get('linear_programming', {}).get('tolerance'),
            'max_weight': opt_config.get('bounds', {}).get('max_weight'),
            'min_weight': opt_config.get('bounds', {}).get('min_weight'),
            'volume_preservation_strength': opt_config.get('constraints', {}).get('volume_preservation'),
            'rank_penalty_weight': opt_config.get('linear_programming', {}).get('rank_penalty_weight'),
            'rank_preservation_strength': getattr(analyzer, 'rank_preservation_strength', None),
            'prefer_slacks_first': opt_config.get('subset_search', {}).get('prefer_slacks_first'),
            'bic_percentiles': bic_percentiles,  # Store both BIC percentiles
            # Extended optimization parameters
            'lambda_penalty': opt_config.get('linear_programming', {}).get('lambda_penalty'),
            'volume_weighted_penalties': opt_config.get('linear_programming', {}).get('volume_weighted_penalties'),
            'volume_weighting_exponent': opt_config.get('linear_programming', {}).get('volume_weighting_exponent'),
            'subset_search_enabled': opt_config.get('subset_search', {}).get('enabled'),
            'subset_search_strategy': opt_config.get('subset_search', {}).get('strategy'),
            'subset_search_max_tests': opt_config.get('subset_search', {}).get('max_attempts'),
            'subset_search_trigger_on_slack': opt_config.get('subset_search', {}).get('trigger_on_slack'),
            'subset_search_max_slack_threshold': opt_config.get('subset_search', {}).get('max_slack_threshold'),
            'bayesian_max_iterations': opt_config.get('bayesian', {}).get('max_iterations'),
            'bayesian_learning_rate': opt_config.get('bayesian', {}).get('learning_rate'),
            'violation_penalty_weight': opt_config.get('bayesian', {}).get('violation_penalty_weight'),
            'structural_infeasibility_summary': analyzer.get_structural_infeasibility_summary(),
            'impact_thresholds': {
                'high_pp': config.get(
                    'output',
                    'impact_thresholds',
                    'high_pp',
                    default=config.get('output', 'distortion_thresholds', 'high_distortion_pp', default=1.0)
                ),
                'low_pp': config.get(
                    'output',
                    'impact_thresholds',
                    'low_pp',
                    default=config.get('output', 'distortion_thresholds', 'low_distortion_pp', default=0.25)
                ),
            },
        }
        
        # Get weights data if debug mode (same weights for all rate types)
        weights_df = None
        privacy_validation_df = None
        export_csv = getattr(args, 'export_balanced_csv', False)
        
        if debug_mode:
            weights_df = analyzer.get_weights_dataframe()
            if not weights_df.empty:
                logger.info(f"Captured weights data: {len(weights_df)} weight entries")
        
        # Build privacy validation dataframe if debug mode OR CSV export is requested (based on total_col for rate analysis)
        if (include_privacy_validation or debug_mode or export_csv) and consistent_weights:
            privacy_validation_df = analyzer.build_privacy_validation_dataframe(df, total_col, dimensions)
            if not privacy_validation_df.empty:
                logger.info(f"Built privacy validation data: {len(privacy_validation_df)} validation entries")
                if 'Structural_Infeasible_Category' in privacy_validation_df.columns:
                    structural_rows = int((privacy_validation_df['Structural_Infeasible_Category'] == 'Yes').sum())
                    structural_categories = int(
                        privacy_validation_df.loc[
                            privacy_validation_df['Structural_Infeasible_Category'] == 'Yes',
                            ['Dimension', 'Category', 'Time_Period']
                        ].drop_duplicates().shape[0]
                    )
                    metadata['structural_infeasible_validation_rows'] = structural_rows
                    metadata['structural_infeasible_validation_categories'] = structural_categories
        
        # Build method breakdown tab (same for all rate types since weights are shared)
        method_breakdown_df = None
        if consistent_weights:
            rows = []
            dims_all = list(dimensions)
            used_dims = set(getattr(analyzer, 'global_dimensions_used', []))
            removed_dims = set(getattr(analyzer, 'removed_dimensions', []))
            per_dim_dict: Dict[str, Dict[str, float]] = getattr(analyzer, 'per_dimension_weights', {})
            weight_methods: Dict[str, str] = getattr(analyzer, 'weight_methods', {})
            global_w = getattr(analyzer, 'global_weights', {})
            peers = set(global_w.keys())
            for d, wmap in per_dim_dict.items():
                peers.update(wmap.keys())
            for dim in dims_all:
                # Use weight_methods if available, otherwise determine from context
                if dim in weight_methods:
                    method = weight_methods[dim]
                elif dim in per_dim_dict:
                    method = 'Per-Dimension-LP'
                elif dim in used_dims:
                    method = 'Global-LP'
                elif dim in removed_dims:
                    method = 'Global weights (dropped in LP)'
                else:
                    method = 'Global weights'
                # rows per peer
                peer_list = sorted(peers)
                for p in peer_list:
                    if dim in per_dim_dict and p in per_dim_dict[dim]:
                        mult = float(per_dim_dict[dim][p])
                    else:
                        mult = float(global_w.get(p, {}).get('multiplier', 1.0))
                    global_weight_pct = global_w.get(p, {}).get('weight', None)
                    rows.append({
                        'Dimension': dim,
                        'Method': method,
                        'Peer': p,
                        'Multiplier': round(mult, 6),
                        'Global_Weight_%': round(global_weight_pct, 4) if isinstance(global_weight_pct, (int, float)) else None
                    })
            if rows:
                method_breakdown_df = pd.DataFrame(rows)
                logger.info(f"Built method breakdown data: {len(method_breakdown_df)} entries")
        
        # ===================================
        # Preset Comparison (Phase 2 feature)
        # ===================================
        preset_comparison_df = None
        if include_preset_comparison:
            logger.info("\n=== Running Preset Comparison ===")
            preset_comparison_df = run_preset_comparison(
                df=df,
                metric_col=total_col,
                entity_col=entity_col,
                dimensions=dimensions,
                target_entity=resolved_entity,
                time_col=time_col,
                analysis_type='rate',
                logger=logger,
                total_col=total_col,
                numerator_cols=numerator_cols
            )
            if preset_comparison_df is not None and not preset_comparison_df.empty:
                logger.info(f"Preset comparison complete. Analyzed {len(preset_comparison_df)} presets.")
                metadata['preset_comparison'] = preset_comparison_df.to_dict('records')
            else:
                logger.warning("Preset comparison returned no results.")
        
        # ===================================
        # Impact Analysis (Phase 2 feature)
        # ===================================
        impact_df = None
        impact_summary_df = None
        if include_impact_summary:
            logger.info("\n=== Computing Impact Analysis ===")
            try:
                # Calculate impact for all categories
                impact_df = analyzer.calculate_rate_impact(df, total_col, numerator_cols, dimensions)
                
                if impact_df is not None and not impact_df.empty:
                    logger.info(f"Impact analysis complete. Analyzed {len(impact_df)} categories.")
                    
                    # Calculate summary statistics for each rate type
                    impact_summary = {}
                    
                    # Find impact columns
                    rate_cols = [c for c in impact_df.columns if c.endswith('_Impact_PP')]
                    
                    for col in rate_cols:
                        rate_name = col.replace('_Impact_PP', '')
                        impact_summary[f'{rate_name}_mean_abs_impact_pp'] = round(impact_df[col].abs().mean(), 4)
                        impact_summary[f'{rate_name}_max_abs_impact_pp'] = round(impact_df[col].abs().max(), 4)
                    
                    if rate_cols:
                        impact_summary['mean_abs_impact_pp'] = round(impact_df[rate_cols].abs().stack().mean(), 4)
                    
                    metadata['impact_summary'] = impact_summary
                    
                    # Create summary by dimension (taking the max impact across all rates)
                    if 'Dimension' in impact_df.columns:
                        dimension_summary = []
                        for dim in impact_df['Dimension'].unique():
                            dim_data = impact_df[impact_df['Dimension'] == dim]
                            
                            # Calculate max impact across all rate columns for this dimension
                            max_impacts = []
                            mean_impacts = []
                            for col in rate_cols:
                                max_impacts.append(dim_data[col].abs().max())
                                mean_impacts.append(dim_data[col].abs().mean())
                            
                            dimension_summary.append({
                                'Dimension': dim,
                                'Mean_Abs_Impact_PP': round(sum(mean_impacts) / len(mean_impacts), 4),
                                'Max_Abs_Impact_PP': round(max(max_impacts), 4),
                                'Categories': len(dim_data),
                            })
                        
                        impact_summary_df = pd.DataFrame(dimension_summary)
                        logger.info("Impact summary by dimension:")
                        for row in dimension_summary:
                            logger.info(f"  {row['Dimension']}: Mean Abs Impact = {row['Mean_Abs_Impact_PP']:.4f} pp, Max = {row['Max_Abs_Impact_PP']:.4f} pp")
                    
                    metadata['impact_details'] = impact_df.to_dict('records')
                else:
                    logger.warning("Impact analysis returned no results.")
            except Exception as e:
                logger.error(f"Failed to compute impact analysis: {e}")
        
        # Generate output file
        entity_name = resolved_entity.replace(' ', '_') if resolved_entity else 'PEER_ONLY'
        if args.output:
            analysis_output_file = args.output
        elif len(all_results) > 1:
            analysis_output_file = f"benchmark_multi_rate_{entity_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            analysis_output_file = f"benchmark_{rate_types[0]}_rate_{entity_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        publication_output = None
        
        # Generate report with all rate types
        if output_format in ('analysis', 'both'):
            generate_multi_rate_excel_report(
                all_results,
                analysis_output_file,
                resolved_entity or 'PEER-ONLY',
                logger,
                metadata,
                weights_df,
                numerator_cols,
                bic_percentiles,
                privacy_validation_df,
                method_breakdown_df,
                secondary_results=secondary_results_df,
                preset_comparison_df=preset_comparison_df,
                impact_df=impact_df,
                impact_summary_df=impact_summary_df,
                validation_issues=validation_issues
            )
        
        if output_format in ('publication', 'both'):
            publication_results: Dict[str, pd.DataFrame] = {}
            if len(all_results) > 1:
                for dimension in dimensions:
                    approval_df = all_results.get('approval', {}).get(dimension)
                    fraud_df = all_results.get('fraud', {}).get(dimension)
                    if approval_df is None and fraud_df is None:
                        continue
                    if approval_df is None:
                        combined_df = fraud_df.copy()
                    else:
                        combined_df = approval_df[['Category']].copy()
                        for col in approval_df.columns:
                            if col != 'Category':
                                combined_df[f'Approval_{col}'] = approval_df[col]
                        if fraud_df is not None:
                            for col in fraud_df.columns:
                                if col != 'Category':
                                    combined_df[f'Fraud_{col}'] = fraud_df[col]
                    publication_results[dimension] = combined_df
            else:
                rate_type = rate_types[0]
                publication_results = all_results.get(rate_type, {})
            
            pub_path = Path(analysis_output_file)
            publication_output = str(pub_path.with_name(f"{pub_path.stem}_publication{pub_path.suffix}"))
            report_generator = ReportGenerator(config)
            report_generator.generate_publication_workbook(
                results=publication_results,
                output_file=publication_output,
                analysis_type='rate',
                metadata=metadata,
                fraud_in_bps=fraud_in_bps
            )
        
        # Export balanced CSV if requested
        csv_output = None
        if getattr(args, 'export_balanced_csv', False):
            # Store secondary_metrics in analyzer for export function to access
            analyzer.secondary_metrics = secondary_metrics
            export_balanced_csv(
                None, analysis_output_file, logger, 
                analysis_type='rate', 
                all_results=all_results,
                df=df,
                analyzer=analyzer,
                dimensions=dimensions,
                total_col=total_col,
                numerator_cols=numerator_cols,
                include_calculated=include_calculated_metrics
            )
            csv_output = analysis_output_file.rsplit('.', 1)[0] + '_balanced.csv'

        # Build report paths list for audit/summary
        report_paths = build_report_paths(output_format, analysis_output_file, publication_output)

        # Audit log (dedicated)
        if include_audit_log:
            write_audit_log(
                config,
                analysis_output_file=analysis_output_file,
                metadata=metadata,
                report_paths=report_paths,
                dimensions_analyzed=len(dimensions),
                csv_output=csv_output,
                impact_df=impact_df,
                privacy_validation_df=privacy_validation_df,
                validation_issues=validation_issues,
            )
        
        # Print summary
        print(f"\n{'='*80}")
        print(f"RATE ANALYSIS COMPLETE")
        print(f"{'='*80}")
        if resolved_entity:
            print(f"Entity: {resolved_entity}")
        else:
            print(f"Mode: PEER-ONLY (No Target Entity)")
        print(f"Rate Types Analyzed: {', '.join([rt.upper() for rt in all_results.keys()])}")
        print(f"Dimensions Analyzed: {len(dimensions)}")
        print(f"Report: {', '.join(report_paths)}")
        if len(all_results) > 1:
            print(f"Note: Both rate types use same privacy-constrained weights (based on {args.total_col})")
        print(f"{'='*80}\n")
        
        return 0
        
    except Exception as e:
        logger.error(f"Analysis failed: {e}", exc_info=True)
        return 1


def _save_workbook_with_retries(
    wb: Any,
    output_file: str,
    logger: logging.Logger,
    max_attempts: int = 3,
) -> None:
    """Save workbook with retries and dedicated temp dir to avoid transient IO_WRITE failures."""
    output_path = Path(output_file)
    if output_path.parent:
        output_path.parent.mkdir(parents=True, exist_ok=True)

    tmp_dir = (Path.cwd() / ".openpyxl_tmp").resolve()
    tmp_dir.mkdir(parents=True, exist_ok=True)

    original_tempdir = tempfile.tempdir
    last_error: Optional[Exception] = None
    try:
        tempfile.tempdir = str(tmp_dir)
        backoff_sec = 0.5
        for attempt in range(1, max_attempts + 1):
            try:
                wb.save(str(output_path))
                return
            except Exception as exc:
                last_error = exc
                logger.warning(
                    "Workbook save attempt %s/%s failed for '%s': %s",
                    attempt,
                    max_attempts,
                    output_file,
                    exc,
                )
                gc.collect()
                if attempt < max_attempts:
                    time.sleep(backoff_sec)
                    backoff_sec *= 2
    finally:
        tempfile.tempdir = original_tempdir

    if last_error is not None:
        raise RuntimeError(f"IO_WRITE: Failed to save workbook after {max_attempts} attempts ({last_error})")
    raise RuntimeError(f"IO_WRITE: Failed to save workbook after {max_attempts} attempts")


def generate_excel_report(
    results: Dict[str, Any],
    output_file: str,
    entity_name: str,
    analysis_type: str,
    logger: logging.Logger,
    metadata: Optional[Dict[str, Any]] = None,
    weights_df: Optional[Any] = None,
    method_breakdown_df: Optional[pd.DataFrame] = None,
    privacy_validation_df: Optional[pd.DataFrame] = None,
    secondary_results: Optional[Dict[str, Any]] = None,
    preset_comparison_df: Optional[pd.DataFrame] = None,
    impact_df: Optional[pd.DataFrame] = None,
    impact_summary_df: Optional[pd.DataFrame] = None,
    validation_issues: Optional[Any] = None
) -> None:
    """Generate Excel report with dimensional analysis results."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise

    include_debug = True
    include_privacy_validation = True
    include_impact_summary = True
    include_preset_comparison = True
    if metadata:
        include_debug = bool(metadata.get('include_debug_sheets', metadata.get('debug_mode', True)))
        include_privacy_validation = bool(metadata.get('include_privacy_validation', True))
        include_impact_summary = bool(metadata.get('include_impact_summary', True))
        include_preset_comparison = bool(metadata.get('include_preset_comparison', True))

    if not include_debug:
        weights_df = None
    if not include_privacy_validation:
        privacy_validation_df = None
    if not include_impact_summary:
        impact_df = None
        impact_summary_df = None
    if not include_preset_comparison:
        preset_comparison_df = None
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Summary sheet with enhanced metadata
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 60
    
    row = 1
    ws_summary[f'A{row}'] = f"{analysis_type.upper()} ANALYSIS SUMMARY"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 2
    
    # Basic information
    ws_summary[f'A{row}'] = "BASIC INFORMATION"
    ws_summary[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    ws_summary[f'A{row}'] = "Target Entity:"
    ws_summary[f'B{row}'] = entity_name
    row += 1
    
    ws_summary[f'A{row}'] = "Analysis Type:"
    ws_summary[f'B{row}'] = analysis_type
    row += 1
    
    ws_summary[f'A{row}'] = "Timestamp:"
    ws_summary[f'B{row}'] = metadata.get('timestamp', datetime.now()).strftime('%Y-%m-%d %H:%M:%S') if metadata else datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    row += 2
    
    # Input parameters
    if metadata:
        ws_summary[f'A{row}'] = "INPUT PARAMETERS"
        ws_summary[f'A{row}'].font = Font(bold=True, size=11)
        row += 1

        def write_input(label: str, value: Any):
            nonlocal row
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value if value is not None else 'N/A'
            row += 1
        
        write_input("CSV Path:", metadata.get('input_csv'))
        write_input("Output File:", output_file)
        write_input("Entity Column (arg):", metadata.get('entity_col_arg', metadata.get('entity_column')))
        # Share/Rate specific inputs
        if metadata.get('analysis_type') == 'share':
            write_input("Metric:", metadata.get('metric'))
        else:
            write_input("Rate Type:", metadata.get('rate_type'))
            write_input("Numerator Column:", metadata.get('numerator_col'))
            write_input("Total Column:", metadata.get('total_col'))
            write_input("Fraud Mode:", metadata.get('fraud_mode'))
        # Dimension selection
        write_input("Dimensions Mode:", metadata.get('dimensions_mode'))
        dims_req = metadata.get('dimensions_requested')
        if dims_req:
            write_input("Dimensions Requested:", ", ".join(map(str, dims_req)))
        # General controls
        write_input("BIC Percentile:", metadata.get('bic_percentile'))
        write_input("Log Level:", metadata.get('log_level'))
        write_input("Preset:", metadata.get('preset'))
        write_input("Debug Mode:", 'ENABLED' if metadata.get('debug_mode') else 'DISABLED')
        write_input("Consistent Weights:", 'ENABLED' if metadata.get('consistent_weights') else 'DISABLED')
        write_input("Enforce Single Weight Set:", metadata.get('enforce_single_weight_set'))
        # Weight algorithm parameters
        write_input("Max Iterations:", metadata.get('max_iterations'))
        write_input("Tolerance (pp):", metadata.get('tolerance_pp'))
        write_input("Max Weight:", metadata.get('max_weight'))
        write_input("Min Weight:", metadata.get('min_weight'))
        write_input("Volume Preservation (input):", metadata.get('volume_preservation_strength'))
        write_input("Rank Preservation Strength:", metadata.get('rank_preservation_strength'))
        
        # Extended Optimization Parameters
        write_input("Lambda Penalty:", metadata.get('lambda_penalty'))
        write_input("Volume Weighted Penalties:", metadata.get('volume_weighted_penalties'))
        write_input("Volume Weighting Exponent:", metadata.get('volume_weighting_exponent'))
        write_input("Subset Search Enabled:", metadata.get('subset_search_enabled'))
        write_input("Subset Search Strategy:", metadata.get('subset_search_strategy'))
        write_input("Subset Search Max Tests:", metadata.get('subset_search_max_tests'))
        write_input("Subset Search Trigger on Slack:", metadata.get('subset_search_trigger_on_slack'))
        write_input("Subset Search Max Slack Threshold:", metadata.get('subset_search_max_slack_threshold'))
        write_input("Bayesian Max Iterations:", metadata.get('bayesian_max_iterations'))
        write_input("Bayesian Learning Rate:", metadata.get('bayesian_learning_rate'))
        write_input("Violation Penalty Weight:", metadata.get('violation_penalty_weight'))
        
        row += 1
    
    # Data information
    if metadata:
        ws_summary[f'A{row}'] = "DATA INFORMATION"
        ws_summary[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        ws_summary[f'A{row}'] = "Total Records:"
        ws_summary[f'B{row}'] = metadata.get('total_records', 'N/A')
        row += 1
        
        ws_summary[f'A{row}'] = "Entity Column:"
        ws_summary[f'B{row}'] = metadata.get('entity_column', 'N/A')
        row += 1
        
        ws_summary[f'A{row}'] = "Unique Entities in Data:"
        ws_summary[f'B{row}'] = metadata.get('unique_entities', 'N/A')
        row += 1
        
        ws_summary[f'A{row}'] = "Peer Count:"
        ws_summary[f'B{row}'] = metadata.get('peer_count', 'N/A')
        row += 1
        
        if 'metric' in metadata:
            ws_summary[f'A{row}'] = "Metric Analyzed:"
            ws_summary[f'B{row}'] = metadata.get('metric', 'N/A')
            row += 1
        
        if metadata.get('secondary_metrics'):
            ws_summary[f'A{row}'] = "Secondary Metrics:"
            ws_summary[f'B{row}'] = ', '.join(metadata.get('secondary_metrics', []))
            row += 1
        
        if 'rate_type' in metadata:
            ws_summary[f'A{row}'] = "Rate Type:"
            ws_summary[f'B{row}'] = metadata.get('rate_type', 'N/A')
            row += 1
            
            ws_summary[f'A{row}'] = "Numerator Column:"
            ws_summary[f'B{row}'] = metadata.get('numerator_col', 'N/A')
            row += 1
            
            ws_summary[f'A{row}'] = "Total Column:"
            ws_summary[f'B{row}'] = metadata.get('total_col', 'N/A')
            row += 1
        
        row += 1
        
        # Privacy & methodology
        ws_summary[f'A{row}'] = "PRIVACY & METHODOLOGY"
        ws_summary[f'A{row}'].font = Font(bold=True, size=11)
        row += 1
        
        ws_summary[f'A{row}'] = "Privacy Compliance:"
        ws_summary[f'B{row}'] = "Mastercard Control 3.2"
        row += 1
        
        # Determine which privacy rule applies (based on PEER count, not including target)
        peer_count = metadata.get('peer_count', 0)
        merchant_mode = bool(metadata.get('merchant_mode', False)) if metadata else False
        rule_name = PrivacyValidator.select_rule(peer_count, merchant_mode=merchant_mode)
        if rule_name == 'insufficient':
            min_required = 4 if merchant_mode else 5
            privacy_rule = (
                f"WARNING: Only {peer_count} peers - insufficient for privacy compliance "
                f"(minimum {min_required} required)"
            )
        else:
            rule_cfg = PrivacyValidator.get_rule_config(rule_name)
            min_entities = rule_cfg.get('min_entities', 'N/A') if rule_cfg else 'N/A'
            max_conc = rule_cfg.get('max_concentration', 'N/A') if rule_cfg else 'N/A'
            privacy_rule = f"{rule_name} Rule ({min_entities} peers min, {max_conc}% max concentration)"
        
        ws_summary[f'A{row}'] = "Applied Privacy Rule:"
        ws_summary[f'B{row}'] = privacy_rule
        row += 1
        
        ws_summary[f'A{row}'] = "BIC Percentile:"
        bic_pct = metadata.get('bic_percentile', 0.85)
        ws_summary[f'B{row}'] = f"{int(bic_pct * 100)}th percentile"
        row += 1
        
        ws_summary[f'A{row}'] = "Balanced Average Method:"
        ws_summary[f'B{row}'] = "Weighted (sum of peer values / sum of peer totals)"
        row += 1
        
        if metadata.get('debug_mode'):
            ws_summary[f'A{row}'] = "Debug Mode:"
            ws_summary[f'B{row}'] = "ENABLED (includes unweighted averages)"
            row += 1
        
        if metadata.get('preset'):
            ws_summary[f'A{row}'] = "Preset Used:"
            ws_summary[f'B{row}'] = metadata.get('preset')
            row += 1

        structural_summary = metadata.get('structural_infeasibility_summary', {}) if metadata else {}
        ws_summary[f'A{row}'] = "Structural Infeasibility:"
        if structural_summary.get('has_structural_infeasibility'):
            ws_summary[f'B{row}'] = "DETECTED"
        else:
            ws_summary[f'B{row}'] = "Not detected"
        row += 1
        if structural_summary.get('has_structural_infeasibility'):
            ws_summary[f'A{row}'] = "Infeasible Dimensions/Categories:"
            ws_summary[f'B{row}'] = (
                f"{structural_summary.get('infeasible_dimensions', 0)} / "
                f"{structural_summary.get('infeasible_categories', 0)}"
            )
            row += 1
            ws_summary[f'A{row}'] = "Worst Structural Margin (pp):"
            ws_summary[f'B{row}'] = structural_summary.get('worst_margin_pp')
            row += 1
            ws_summary[f'A{row}'] = "Top Structural Dimension:"
            ws_summary[f'B{row}'] = structural_summary.get('top_infeasible_dimension')
            row += 1
            ws_summary[f'A{row}'] = "Top Structural Category:"
            ws_summary[f'B{row}'] = structural_summary.get('top_infeasible_category')
            row += 1
        if metadata.get('structural_infeasible_validation_categories') is not None:
            ws_summary[f'A{row}'] = "Validation Rows in Structurally Infeasible Categories:"
            ws_summary[f'B{row}'] = metadata.get('structural_infeasible_validation_rows')
            row += 1
            ws_summary[f'A{row}'] = "Structurally Infeasible Categories in Validation:"
            ws_summary[f'B{row}'] = metadata.get('structural_infeasible_validation_categories')
            row += 1
        
        row += 1

    # If Rank Changes available, add a short top-movers section to Summary
    try:
        rank_df = None
        if metadata and 'analyzer_ref' in metadata:
            ana = metadata['analyzer_ref']
            rank_df = getattr(ana, 'rank_changes_df', None)
        if rank_df is not None and hasattr(rank_df, 'empty') and not rank_df.empty:
            ws_summary[f'A{row}'] = "RANK CHANGES (Top Movers)"
            ws_summary[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            # Top 5 by absolute Delta
            top = rank_df.copy()
            top['Abs_Delta'] = (top['Delta']).abs()
            top = top.sort_values(['Abs_Delta', 'Adjusted_Rank'], ascending=[False, True]).head(5)
            for _, r in top.iterrows():
                ws_summary[f'A{row}'] = str(r['Peer'])
                ws_summary[f'B{row}'] = f"Base->Adj: {int(r['Base_Rank'])}->{int(r['Adjusted_Rank'])} (Delta {int(r['Delta'])})"
                row += 1
            row += 1
    except Exception as e:
        logger.warning(f"Could not add Rank Changes summary: {e}")

    # Dimensions analyzed
    ws_summary[f'A{row}'] = "DIMENSIONS ANALYZED"
    ws_summary[f'A{row}'].font = Font(bold=True, size=11)
    row += 1
    
    ws_summary[f'A{row}'] = "Total Dimensions:"
    ws_summary[f'B{row}'] = len(results)
    row += 1
    row += 1
    
    ws_summary[f'A{row}'] = "Dimension"
    ws_summary[f'B{row}'] = "Categories"
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'].font = Font(bold=True)
    ws_summary[f'A{row}'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    ws_summary[f'B{row}'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for dim_name, dim_df in results.items():
        row += 1
        ws_summary[f'A{row}'] = dim_name
        ws_summary[f'B{row}'] = len(dim_df)
    
    # Create dimension sheets
    for dim_name, dim_df in results.items():
        sheet_name = dim_name[:31].replace('/', '_').replace('\\', '_')
        ws = wb.create_sheet(sheet_name)
        
        ws['A1'] = f"Dimension: {dim_name}"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Write headers
        for c_idx, col_name in enumerate(dim_df.columns, start=1):
            cell = ws.cell(row=3, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Write data
        for r_idx, row_data in enumerate(dim_df.itertuples(index=False), start=4):
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Auto-size columns for this dimension sheet
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Create secondary metrics summary sheet if available
    if secondary_results is not None and isinstance(secondary_results, pd.DataFrame) and not secondary_results.empty:
        logger.info("Adding Secondary Metrics Summary sheet")
        ws = wb.create_sheet("Secondary Metrics")
        
        ws['A1'] = "Secondary Metrics Summary"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f"(Using weights from primary metric: {metadata.get('metric', 'N/A')})"
        ws['A2'].font = Font(italic=True, size=9)
        
        # Write headers
        for c_idx, col_name in enumerate(secondary_results.columns, start=1):
            cell = ws.cell(row=4, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Write data
        for r_idx, row_data in enumerate(secondary_results.itertuples(index=False), start=5):
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Add Peer Weights tab if provided (outside the dimension loop)
    if weights_df is not None and hasattr(weights_df, 'empty') and not weights_df.empty:
        ws_weights = wb.create_sheet("Peer Weights")
        ws_weights['A1'] = "Peer Weights"
        ws_weights['A1'].font = Font(bold=True, size=12)

        ws_weights['A2'] = "Adjusted_Share = Adjusted_Volume / Sum(All Adjusted_Volumes) [capped at Max_Concentration]"
        ws_weights['A2'].font = Font(italic=True, size=9)
        ws_weights['A3'] = "Peer_Balanced_Avg = Sum(Metric * Adjusted_Share) - ensures no peer > Max_Concentration"
        ws_weights['A3'].font = Font(italic=True, size=9)

        # Write headers
        for c_idx, col_name in enumerate(weights_df.columns, start=1):
            cell = ws_weights.cell(row=5, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data
        for r_idx, row_data in enumerate(weights_df.itertuples(index=False), start=6):
            for c_idx, value in enumerate(row_data, start=1):
                ws_weights.cell(row=r_idx, column=c_idx, value=value)

        # Auto-size columns for Peer Weights
        for col in ws_weights.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws_weights.column_dimensions[column].width = min(max_length + 2, 50)

        logger.info("Added Peer Weights tab with weight calculations and contributions")

    # Add final Weight Methods tab if provided
    if method_breakdown_df is not None and not method_breakdown_df.empty:
        ws_methods = wb.create_sheet("Weight Methods")
        ws_methods['A1'] = "Dimension Weighting Methods"
        ws_methods['A1'].font = Font(bold=True, size=12)
        
        # Write headers
        for c_idx, col_name in enumerate(method_breakdown_df.columns, start=1):
            cell = ws_methods.cell(row=3, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Write data
        for r_idx, row_data in enumerate(method_breakdown_df.itertuples(index=False), start=4):
            for c_idx, value in enumerate(row_data, start=1):
                ws_methods.cell(row=r_idx, column=c_idx, value=value)
        
        # Auto-size columns
        for col in ws_methods.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws_methods.column_dimensions[column].width = min(max_length + 2, 60)

    # Subset Search tab if attempts were recorded
    try:
        attempts = metadata.get('subset_search_results') if metadata else None
        if attempts:
            ws_ss = wb.create_sheet("Subset Search")
            ws_ss['A1'] = "Auto Subset Search Attempts"
            ws_ss['A1'].font = Font(bold=True, size=12)
            # Stable columns aligned with analyzer keys
            cols_sorted = ['Attempt', 'Count', 'Dimensions', 'Success', 'Max_Slack', 'Sum_Slack', 'Method', 'Note']
            for c_idx, col_name in enumerate(cols_sorted, start=1):
                cell = ws_ss.cell(row=3, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for r_idx, d in enumerate(attempts, start=4):
                ws_ss.cell(row=r_idx, column=1, value=d.get('Attempt'))
                ws_ss.cell(row=r_idx, column=2, value=d.get('Count'))
                # Join list of dims into a string for Excel
                dims_val = d.get('Dimensions')
                if isinstance(dims_val, (list, tuple)):
                    dims_val = ", ".join(map(str, dims_val))
                ws_ss.cell(row=r_idx, column=3, value=dims_val)
                ws_ss.cell(row=r_idx, column=4, value=d.get('Success'))
                ws_ss.cell(row=r_idx, column=5, value=d.get('Max_Slack'))
                ws_ss.cell(row=r_idx, column=6, value=d.get('Sum_Slack'))
                ws_ss.cell(row=r_idx, column=7, value=d.get('Method'))
                ws_ss.cell(row=r_idx, column=8, value=d.get('Note'))
            for col in ws_ss.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_ss.column_dimensions[column].width = min(max_length + 2, 80)
    except Exception as e:
        logger.warning(f"Could not add Subset Search sheet: {e}")

    # New: Structural infeasibility diagnostics tabs if available via analyzer metadata
    try:
        structural_detail = None
        structural_summary = None
        if metadata and 'analyzer_ref' in metadata:
            ana = metadata['analyzer_ref']
            structural_detail = getattr(ana, 'structural_detail_df', None)
            structural_summary = getattr(ana, 'structural_summary_df', None)
        # fallback if metadata carries dataframes directly in future
        if structural_summary is not None and hasattr(structural_summary, 'empty') and not structural_summary.empty:
            ws_sum = wb.create_sheet("Structural Summary")
            ws_sum['A1'] = "Structural Infeasibility by Dimension"
            ws_sum['A1'].font = Font(bold=True, size=12)
            df_sum = structural_summary
            for c_idx, col_name in enumerate(df_sum.columns, start=1):
                cell = ws_sum.cell(row=3, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for r_idx, row_data in enumerate(df_sum.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row_data, start=1):
                    ws_sum.cell(row=r_idx, column=c_idx, value=value)
            for col in ws_sum.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_sum.column_dimensions[column].width = min(max_length + 2, 60)
        if structural_detail is not None and hasattr(structural_detail, 'empty') and not structural_detail.empty:
            ws_det = wb.create_sheet("Structural Detail")
            ws_det['A1'] = "Per-Category Structural Feasibility"
            ws_det['A1'].font = Font(bold=True, size=12)
            df_det = structural_detail
            for c_idx, col_name in enumerate(df_det.columns, start=1):
                cell = ws_det.cell(row=3, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for r_idx, row_data in enumerate(df_det.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row_data, start=1):
                    ws_det.cell(row=r_idx, column=c_idx, value=value)
            for col in ws_det.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_det.column_dimensions[column].width = min(max_length + 2, 80)
    except Exception as e:
        logger.warning(f"Could not add structural diagnostics sheets: {e}")

    # Rank Changes full sheet (Milestone 2)
    try:
        rank_df = None
        if metadata and 'analyzer_ref' in metadata:
            ana = metadata['analyzer_ref']
            rank_df = getattr(ana, 'rank_changes_df', None)
        if rank_df is not None and hasattr(rank_df, 'empty') and not rank_df.empty:
            ws_rank = wb.create_sheet("Rank Changes")
            ws_rank['A1'] = "Peer Rank Changes (Baseline vs Adjusted)"
            ws_rank['A1'].font = Font(bold=True, size=12)
            for c_idx, col_name in enumerate(rank_df.columns, start=1):
                cell = ws_rank.cell(row=3, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for r_idx, row_data in enumerate(rank_df.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row_data, start=1):
                    ws_rank.cell(row=r_idx, column=c_idx, value=value)
            for col in ws_rank.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_rank.column_dimensions[column].width = min(max_length + 2, 60)
    except Exception as e:
        logger.warning(f"Could not add Rank Changes sheet: {e}")

    # Privacy Validation sheet (debug mode)
    if privacy_validation_df is not None and not privacy_validation_df.empty:
        try:
            ws_validation = wb.create_sheet("Privacy Validation")
            ws_validation['A1'] = "Privacy Compliance Validation"
            ws_validation['A1'].font = Font(bold=True, size=12)
            ws_validation['A2'] = "Detailed breakdown showing original and balanced volume shares for each dimension-category-(time) combination"
            ws_validation['A2'].font = Font(italic=True)
            for c_idx, col_name in enumerate(privacy_validation_df.columns, start=1):
                cell = ws_validation.cell(row=4, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for r_idx, row_data in enumerate(privacy_validation_df.itertuples(index=False), start=5):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws_validation.cell(row=r_idx, column=c_idx, value=value)
                    # Highlight violations in red
                    if c_idx == privacy_validation_df.columns.get_loc('Compliant') + 1 and value == 'No':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            for col in ws_validation.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_validation.column_dimensions[column].width = min(max_length + 2, 50)
            logger.info("Added Privacy Validation tab")
        except Exception as e:
            logger.warning(f"Could not add Privacy Validation sheet: {e}")

    # Add Preset Comparison sheet if available
    if preset_comparison_df is not None and not preset_comparison_df.empty:
        try:
            ws_comparison = wb.create_sheet("Preset Comparison")
            
            # Add title
            ws_comparison['A1'] = "PRESET COMPARISON"
            ws_comparison['A1'].font = Font(bold=True, size=14)
            ws_comparison.merge_cells('A1:E1')
            
            # Write data starting at row 3
            for r_idx, row in enumerate(preset_comparison_df.itertuples(index=False), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_comparison.cell(row=r_idx, column=c_idx, value=value)
                    # Highlight best preset
                    if c_idx == len(row) and value == '*':
                        cell.font = Font(color="228B22", bold=True)
                        ws_comparison.cell(row=r_idx, column=1).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            # Add headers
            for c_idx, col_name in enumerate(preset_comparison_df.columns, 1):
                cell = ws_comparison.cell(row=2, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-fit columns (skip row 1 which has merged cells)
            from openpyxl.utils import get_column_letter
            for c_idx in range(1, len(preset_comparison_df.columns) + 1):
                max_length = 0
                column_letter = get_column_letter(c_idx)
                for r_idx in range(2, len(preset_comparison_df) + 4):  # Skip merged row 1
                    cell_value = ws_comparison.cell(row=r_idx, column=c_idx).value
                    try:
                        if cell_value and len(str(cell_value)) > max_length:
                            max_length = len(str(cell_value))
                    except Exception:
                        pass
                ws_comparison.column_dimensions[column_letter].width = min(max_length + 2, 30)
            
            logger.info("Added Preset Comparison tab")
        except Exception as e:
            logger.warning(f"Could not add Preset Comparison sheet: {e}")

    # Add Impact Analysis sheet if available
    if impact_df is not None and not impact_df.empty:
        try:
            ws_impact = wb.create_sheet("Impact Analysis")
            
            # Add title
            ws_impact['A1'] = "IMPACT ANALYSIS"
            ws_impact['A1'].font = Font(bold=True, size=14)
            
            row_offset = 2
            
            # Add summary if available
            if impact_summary_df is not None and not impact_summary_df.empty:
                ws_impact.cell(row=row_offset, column=1, value="SUMMARY BY DIMENSION").font = Font(bold=True, size=11)
                row_offset += 1
                
                # Headers
                for c_idx, col_name in enumerate(impact_summary_df.columns, 1):
                    cell = ws_impact.cell(row=row_offset, column=c_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                row_offset += 1
                
                # Data
                for _, row in impact_summary_df.iterrows():
                    for c_idx, value in enumerate(row, 1):
                        ws_impact.cell(row=row_offset, column=c_idx, value=value)
                    row_offset += 1
                
                row_offset += 2
            
            # Add detailed data
            ws_impact.cell(row=row_offset, column=1, value="DETAILED IMPACT BY CATEGORY").font = Font(bold=True, size=11)
            row_offset += 1
            
            # Headers
            for c_idx, col_name in enumerate(impact_df.columns, 1):
                cell = ws_impact.cell(row=row_offset, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            row_offset += 1
            
            # Data with conditional formatting
            for _, row in impact_df.iterrows():
                for c_idx, col_name in enumerate(impact_df.columns, 1):
                    value = row[col_name]
                    cell = ws_impact.cell(row=row_offset, column=c_idx, value=value)
                    
                    # Highlight impact values
                    if col_name == 'Impact_PP' and isinstance(value, (int, float)):
                        high_threshold = metadata.get('impact_thresholds', {}).get(
                            'high_pp',
                            metadata.get('distortion_thresholds', {}).get('high_distortion_pp', 1.0)
                        )
                        low_threshold = metadata.get('impact_thresholds', {}).get(
                            'low_pp',
                            metadata.get('distortion_thresholds', {}).get('low_distortion_pp', 0.25)
                        )

                        if abs(value) > high_threshold:
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif abs(value) < low_threshold:
                            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                row_offset += 1
            
            # Auto-fit columns
            from openpyxl.utils import get_column_letter
            for c_idx in range(1, len(impact_df.columns) + 1):
                max_length = 0
                column_letter = get_column_letter(c_idx)
                for r_idx in range(1, row_offset):
                    cell_value = ws_impact.cell(row=r_idx, column=c_idx).value
                    try:
                        if cell_value and len(str(cell_value)) > max_length:
                            max_length = len(str(cell_value))
                    except Exception:
                        pass
                ws_impact.column_dimensions[column_letter].width = min(max_length + 2, 35)
            
            logger.info("Added Impact Analysis tab")
        except Exception as e:
            logger.warning(f"Could not add Impact Analysis sheet: {e}")

    # Add Data Quality sheet if validation ran
    if validation_issues is not None:
        try:
            report_generator = ReportGenerator(None)
            has_errors = any(getattr(i, 'severity', None) == ValidationSeverity.ERROR for i in validation_issues)
            report_generator.add_data_quality_sheet(wb, validation_issues, passed=not has_errors)
        except Exception as e:
            logger.warning(f"Could not add Data Quality sheet: {e}")

    _save_workbook_with_retries(wb, output_file, logger)
    logger.info(f"Report saved to: {output_file}")


def generate_multi_rate_excel_report(
    all_results: Dict[str, Dict[str, Any]],
    output_file: str,
    entity_name: str,
    logger: logging.Logger,
    metadata: Dict[str, Any],
    weights_df: Optional[Any] = None,
    numerator_cols: Optional[Dict[str, str]] = None,
    bic_percentiles: Optional[Dict[str, float]] = None,
    privacy_validation_df: Optional[pd.DataFrame] = None,
    method_breakdown_df: Optional[pd.DataFrame] = None,
    secondary_results: Optional[Dict[str, Any]] = None,
    preset_comparison_df: Optional[pd.DataFrame] = None,
    impact_df: Optional[pd.DataFrame] = None,
    impact_summary_df: Optional[pd.DataFrame] = None,
    validation_issues: Optional[Any] = None
) -> None:
    """Generate Excel report with multiple rate types using shared weights.
    
    For multi-rate analysis, combines both rate types into the same dimension sheets
    with side-by-side columns for easy comparison.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise
    
    wb = Workbook()
    wb.remove(wb.active)
    
    rate_types = list(all_results.keys())
    is_multi_rate = len(rate_types) > 1
    debug_mode = metadata.get('debug_mode', False)
    
    # Summary sheet with enhanced metadata
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 60
    
    row = 1
    if is_multi_rate:
        ws_summary[f'A{row}'] = "MULTI-RATE ANALYSIS SUMMARY"
    else:
        ws_summary[f'A{row}'] = f"{rate_types[0].upper()} RATE ANALYSIS SUMMARY"
    ws_summary[f'A{row}'].font = Font(bold=True, size=14)
    row += 2
    
    # Basic information
    ws_summary[f'A{row}'] = "BASIC INFORMATION"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_summary[f'A{row}'] = "Entity"
    ws_summary[f'B{row}'] = entity_name
    row += 1
    
    if is_multi_rate:
        ws_summary[f'A{row}'] = "Analysis Type"
        ws_summary[f'B{row}'] = "Multi-Rate (Approval & Fraud)"
        row += 1
        
        ws_summary[f'A{row}'] = "Approval Column"
        ws_summary[f'B{row}'] = numerator_cols.get('approval', 'N/A') if numerator_cols else 'N/A'
        row += 1
        
        ws_summary[f'A{row}'] = "Fraud Column"
        ws_summary[f'B{row}'] = numerator_cols.get('fraud', 'N/A') if numerator_cols else 'N/A'
        row += 1
        
        ws_summary[f'A{row}'] = "Total Column (Shared Denominator)"
        ws_summary[f'B{row}'] = metadata.get('total_col', 'N/A')
        row += 1
        
        ws_summary[f'A{row}'] = "Approval BIC Percentile"
        ws_summary[f'B{row}'] = f"{bic_percentiles.get('approval', 0.85)*100:.0f}th (higher is better)" if bic_percentiles else "85th"
        row += 1
        
        ws_summary[f'A{row}'] = "Fraud BIC Percentile"
        ws_summary[f'B{row}'] = f"{bic_percentiles.get('fraud', 0.15)*100:.0f}th (lower is better)" if bic_percentiles else "15th"
        row += 1
    else:
        rate_type = rate_types[0]
        ws_summary[f'A{row}'] = "Analysis Type"
        ws_summary[f'B{row}'] = f"{rate_type.capitalize()} Rate"
        row += 1
        
        ws_summary[f'A{row}'] = f"{rate_type.capitalize()} Column"
        ws_summary[f'B{row}'] = numerator_cols.get(rate_type, 'N/A') if numerator_cols else 'N/A'
        row += 1
        
        ws_summary[f'A{row}'] = "Total Column"
        ws_summary[f'B{row}'] = metadata.get('total_col', 'N/A')
        row += 1
        
        ws_summary[f'A{row}'] = "BIC Percentile"
        ws_summary[f'B{row}'] = f"{bic_percentiles.get(rate_type, 0.85)*100:.0f}th" if bic_percentiles else "85th"
        row += 1
    
    if metadata.get('secondary_metrics'):
        ws_summary[f'A{row}'] = "Secondary Metrics:"
        ws_summary[f'B{row}'] = ', '.join(metadata.get('secondary_metrics', []))
        row += 1

    ws_summary[f'A{row}'] = "Dimensions Analyzed"
    ws_summary[f'B{row}'] = ', '.join(metadata.get('dimension_names', []))
    row += 1
    
    ws_summary[f'A{row}'] = "Total Records"
    ws_summary[f'B{row}'] = metadata.get('total_records', 'N/A')
    row += 1
    
    ws_summary[f'A{row}'] = "Unique Entities"
   
    ws_summary[f'B{row}'] = metadata.get('unique_entities', 'N/A')
    row += 1
    
    ws_summary[f'A{row}'] = "Peer Count"
    ws_summary[f'B{row}'] = metadata.get('peer_count', 'N/A')
    row += 1
    
    if metadata.get('consistent_weights', False):
        ws_summary[f'A{row}'] = "Weight Strategy"
        if is_multi_rate:
            ws_summary[f'B{row}'] = "Global weights (shared across all dimensions and rate types, based on total_col)"
        else:
            ws_summary[f'B{row}'] = "Global weights (shared across all dimensions)"
        ws_summary[f'B{row}'].font = Font(bold=True, color="0000FF")
        row += 1
    
    ws_summary[f'A{row}'] = "Timestamp"
    ws_summary[f'B{row}'] = str(metadata.get('timestamp', ''))
    row += 2

    # Optimization Parameters
    ws_summary[f'A{row}'] = "OPTIMIZATION PARAMETERS"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1

    def write_param(label, key, default='N/A'):
        nonlocal row
        ws_summary[f'A{row}'] = label
        val = metadata.get(key)
        ws_summary[f'B{row}'] = val if val is not None else default
        row += 1

    write_param("Preset", 'preset')
    write_param("Enforce Single Weight Set", 'enforce_single_weight_set')
    write_param("Max Iterations", 'max_iterations')
    write_param("Tolerance (pp)", 'tolerance_pp')
    write_param("Max Weight", 'max_weight')
    write_param("Min Weight", 'min_weight')
    write_param("Volume Preservation", 'volume_preservation_strength')
    write_param("Rank Preservation", 'rank_preservation_strength')
    write_param("Prefer Slacks First", 'prefer_slacks_first')
    
    # Extended Optimization Parameters
    write_param("Lambda Penalty", 'lambda_penalty')
    write_param("Volume Weighted Penalties", 'volume_weighted_penalties')
    write_param("Volume Weighting Exponent", 'volume_weighting_exponent')
    write_param("Subset Search Enabled", 'subset_search_enabled')
    write_param("Subset Search Strategy", 'subset_search_strategy')
    write_param("Subset Search Max Tests", 'subset_search_max_tests')
    write_param("Subset Search Trigger on Slack", 'subset_search_trigger_on_slack')
    write_param("Subset Search Max Slack Threshold", 'subset_search_max_slack_threshold')
    write_param("Bayesian Max Iterations", 'bayesian_max_iterations')
    write_param("Bayesian Learning Rate", 'bayesian_learning_rate')
    write_param("Violation Penalty Weight", 'violation_penalty_weight')
    
    row += 2

    structural_summary = metadata.get('structural_infeasibility_summary', {}) if metadata else {}
    ws_summary[f'A{row}'] = "STRUCTURAL FEASIBILITY"
    ws_summary[f'A{row}'].font = Font(bold=True)
    row += 1
    ws_summary[f'A{row}'] = "Structural Infeasibility"
    ws_summary[f'B{row}'] = "DETECTED" if structural_summary.get('has_structural_infeasibility') else "Not detected"
    row += 1
    if structural_summary.get('has_structural_infeasibility'):
        ws_summary[f'A{row}'] = "Infeasible Dimensions/Categories"
        ws_summary[f'B{row}'] = (
            f"{structural_summary.get('infeasible_dimensions', 0)} / "
            f"{structural_summary.get('infeasible_categories', 0)}"
        )
        row += 1
        ws_summary[f'A{row}'] = "Worst Structural Margin (pp)"
        ws_summary[f'B{row}'] = structural_summary.get('worst_margin_pp')
        row += 1
        ws_summary[f'A{row}'] = "Top Structural Dimension"
        ws_summary[f'B{row}'] = structural_summary.get('top_infeasible_dimension')
        row += 1
        ws_summary[f'A{row}'] = "Top Structural Category"
        ws_summary[f'B{row}'] = structural_summary.get('top_infeasible_category')
        row += 1
    if metadata.get('structural_infeasible_validation_categories') is not None:
        ws_summary[f'A{row}'] = "Validation Rows in Structurally Infeasible Categories"
        ws_summary[f'B{row}'] = metadata.get('structural_infeasible_validation_rows')
        row += 1
        ws_summary[f'A{row}'] = "Structurally Infeasible Categories in Validation"
        ws_summary[f'B{row}'] = metadata.get('structural_infeasible_validation_categories')
        row += 1

    row += 1
    
    # Add note about shared weights for multi-rate
    if is_multi_rate:
        ws_summary[f'A{row}'] = "IMPORTANT NOTE"
        ws_summary[f'A{row}'].font = Font(bold=True, color="FF0000")
        row += 1
        ws_summary[f'A{row}'] = "Privacy-constrained weights are calculated ONCE based on the total column."
       
        row += 1
        ws_summary[f'A{row}'] = "The same weights are applied to both approval and fraud rate calculations,"
        row += 1
        ws_summary[f'A{row}'] = "ensuring consistent privacy compliance across both analyses."
        row += 2
    
    # Create dimension sheets - combine rate types if multi-rate
    if is_multi_rate:
        # Get all dimensions from first rate type (should be same for all)
        dimensions = list(all_results[rate_types[0]].keys())
        
        for dimension in dimensions:
            sheet_name = dimension[:31] if len(dimension) > 31 else dimension
            ws = wb.create_sheet(sheet_name)
            
            # Get results for both rate types
            approval_df = all_results.get('approval', {}).get(dimension)
            fraud_df = all_results.get('fraud', {}).get(dimension)
            
            if approval_df is None or fraud_df is None:
                logger.warning(f"Missing data for dimension {dimension}")
                continue
            
            # Merge the dataframes side by side
            # Start with category column
            combined_df = approval_df[['Category']].copy()
            
            # Add approval rate columns
            for col in approval_df.columns:
                if col != 'Category':
                    if 'Unweighted' in col or 'Original' in col or 'Peer_Count' in col:
                        if debug_mode:
                            combined_df[f'Approval_{col}'] = approval_df[col]
                    else:
                        combined_df[f'Approval_{col}'] = approval_df[col]
            
            # Add fraud rate columns
            for col in fraud_df.columns:
                if col != 'Category':
                    if 'Unweighted' in col or 'Original' in col or 'Peer_Count' in col:
                        if debug_mode:
                            combined_df[f'Fraud_{col}'] = fraud_df[col]
                    else:
                        combined_df[f'Fraud_{col}'] = fraud_df[col]
            
            # Write headers with formatting
            for col_idx, col_name in enumerate(combined_df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True)
                
                # Color code by rate type
                if 'Approval_' in col_name:
                    cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Green
                elif 'Fraud_' in col_name:
                    cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")  # Orange
                else:
                    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Gray
            
            # Write data
            for row_idx, row_data in enumerate(combined_df.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
    else:
        # Single rate type - use original logic
        rate_type = rate_types[0]
        results = all_results[rate_type]
        
        for dimension, result_df in results.items():
            sheet_name = dimension[:31] if len(dimension) > 31 else dimension
            ws = wb.create_sheet(sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(result_df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            for row_idx, row_data in enumerate(result_df.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Create secondary metrics summary sheet if available
    if secondary_results is not None and isinstance(secondary_results, pd.DataFrame) and not secondary_results.empty:
        logger.info("Adding Secondary Metrics Summary sheet")
        ws = wb.create_sheet("Secondary Metrics")
        
        ws['A1'] = "Secondary Metrics Summary"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A2'] = f"(Using weights from total column: {metadata.get('total_col', 'N/A')})"
        ws['A2'].font = Font(italic=True, size=9)
        
        # Write headers
        for c_idx, col_name in enumerate(secondary_results.columns, start=1):
            cell = ws.cell(row=4, column=c_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Write data
        for r_idx, row_data in enumerate(secondary_results.itertuples(index=False), start=5):
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Add weights tab if available (same for all rate types)
    if weights_df is not None and not weights_df.empty:
        ws_weights = wb.create_sheet("Peer Weights")
        
        # Write headers
        for col_idx, col_name in enumerate(weights_df.columns, start=1):
            cell = ws_weights.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(weights_df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws_weights.cell(row=row_idx, column=col_idx, value=value)
        
        logger.info("Added Peer Weights tab with weight calculations and contributions")
    
    # Privacy Validation sheet (debug mode - same for all rate types)
    if privacy_validation_df is not None and not privacy_validation_df.empty:
        try:
            ws_validation = wb.create_sheet("Privacy Validation")
            ws_validation['A1'] = "Privacy Compliance Validation"
            ws_validation['A1'].font = Font(bold=True, size=12)
            ws_validation['A2'] = "Detailed breakdown showing original and balanced volume shares for each dimension-category-(time) combination"
            ws_validation['A2'].font = Font(italic=True)
            for c_idx, col_name in enumerate(privacy_validation_df.columns, start=1):
                cell = ws_validation.cell(row=4, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for r_idx, row_data in enumerate(privacy_validation_df.itertuples(index=False), start=5):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws_validation.cell(row=r_idx, column=c_idx, value=value)
                    # Highlight violations in red
                    if c_idx == privacy_validation_df.columns.get_loc('Compliant') + 1 and value == 'No':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            for col in ws_validation.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_validation.column_dimensions[column].width = min(max_length + 2, 50)
            logger.info("Added Privacy Validation tab")
        except Exception as e:
            logger.warning(f"Could not add Privacy Validation sheet: {e}")
    
    # Add final Weight Methods tab if provided (same for all rate types)
    if method_breakdown_df is not None and not method_breakdown_df.empty:
        try:
            ws_methods = wb.create_sheet("Weight Methods")
            ws_methods['A1'] = "Dimension Weighting Methods"
            ws_methods['A1'].font = Font(bold=True, size=12)
            ws_methods['A2'] = "Shows which calculation method was used for each dimension (Global LP, Per-Dimension LP, or Per-Dimension Bayesian)"
            ws_methods['A2'].font = Font(italic=True)
            
            # Write headers
            for c_idx, col_name in enumerate(method_breakdown_df.columns, start=1):
                cell = ws_methods.cell(row=4, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Write data
            for r_idx, row_data in enumerate(method_breakdown_df.itertuples(index=False), start=5):
                for c_idx, value in enumerate(row_data, start=1):
                    ws_methods.cell(row=r_idx, column=c_idx, value=value)
            
            # Auto-size columns
            for col in ws_methods.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws_methods.column_dimensions[column].width = min(max_length + 2, 60)
            
            logger.info("Added Weight Methods tab showing calculation methods per dimension")
        except Exception as e:
            logger.warning(f"Could not add Weight Methods sheet: {e}")
    
    # Add Preset Comparison sheet if available
    if preset_comparison_df is not None and not preset_comparison_df.empty:
        try:
            ws_comparison = wb.create_sheet("Preset Comparison")
            
            # Add title
            ws_comparison['A1'] = "PRESET COMPARISON"
            ws_comparison['A1'].font = Font(bold=True, size=14)
            ws_comparison.merge_cells('A1:E1')
            
            # Write data starting at row 3
            for r_idx, row in enumerate(preset_comparison_df.itertuples(index=False), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_comparison.cell(row=r_idx, column=c_idx, value=value)
                    # Highlight best preset
                    if c_idx == len(row) and value == '*':
                        cell.font = Font(color="228B22", bold=True)
                        ws_comparison.cell(row=r_idx, column=1).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            
            # Add headers
            for c_idx, col_name in enumerate(preset_comparison_df.columns, 1):
                cell = ws_comparison.cell(row=2, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Auto-fit columns (skip row 1 which has merged cells)
            from openpyxl.utils import get_column_letter
            for c_idx in range(1, len(preset_comparison_df.columns) + 1):
                max_length = 0
                column_letter = get_column_letter(c_idx)
                for r_idx in range(2, len(preset_comparison_df) + 4):  # Skip merged row 1
                    cell_value = ws_comparison.cell(row=r_idx, column=c_idx).value
                    try:
                        if cell_value and len(str(cell_value)) > max_length:
                            max_length = len(str(cell_value))
                    except Exception:
                        pass
                ws_comparison.column_dimensions[column_letter].width = min(max_length + 2, 30)
            
            logger.info("Added Preset Comparison tab")
        except Exception as e:
            logger.warning(f"Could not add Preset Comparison sheet: {e}")
    
    # Add Impact Analysis sheet if available
    if impact_df is not None and not impact_df.empty:
        try:
            ws_impact = wb.create_sheet("Impact Analysis")
            
            # Add title
            ws_impact['A1'] = "IMPACT ANALYSIS"
            ws_impact['A1'].font = Font(bold=True, size=14)
            
            row_offset = 2
            
            # Add summary if available
            if impact_summary_df is not None and not impact_summary_df.empty:
                ws_impact.cell(row=row_offset, column=1, value="SUMMARY BY DIMENSION").font = Font(bold=True, size=11)
                row_offset += 1
                
                # Headers
                for c_idx, col_name in enumerate(impact_summary_df.columns, 1):
                    cell = ws_impact.cell(row=row_offset, column=c_idx, value=col_name)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                row_offset += 1
                
                # Data
                for _, row in impact_summary_df.iterrows():
                    for c_idx, value in enumerate(row, 1):
                        ws_impact.cell(row=row_offset, column=c_idx, value=value)
                    row_offset += 1
                
                row_offset += 2
            
            # Add detailed data
            ws_impact.cell(row=row_offset, column=1, value="DETAILED IMPACT BY CATEGORY").font = Font(bold=True, size=11)
            row_offset += 1
            
            # Headers
            for c_idx, col_name in enumerate(impact_df.columns, 1):
                cell = ws_impact.cell(row=row_offset, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            row_offset += 1
            
            # Data with conditional formatting
            for _, row in impact_df.iterrows():
                for c_idx, col_name in enumerate(impact_df.columns, 1):
                    value = row[col_name]
                    cell = ws_impact.cell(row=row_offset, column=c_idx, value=value)
                    
                    # Highlight impact values using suffix check
                    if col_name.endswith('_Impact_PP') and isinstance(value, (int, float)):
                        high_threshold = metadata.get('impact_thresholds', {}).get(
                            'high_pp',
                            metadata.get('distortion_thresholds', {}).get('high_distortion_pp', 1.0)
                        )
                        low_threshold = metadata.get('impact_thresholds', {}).get(
                            'low_pp',
                            metadata.get('distortion_thresholds', {}).get('low_distortion_pp', 0.25)
                        )
                        if abs(value) > high_threshold:
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif abs(value) < low_threshold:
                            cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                row_offset += 1
            
            # Auto-fit columns
            from openpyxl.utils import get_column_letter
            for c_idx in range(1, len(impact_df.columns) + 1):
                max_length = 0
                column_letter = get_column_letter(c_idx)
                for r_idx in range(1, row_offset):
                    cell_value = ws_impact.cell(row=r_idx, column=c_idx).value
                    try:
                        if cell_value and len(str(cell_value)) > max_length:
                            max_length = len(str(cell_value))
                    except Exception:
                        pass
                ws_impact.column_dimensions[column_letter].width = min(max_length + 2, 35)
            
            logger.info("Added Impact Analysis tab")
        except Exception as e:
            logger.warning(f"Could not add Impact Analysis sheet: {e}")
    
    # Save workbook
    if validation_issues is not None:
        try:
            report_generator = ReportGenerator(None)
            has_errors = any(getattr(i, 'severity', None) == ValidationSeverity.ERROR for i in validation_issues)
            report_generator.add_data_quality_sheet(wb, validation_issues, passed=not has_errors)
        except Exception as e:
            logger.warning(f"Could not add Data Quality sheet: {e}")
    _save_workbook_with_retries(wb, output_file, logger)
    logger.info(f"Report saved to: {output_file}")


def get_balanced_metrics_df(
    df: pd.DataFrame,
    analyzer: DimensionalAnalyzer,
    dimensions: list,
    metric_col: Optional[str] = None,
    secondary_metrics: Optional[list] = None,
    total_col: Optional[str] = None,
    numerator_cols: Optional[Dict[str, str]] = None
) -> pd.DataFrame:
    """
    Calculate balanced metrics for primary and secondary metrics.
    Returns a DataFrame with columns: Dimension, Category, [Time], Balanced_{Metric}...
    """
    rows = []
    entity_col = analyzer.entity_column
    
    # Get global weights or per-dimension weights
    def get_weight(dimension: str, peer: str) -> float:
        """Get weight multiplier for a peer in a dimension."""
        if dimension in analyzer.per_dimension_weights and peer in analyzer.per_dimension_weights[dimension]:
            return float(analyzer.per_dimension_weights[dimension][peer])
        if hasattr(analyzer, 'global_weights') and peer in analyzer.global_weights:
            return float(analyzer.global_weights[peer].get('multiplier', 1.0))
        return 1.0
    
    # Check if time column is available
    time_col = analyzer.time_column if hasattr(analyzer, 'time_column') else None
    has_time = time_col and time_col in df.columns
    
    # Build list of metrics to calculate
    metrics_to_calculate = []
    
    # Share analysis primary metric
    if metric_col:
        metrics_to_calculate.append(('Primary', metric_col))
        
    # Rate analysis metrics
    if total_col:
        metrics_to_calculate.append(('Total', total_col))
        if numerator_cols:
            for key, col in numerator_cols.items():
                if col in df.columns:
                    metrics_to_calculate.append((f'Approval_{key.capitalize()}' if key == 'approval' else f'Fraud_{key.capitalize()}' if key == 'fraud' else key.capitalize(), col))

    # Secondary metrics
    if secondary_metrics:
        for sec_metric in secondary_metrics:
            if sec_metric in df.columns:
                metrics_to_calculate.append(('Secondary', sec_metric))
    
    # Process each dimension
    for dimension in dimensions:
        # Aggregate data by entity, dimension category, and optionally time
        group_cols = [entity_col, dimension]
        # Only add time_col if it's different from the current dimension
        if has_time and time_col != dimension:
            group_cols.append(time_col)
        
        # Build aggregation dict - exclude columns that are part of groupby
        agg_dict = {}
        for _, metric in metrics_to_calculate:
            if metric in df.columns and metric not in group_cols:
                agg_dict[metric] = 'sum'
        
        if not agg_dict:
            continue
            
        entity_dim_agg = df.groupby(group_cols).agg(agg_dict).reset_index()
        
        # Get unique categories
        categories = entity_dim_agg[dimension].unique()
        
        # Get unique time periods if applicable
        time_periods = sorted(entity_dim_agg[time_col].unique()) if has_time else [None]
        
        for category in categories:
            for time_period in time_periods:
                # Filter to this category (and time period if applicable)
                if has_time:
                    cat_df = entity_dim_agg[
                        (entity_dim_agg[dimension] == category) & 
                        (entity_dim_agg[time_col] == time_period)
                    ]
                else:
                    cat_df = entity_dim_agg[entity_dim_agg[dimension] == category]
                
                if cat_df.empty:
                    continue
                
                # Build row data
                row_data = {
                    'Dimension': dimension,
                    'Category': category,
                }
                
                # Add time period if applicable
                if has_time:
                    row_data[time_col] = time_period
                
                # Calculate balanced metric for each metric column
                for metric_type, metric in metrics_to_calculate:
                    if metric not in cat_df.columns:
                        continue
                        
                    balanced_metric = 0.0
                    for _, row in cat_df.iterrows():
                        peer = row[entity_col]
                        weight = get_weight(dimension, peer)
                        balanced_metric += row[metric] * weight
                    
                    # Column name based on metric type
                    if metric_type == 'Primary':
                        col_name = f'Balanced_{metric}'
                    elif metric_type == 'Total':
                        col_name = 'Balanced_Total'
                    elif metric_type.startswith('Approval'):
                        col_name = 'Balanced_Approval_Total' # Standardize name for rate analysis
                    elif metric_type.startswith('Fraud'):
                        col_name = 'Balanced_Fraud_Total' # Standardize name for rate analysis
                    elif metric_type == 'Secondary':
                        col_name = metric
                    else:
                        col_name = f'Balanced_{metric}'
                        
                    row_data[col_name] = round(balanced_metric, 2)
                
                rows.append(row_data)
    
    if not rows:
        return pd.DataFrame()
    
    # Create DataFrame
    result_df = pd.DataFrame(rows)
    
    # Sort by Dimension, Time (if present), then Category
    sort_cols = ['Dimension']
    if has_time and time_col in result_df.columns:
        sort_cols.append(time_col)
    sort_cols.append('Category')
    
    return result_df.sort_values(sort_cols)


def export_balanced_csv(
    results: Optional[Dict[str, pd.DataFrame]], 
    output_file: str, 
    logger: logging.Logger,
    analysis_type: str = 'share',
    all_results: Optional[Dict[str, Dict[str, pd.DataFrame]]] = None,
    df: Optional[pd.DataFrame] = None,
    analyzer: Optional[Any] = None,
    dimensions: Optional[list] = None,
    total_col: Optional[str] = None,
    numerator_cols: Optional[Dict[str, str]] = None,
    metric_col: Optional[str] = None,
    secondary_metrics: Optional[list] = None,
    include_calculated: bool = False
) -> None:
    """
    Export balanced metrics to CSV in concatenated dimension format.
    
    For rate analysis: exports dimension, category, balanced_total, balanced_approval_total, balanced_fraud_total
    The balanced totals are weighted sums: sum(peer_value * weight) across all peers
    
    For share analysis: exports dimension, category, balanced metric values (for primary and secondary metrics)
    The balanced metrics are weighted sums: sum(peer_metric_value * weight) across all peers
    
    Args:
        results: Results dictionary for share analysis (dimension -> DataFrame)
        output_file: Base output file path (will change extension to .csv)
        logger: Logger instance
        analysis_type: Type of analysis ('share' or 'rate')
        all_results: For rate analysis, dict of rate_type -> results (e.g., {'approval': {...}, 'fraud': {...}})
        df: Source dataframe for recalculating weighted totals
        analyzer: Analyzer instance with weights
        dimensions: List of dimensions analyzed
        total_col: Total column name (for rate analysis)
        numerator_cols: Dict mapping rate_type to numerator column name
        metric_col: Primary metric column name (for share analysis)
        secondary_metrics: List of secondary metric columns (for share analysis)
    """
    # Create CSV filename from Excel filename
    csv_output = output_file.rsplit('.', 1)[0] + '_balanced.csv'
    
    if analysis_type == 'rate' and all_results and df is not None and analyzer is not None:
        # Rate analysis: calculate weighted totals for each dimension-category
        export_rows = []
        
        # Get entity column and weights
        entity_col = analyzer.entity_column
        
        # Get global weights or per-dimension weights
        def get_weight(dimension: str, peer: str) -> float:
            """Get weight multiplier for a peer in a dimension."""
            weight = 1.0
            if hasattr(analyzer, 'global_weights') and peer in analyzer.global_weights:
                weight = float(analyzer.global_weights[peer].get('multiplier', 1.0))
            if dimension in analyzer.per_dimension_weights and peer in analyzer.per_dimension_weights[dimension]:
                weight = float(analyzer.per_dimension_weights[dimension][peer])
            return weight
        
        # Check if time column is available
        time_col = analyzer.time_column if hasattr(analyzer, 'time_column') else None
        has_time = time_col and time_col in df.columns
        
        # Get secondary metrics from caller
        secondary_metrics_list = getattr(analyzer, 'secondary_metrics', None)
        
        # Process each dimension
        for dimension in dimensions:
            # Aggregate data by entity, dimension category, and optionally time
            group_cols = [entity_col, dimension]
            # Only add time_col if it's different from the current dimension
            if has_time and time_col != dimension:
                group_cols.append(time_col)
            
            # Build aggregation dict - exclude columns that are part of groupby
            agg_dict = {}
            if total_col not in group_cols:
                agg_dict[total_col] = 'sum'
            
            # Add numerator columns (exclude those in group_cols)
            if numerator_cols:
                for num_col in numerator_cols.values():
                    if num_col in df.columns and num_col not in group_cols:
                        agg_dict[num_col] = 'sum'
            
            # Add secondary metrics (exclude those in group_cols)
            if secondary_metrics_list:
                for sec_metric in secondary_metrics_list:
                    if sec_metric in df.columns and sec_metric not in group_cols:
                        agg_dict[sec_metric] = 'sum'
            
            if not agg_dict:
                continue
                
            entity_dim_agg = df.groupby(group_cols).agg(agg_dict).reset_index()
            
            # Get unique categories
            categories = entity_dim_agg[dimension].unique()
            
            # Get unique time periods if applicable
            time_periods = sorted(entity_dim_agg[time_col].unique()) if has_time else [None]
            
            for category in categories:
                for time_period in time_periods:
                    # Filter to this category (and time period if applicable)
                    if has_time:
                        cat_df = entity_dim_agg[
                            (entity_dim_agg[dimension] == category) & 
                            (entity_dim_agg[time_col] == time_period)
                        ]
                    else:
                        cat_df = entity_dim_agg[entity_dim_agg[dimension] == category]
                    
                    if cat_df.empty:
                        continue
                    
                    # Calculate weighted totals
                    balanced_total = 0.0
                    balanced_approval = 0.0
                    balanced_fraud = 0.0
                    secondary_balanced = {}
                    
                    # Calculate raw totals if requested
                    raw_total = 0.0
                    raw_approval = 0.0
                    raw_fraud = 0.0
                    secondary_raw = {}

                    # Peer-only totals for weight-effect calculations
                    raw_peer_total = 0.0
                    raw_peer_approval = 0.0
                    raw_peer_fraud = 0.0
                    balanced_peer_total = 0.0
                    balanced_peer_approval = 0.0
                    balanced_peer_fraud = 0.0
                    
                    # Initialize secondary metrics dict
                    if secondary_metrics_list:
                        for sec_metric in secondary_metrics_list:
                            if sec_metric in cat_df.columns:
                                secondary_balanced[sec_metric] = 0.0
                                if include_calculated:
                                    secondary_raw[sec_metric] = 0.0
                    
                    for _, row in cat_df.iterrows():
                        peer = row[entity_col]
                        weight = get_weight(dimension, peer)
                        is_target = analyzer.target_entity is not None and peer == analyzer.target_entity

                        # Balanced totals should always be peer-only
                        if not is_target:
                            # Weighted total (denominator)
                            balanced_total += row[total_col] * weight
                            balanced_peer_total += row[total_col] * weight

                            # Weighted approval numerator
                            if approval_col := numerator_cols.get('approval'):
                                if approval_col in row.index:
                                    balanced_approval += row[approval_col] * weight
                                    balanced_peer_approval += row[approval_col] * weight

                            # Weighted fraud numerator
                            if fraud_col := numerator_cols.get('fraud'):
                                if fraud_col in row.index:
                                    balanced_fraud += row[fraud_col] * weight
                                    balanced_peer_fraud += row[fraud_col] * weight

                            # Weighted secondary metrics
                            for sec_metric in secondary_balanced.keys():
                                if sec_metric in row.index:
                                    secondary_balanced[sec_metric] += row[sec_metric] * weight

                        # Raw totals if requested (peer-only)
                        if include_calculated and not is_target:
                            raw_total += row[total_col]
                            raw_peer_total += row[total_col]

                            if approval_col := numerator_cols.get('approval'):
                                if approval_col in row.index:
                                    raw_approval += row[approval_col]
                                    raw_peer_approval += row[approval_col]

                            if fraud_col := numerator_cols.get('fraud'):
                                if fraud_col in row.index:
                                    raw_fraud += row[fraud_col]
                                    raw_peer_fraud += row[fraud_col]

                            for sec_metric in secondary_raw.keys():
                                if sec_metric in row.index:
                                    secondary_raw[sec_metric] += row[sec_metric]
                    
                    # Add row to export
                    row_data = {
                        'Dimension': dimension,
                        'Category': category,
                    }
                    
                    if include_calculated:
                        # Add raw and balanced rates and impact
                        if total_col:
                            row_data['Raw_Total'] = raw_total
                            row_data['Balanced_Total'] = balanced_total
                        
                        # Approval
                        if numerator_cols.get('approval'):
                            row_data['Balanced_Approval_Total'] = balanced_approval
                            # Calculate rates in %
                            raw_rate = (raw_peer_approval / raw_peer_total * 100) if raw_peer_total > 0 else 0.0
                            bal_rate = (balanced_peer_approval / balanced_peer_total * 100) if balanced_peer_total > 0 else 0.0
                            row_data['Raw_Approval_Rate_%'] = raw_rate
                            row_data['Balanced_Approval_Rate_%'] = bal_rate
                            row_data['Approval_Impact_PP'] = bal_rate - raw_rate
                            
                        # Fraud
                        if numerator_cols.get('fraud'):
                            row_data['Balanced_Fraud_Total'] = balanced_fraud
                            # Calculate rates in pp for consistency with impact
                            raw_fraud_rate = (raw_peer_fraud / raw_peer_total * 100) if raw_peer_total > 0 else 0.0
                            bal_fraud_rate = (balanced_peer_fraud / balanced_peer_total * 100) if balanced_peer_total > 0 else 0.0
                            row_data['Raw_Fraud_Rate_%'] = raw_fraud_rate
                            row_data['Balanced_Fraud_Rate_%'] = bal_fraud_rate
                            row_data['Fraud_Impact_PP'] = bal_fraud_rate - raw_fraud_rate
                    
                    # Add time period if applicable
                    if has_time:
                        row_data[time_col] = time_period
                    
                    # Use dynamic column names based on input columns
                    row_data[total_col] = round(balanced_total, 2)
                    
                    if numerator_cols:
                        if approval_col := numerator_cols.get('approval'):
                            row_data[approval_col] = round(balanced_approval, 2) if balanced_approval > 0 else None
                        if fraud_col := numerator_cols.get('fraud'):
                            row_data[fraud_col] = round(balanced_fraud, 2) if balanced_fraud > 0 else None
                    
                    # Add secondary metrics to row
                    for sec_metric, sec_value in secondary_balanced.items():
                        row_data[sec_metric] = round(sec_value, 2)
                    
                    export_rows.append(row_data)
        
        if not export_rows:
            logger.warning("No data to export for rate analysis CSV")
            return
        
        # Create DataFrame and export
        export_df = pd.DataFrame(export_rows)
        
        # Sort by Dimension, Time (if present), then Category
        sort_cols = ['Dimension']
        if has_time and time_col in export_df.columns:
            sort_cols.append(time_col)
        sort_cols.append('Category')
        
        export_df = export_df.sort_values(sort_cols)
        export_df.to_csv(csv_output, index=False)
        logger.info(f"Balanced rate data CSV exported to: {csv_output}")
        print(f"Balanced CSV: {csv_output}")
        
    elif analysis_type == 'share' and results and df is not None and analyzer is not None:
        # Share analysis: calculate balanced metrics for each dimension-category
        export_rows = []
        
        # Get entity column and weights
        entity_col = analyzer.entity_column
        
        # Get global weights or per-dimension weights
        def get_weight(dimension: str, peer: str) -> float:
            """Get weight multiplier for a peer in a dimension."""
            weight = 1.0
            if hasattr(analyzer, 'global_weights') and peer in analyzer.global_weights:
                weight = float(analyzer.global_weights[peer].get('multiplier', 1.0))
            if dimension in analyzer.per_dimension_weights and peer in analyzer.per_dimension_weights[dimension]:
                weight = float(analyzer.per_dimension_weights[dimension][peer])
            return weight
        
        # Check if time column is available
        time_col = analyzer.time_column if hasattr(analyzer, 'time_column') else None
        has_time = time_col and time_col in df.columns
        
        # Build list of metrics to calculate (primary + secondary)
        metrics_to_calculate = []
        if metric_col:
            metrics_to_calculate.append(('Primary', metric_col))
        if secondary_metrics:
            for sec_metric in secondary_metrics:
                if sec_metric in df.columns:
                    metrics_to_calculate.append(('Secondary', sec_metric))
        
        target_entity = analyzer.target_entity
        if include_calculated and not target_entity:
            logger.info("Peer-only mode: skipping target-vs-peer share calculations in balanced CSV.")

        # Process each dimension
        for dimension in dimensions:
            # Aggregate data by entity, dimension category, and optionally time
            group_cols = [entity_col, dimension]
            # Only add time_col if it's different from the current dimension
            if has_time and time_col != dimension:
                group_cols.append(time_col)

            # Build aggregation dict for all metrics - exclude columns that are part of groupby
            agg_dict = {}
            for _, metric in metrics_to_calculate:
                if metric in df.columns and metric not in group_cols:
                    agg_dict[metric] = 'sum'

            if not agg_dict:
                continue

            entity_dim_agg = df.groupby(group_cols).agg(agg_dict).reset_index()

            # Get unique categories
            categories = entity_dim_agg[dimension].unique()

            # Get unique time periods if applicable
            time_periods = sorted(entity_dim_agg[time_col].unique()) if has_time else [None]

            for category in categories:
                for time_period in time_periods:
                    # Filter to this category (and time period if applicable)
                    if has_time:
                        cat_df = entity_dim_agg[
                            (entity_dim_agg[dimension] == category) &
                            (entity_dim_agg[time_col] == time_period)
                        ]
                    else:
                        cat_df = entity_dim_agg[entity_dim_agg[dimension] == category]

                    if cat_df.empty:
                        continue

                    # Calculate metrics
                    balanced_metric_values = {}
                    raw_metric_values = {}

                    # Initialize values
                    for m_type, m_col in metrics_to_calculate:
                        balanced_metric_values[m_col] = 0.0
                        if include_calculated:
                            raw_metric_values[m_col] = 0.0

                    for _, row in cat_df.iterrows():
                        peer = row[entity_col]
                        if analyzer.target_entity is not None and peer == analyzer.target_entity:
                            continue
                        weight = get_weight(dimension, peer)

                        for m_type, m_col in metrics_to_calculate:
                            if m_col in row:
                                val = row[m_col]
                                balanced_metric_values[m_col] += val * weight
                                if include_calculated:
                                    raw_metric_values[m_col] += val

                    # Add row to export
                    row_data = {
                        'Dimension': dimension,
                        'Category': category,
                    }

                    # Add metric values
                    for m_type, m_col in metrics_to_calculate:
                        # Use original metric name for column header
                        clean_name = m_col
                        row_data[f'Balanced_{clean_name}'] = round(balanced_metric_values[m_col], 2)

                        if include_calculated:
                            row_data[f'Raw_{clean_name}'] = round(raw_metric_values[m_col], 2)

                            # Calculate shares and impact
                            if target_entity:
                                target_rows = cat_df[cat_df[entity_col] == target_entity]
                                peer_rows = cat_df[cat_df[entity_col] != target_entity]
                                target_val = float(target_rows[m_col].sum()) if not target_rows.empty else 0.0
                                raw_peer_total = float(peer_rows[m_col].sum()) if not peer_rows.empty else 0.0
                                balanced_peer_total = 0.0
                                for _, prow in peer_rows.iterrows():
                                    peer = prow[entity_col]
                                    weight = get_weight(dimension, peer)
                                    balanced_peer_total += float(prow[m_col]) * weight
                                raw_denom = target_val + raw_peer_total
                                bal_denom = target_val + balanced_peer_total
                                raw_share = (target_val / raw_denom * 100.0) if raw_denom > 0 else 0.0
                                bal_share = (target_val / bal_denom * 100.0) if bal_denom > 0 else 0.0
                                row_data[f'Raw_{clean_name}_Share_%'] = round(raw_share, 4)
                                row_data[f'Balanced_{clean_name}_Share_%'] = round(bal_share, 4)
                                row_data[f'{clean_name}_Impact_PP'] = round(bal_share - raw_share, 4)
                            else:
                                row_data[f'Raw_{clean_name}_Share_%'] = None
                                row_data[f'Balanced_{clean_name}_Share_%'] = None
                                row_data[f'{clean_name}_Impact_PP'] = None

                    # Add time period if applicable
                    if has_time:
                        row_data[time_col] = time_period

                    export_rows.append(row_data)
        
        if not export_rows:
            logger.warning("No data to export for share analysis CSV")
            return
        
        # Create DataFrame and export
        export_df = pd.DataFrame(export_rows)
        
        # Sort by Dimension, Time (if present), then Category
        sort_cols = ['Dimension']
        if has_time and time_col in export_df.columns:
            sort_cols.append(time_col)
        sort_cols.append('Category')
        
        export_df = export_df.sort_values(sort_cols)
        export_df.to_csv(csv_output, index=False)
        logger.info(f"Balanced share metrics CSV exported to: {csv_output}")
        print(f"Balanced CSV: {csv_output}")
    
    else:
        logger.warning("No valid results provided for CSV export")
        return


def main() -> int:
    """Main entry point."""
    parser = create_parser()
    
    if len(sys.argv) == 1:
        parser.print_help()
        return 0
    
    args = parser.parse_args()
    
    # Handle version flag
    if hasattr(args, 'version') and args.version:
        print_version()
        return 0
    
    # Handle config command
    if args.command == 'config':
        return handle_config_command(args)
    
    # Handle deprecated presets command
    if args.command == 'presets':
        list_presets()
        return 0
    
    if not args.command:
        parser.print_help()
        return 0
    
    # Setup logging for analysis commands
    log_level = getattr(args, 'log_level', None) or 'INFO'
    logger = setup_logging(log_level, f"benchmark_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
    
    # Route to appropriate handler
    if args.command == 'share':
        return run_share_analysis(args, logger)
    elif args.command == 'rate':
        return run_rate_analysis(args, logger)
    else:
        parser.print_help()
        return 1


if __name__ == '__main__':
    sys.exit(main())
