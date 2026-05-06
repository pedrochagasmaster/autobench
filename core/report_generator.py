"""
ReportGenerator - Output report generation module.

Generates benchmark reports in multiple formats (Excel, CSV, JSON).
"""

import pandas as pd
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import json
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates benchmark reports in various formats.
    
    Supports:
    - Excel workbooks with multiple sheets
    - CSV exports
    - JSON structured data
    - Metadata and audit trails
    """
    
    def __init__(self, config: Any):
        """
        Initialize report generator.
        
        Parameters:
        -----------
        config : ConfigManager
            Configuration manager instance
        """
        self.config = config
        logger.info("Initialized ReportGenerator")

    @staticmethod
    def _write_header_row(
        worksheet: Any,
        row_index: int,
        headers: List[str],
        font_cls: Any,
        fill_cls: Any,
        fill_color: Optional[str] = None
    ) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_index, column=col_idx, value=header)
            cell.font = font_cls(bold=True)
            if fill_color:
                cell.fill = fill_cls(start_color=fill_color, end_color=fill_color, fill_type="solid")

    @staticmethod
    def _set_column_widths(worksheet: Any, widths: Dict[str, int]) -> None:
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    @staticmethod
    def _resolve_convert_all_rates(metadata: Optional[Dict[str, Any]]) -> bool:
        if not metadata:
            return False
        analysis_label = str(metadata.get('analysis_type', '')).lower()
        rate_types = [str(rt).lower() for rt in metadata.get('rate_types', [])]
        if rate_types and all(rt == 'fraud' for rt in rate_types):
            return True
        return 'fraud_rate' in analysis_label and not rate_types

    @staticmethod
    def _should_convert_rate_column(column_name: str, convert_all_rates: bool) -> bool:
        col_lower = str(column_name).lower().strip()
        non_rate_markers = (
            'impact',
            'effect',
            'distortion',
            'weight',
            'multiplier',
            'total',
            'volume',
            'count',
            'numerator',
            'denominator',
        )
        if any(marker in col_lower for marker in non_rate_markers):
            return False

        rate_patterns = (
            col_lower.endswith('_raw_%'),
            col_lower.endswith('_balanced_%'),
            col_lower in {'target rate (%)', 'balanced peer average (%)', 'bic (%)'},
            'rate' in col_lower,
        )
        if not any(rate_patterns):
            return False
        if convert_all_rates:
            return True
        return 'fraud' in col_lower

    @staticmethod
    def _build_unique_sheet_name(raw_name: str, existing_names: List[str]) -> str:
        base = str(raw_name).strip() or "Sheet"
        candidate = base[:31]
        if candidate not in existing_names:
            return candidate

        suffix = 1
        while True:
            suffix_text = f"_{suffix}"
            trimmed = base[: 31 - len(suffix_text)]
            candidate = f"{trimmed}{suffix_text}"
            if candidate not in existing_names:
                return candidate
            suffix += 1
    
    def generate_report(
        self,
        results: Dict[str, Any],
        output_file: str,
        format: str = 'excel',
        analysis_type: str = 'rate',
        metadata: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Generate benchmark report.
        
        Parameters:
        -----------
        results : Dict
            Analysis results
        output_file : str
            Output file path
        format : str
            Output format ('excel', 'csv', 'json')
        analysis_type : str
            Type of analysis performed
        metadata : Dict, optional
            Additional metadata to include
        """
        logger.info(f"Generating {format} report: {output_file}")
        
        # Early check for Excel dependencies - fail fast with clear message
        if format == 'excel':
            self._ensure_excel_support()
            self._generate_excel_report(results, output_file, analysis_type, metadata)
        elif format == 'csv':
            self._generate_csv_report(results, output_file, analysis_type, metadata)
        elif format == 'json':
            self._generate_json_report(results, output_file, analysis_type, metadata)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        logger.info(f"Report saved to: {output_file}")
    
    def _ensure_excel_support(self) -> None:
        """
        Check Excel dependencies before attempting to generate.
        
        Raises:
        -------
        ImportError
            If openpyxl is not installed, with helpful message.
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "Generating Excel reports requires 'openpyxl'. "
                "Install it with: pip install openpyxl\n"
                "Or use --format csv or --format json instead."
            )
    
    def _generate_excel_report(
        self,
        results: Dict[str, Any],
        output_file: str,
        analysis_type: str,
        metadata: Optional[Dict[str, Any]]
    ) -> None:
        """Generate Excel workbook report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise
        
        # Store styles as instance variables for use in other methods
        self._font_class = Font
        self._fill_class = PatternFill
        self._align_class = Alignment
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary, results, analysis_type, metadata)
        
        # Create sheets for each result
        for i, (metric_name, result_data) in enumerate(results.items()):
            base_name = f"Metric_{i+1}_{metric_name[:20]}"
            sheet_name = self._build_unique_sheet_name(base_name, wb.sheetnames)
            ws = wb.create_sheet(sheet_name)
            self._write_metric_sheet(ws, metric_name, result_data, analysis_type)

        self._write_optional_dataframe_sheet(wb, "Peer Weights", metadata, "weights_df")
        self._write_optional_dataframe_sheet(wb, "Weight Methods", metadata, "method_breakdown_df")
        self._write_optional_dataframe_sheet(wb, "Privacy Validation", metadata, "privacy_validation_df")
        self._write_optional_dataframe_sheet(wb, "Preset Comparison", metadata, "preset_comparison_df")
        self._write_optional_dataframe_sheet(wb, "Impact Analysis", metadata, "impact_df")
        self._write_optional_dataframe_sheet(wb, "Impact Summary", metadata, "impact_summary_df")
        self._write_optional_dataframe_sheet(wb, "Secondary Metrics", metadata, "secondary_results")
        if metadata and "validation_issues" in metadata:
            validation_issues = metadata.get("validation_issues") or []
            passed = not any("ERROR" in str(getattr(issue, "severity", "")) for issue in validation_issues)
            self.add_data_quality_sheet(wb, validation_issues, passed=passed)
        
        # Create Metadata sheet
        if metadata:
            ws_meta = wb.create_sheet("Metadata")
            self._write_metadata_sheet(ws_meta, metadata)
        
        # Save workbook
        wb.save(output_file)
    
    def _write_summary_sheet(
        self,
        worksheet: Any,
        results: Dict[str, Any],
        analysis_type: str,
        metadata: Optional[Dict[str, Any]]
    ) -> None:
        """Write summary information to worksheet."""
        # Header
        worksheet['A1'] = "Benchmark Analysis Summary"
        worksheet['A1'].font = self._font_class(bold=True, size=14)
        
        row = 3
        
        # Analysis info
        worksheet[f'A{row}'] = "Analysis Type:"
        worksheet[f'B{row}'] = analysis_type.upper()
        row += 1
        
        if metadata:
            worksheet[f'A{row}'] = "Entity:"
            worksheet[f'B{row}'] = metadata.get('entity', 'N/A')
            row += 1
            
            worksheet[f'A{row}'] = "Timestamp:"
            worksheet[f'B{row}'] = metadata.get('timestamp', datetime.now().isoformat())
            row += 1
            
            if 'dimensions' in metadata:
                worksheet[f'A{row}'] = "Dimensions:"
                worksheet[f'B{row}'] = ', '.join(metadata['dimensions'])
                row += 1
            
            worksheet[f'A{row}'] = "Privacy Rule:"
            worksheet[f'B{row}'] = f"{metadata.get('participants', 'N/A')} participants, " \
                                   f"{metadata.get('max_concentration', 'N/A')}% max"
            row += 2
            compliance_summary = metadata.get('compliance_summary', {})
            for label, key in [
                ("Compliance Posture:", "compliance_posture"),
                ("Compliance Verdict:", "compliance_verdict"),
                ("Acknowledgement State:", "acknowledgement_state"),
                ("Run Status:", "run_status"),
            ]:
                worksheet[f'A{row}'] = label
                worksheet[f'B{row}'] = metadata.get(key, compliance_summary.get(key, 'N/A'))
                row += 1
            row += 1
        
        # Results summary
        worksheet[f'A{row}'] = "Metrics Analyzed:"
        worksheet[f'A{row}'].font = self._font_class(bold=True)
        row += 1
        
        for metric_name in results.keys():
            worksheet[f'A{row}'] = f"  - {metric_name}"
            row += 1
    
    def _write_metric_sheet(
        self,
        worksheet: Any,
        metric_name: str,
        result_data: Any,
        analysis_type: str
    ) -> None:
        """Write metric results to worksheet."""
        # Header
        worksheet['A1'] = f"Metric: {metric_name}"
        worksheet['A1'].font = self._font_class(bold=True, size=12)
        
        row = 3
        
        if isinstance(result_data, dict):
            # Write dictionary results as key-value pairs
            for key, value in result_data.items():
                worksheet[f'A{row}'] = str(key).replace('_', ' ').title()
                worksheet[f'B{row}'] = value
                row += 1
        elif isinstance(result_data, pd.DataFrame):
            # Write dataframe with headers
            headers = list(result_data.columns)
            for c_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row, column=c_idx, value=header)
                cell.font = self._font_class(bold=True)
            row += 1
            for r_idx, row_data in enumerate(result_data.itertuples(index=False), start=row):
                for c_idx, value in enumerate(row_data, start=1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)

    def _format_metadata_value(self, value: Any) -> Any:
        if hasattr(value, "shape"):
            return f"DataFrame rows={value.shape[0]} cols={value.shape[1]}"
        if isinstance(value, list):
            if value and hasattr(value[0], "severity"):
                return f"ValidationIssue count={len(value)}"
            return json.dumps(value, indent=2, default=str)
        if isinstance(value, dict):
            return json.dumps(value, indent=2, default=str)
        return str(value)

    def _write_optional_dataframe_sheet(
        self,
        workbook: Any,
        sheet_name: str,
        metadata: Optional[Dict[str, Any]],
        metadata_key: str,
    ) -> None:
        if not metadata:
            return
        df = metadata.get(metadata_key)
        if df is None or not hasattr(df, "empty") or df.empty:
            return

        ws = workbook.create_sheet(self._build_unique_sheet_name(sheet_name, workbook.sheetnames))
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(column))
            if hasattr(self, "_font_class"):
                cell.font = self._font_class(bold=True)
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _write_metadata_sheet(
        self,
        worksheet: Any,
        metadata: Dict[str, Any]
    ) -> None:
        """Write metadata to worksheet."""
        worksheet['A1'] = "Analysis Metadata"
        worksheet['A1'].font = self._font_class(bold=True, size=12)
        
        row = 3
        for key, value in metadata.items():
            worksheet[f'A{row}'] = str(key).replace('_', ' ').title()
            worksheet[f'B{row}'] = self._format_metadata_value(value)
            
            row += 1
    
    def add_preset_comparison_sheet(
        self,
        workbook: Any,
        comparison_df: 'pd.DataFrame',
        analysis_type: str = 'share'
    ) -> None:
        """
        Add Preset Comparison sheet to workbook.
        
        Parameters
        ----------
        workbook : openpyxl.Workbook
            Target workbook
        comparison_df : pd.DataFrame
            DataFrame with columns: Preset, Mean_Distortion, Max_Distortion, Time_Seconds, Best
        analysis_type : str
            'share' or 'rate'
        """
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws = workbook.create_sheet("Preset Comparison")
        
        # Header
        ws['A1'] = f"Preset Comparison - {analysis_type.title()} Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Column headers
        headers = ['Preset', 'Mean Distortion (PP)', 'Max Distortion (PP)', 'Time (s)', 'Best']
        self._write_header_row(ws, 3, headers, Font, PatternFill, "CCE5FF")
        
        # Data rows
        for row_idx, row_data in enumerate(dataframe_to_rows(comparison_df, index=False, header=False), 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Highlight best row
                if len(row_data) > 4 and row_data[4] == '*':
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        # Adjust column widths
        self._set_column_widths(ws, {'A': 25, 'B': 20, 'C': 20, 'D': 12, 'E': 8})
        
        logger.info("Added Preset Comparison sheet")
    
    def add_distortion_summary_sheet(
        self,
        workbook: Any,
        summary_df: 'pd.DataFrame',
        analysis_type: str = 'share'
    ) -> None:
        """
        Add Distortion Summary sheet to workbook.
        
        Parameters
        ----------
        workbook : openpyxl.Workbook
            Target workbook
        summary_df : pd.DataFrame
            DataFrame with columns: Dimension, Mean_Distortion, Min, Max, Std
        analysis_type : str
            'share' or 'rate'
        """
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
        
        ws = workbook.create_sheet("Distortion Summary")
        
        # Header
        ws['A1'] = f"Distortion Summary - {analysis_type.title()} Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = "Shows impact of privacy weighting on metric values (percentage points)"
        ws['A2'].font = Font(italic=True, size=10)
        
        # Column headers
        headers = ['Dimension', 'Mean (PP)', 'Min (PP)', 'Max (PP)', 'Std Dev']
        self._write_header_row(ws, 4, headers, Font, PatternFill, "FFF2CC")
        
        # Data rows
        for row_idx, row_data in enumerate(dataframe_to_rows(summary_df, index=False, header=False), 5):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        self._set_column_widths(ws, {'A': 20, 'B': 12, 'C': 12, 'D': 12, 'E': 12})
        
        logger.info("Added Distortion Summary sheet")
    
    def add_data_quality_sheet(
        self,
        workbook: Any,
        validation_issues: List[Any],
        passed: bool = True
    ) -> None:
        """
        Add Data Quality sheet to workbook.
        
        Parameters
        ----------
        workbook : openpyxl.Workbook
            Target workbook
        validation_issues : List[ValidationIssue]
            List of validation issues (each with severity, category, message attributes)
        passed : bool
            Whether validation passed overall
        """
        from openpyxl.styles import Font, PatternFill
        
        ws = workbook.create_sheet("Data Quality")
        
        # Header
        ws['A1'] = "Data Quality Report"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Status
        status_text = "PASSED" if passed else "ISSUES FOUND"
        status_color = "E6FFE6" if passed else "FFCCCC"
        ws['A3'] = f"Status: {status_text}"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        
        if not validation_issues:
            ws['A5'] = "No issues detected."
            return
        
        # Count by severity
        error_count = sum(1 for i in validation_issues if getattr(i, 'severity', None) and 'ERROR' in str(i.severity))
        warn_count = sum(1 for i in validation_issues if getattr(i, 'severity', None) and 'WARN' in str(i.severity))
        info_count = len(validation_issues) - error_count - warn_count
        
        ws['A5'] = f"Errors: {error_count} | Warnings: {warn_count} | Info: {info_count}"
        
        # Column headers
        headers = ['Severity', 'Category', 'Message']
        self._write_header_row(ws, 7, headers, Font, PatternFill, "DDDDDD")
        
        # Issue rows
        for row_idx, issue in enumerate(validation_issues, 8):
            severity = str(getattr(issue, 'severity', 'UNKNOWN'))
            category = str(getattr(issue, 'category', ''))
            message = str(getattr(issue, 'message', str(issue)))
            
            ws.cell(row=row_idx, column=1, value=severity)
            ws.cell(row=row_idx, column=2, value=category)
            ws.cell(row=row_idx, column=3, value=message)
            
            # Color by severity
            if 'ERROR' in severity:
                color = "FFCCCC"
            elif 'WARN' in severity:
                color = "FFF2CC"
            else:
                color = "E6F3FF"
            
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
        
        # Adjust column widths
        self._set_column_widths(ws, {'A': 12, 'B': 20, 'C': 60})
        
        logger.info("Added Data Quality sheet")
    
    def _generate_csv_report(
        self,
        results: Dict[str, Any],
        output_file: str,
        analysis_type: str,
        metadata: Optional[Dict[str, Any]]
    ) -> None:
        """Generate CSV report."""
        # Flatten results to dataframe
        rows = []
        dataframes = []
        
        for metric_name, result_data in results.items():
            if isinstance(result_data, dict):
                row = {'metric': metric_name}
                row.update(result_data)
                rows.append(row)
            elif isinstance(result_data, pd.DataFrame):
                dataframes.append((metric_name, result_data))
        
        if rows:
            df_results = pd.DataFrame(rows)
            df_results.to_csv(output_file, index=False)
        
        if dataframes:
            if not rows and len(dataframes) == 1:
                # Write single dataframe to the requested output file
                _, df = dataframes[0]
                df.to_csv(output_file, index=False)
            else:
                # Save each dataframe result as separate CSV
                output_path = Path(output_file)
                for metric_name, df in dataframes:
                    df_output = output_path.with_name(f"{output_path.stem}_{metric_name}.csv")
                    df.to_csv(df_output, index=False)
    
    def _generate_json_report(
        self,
        results: Dict[str, Any],
        output_file: str,
        analysis_type: str,
        metadata: Optional[Dict[str, Any]]
    ) -> None:
        """Generate JSON report."""
        output_data = {
            'analysis_type': analysis_type,
            'metadata': metadata or {},
            'results': {}
        }
        
        for metric_name, result_data in results.items():
            if isinstance(result_data, dict):
                output_data['results'][metric_name] = result_data
            elif isinstance(result_data, pd.DataFrame):
                output_data['results'][metric_name] = result_data.to_dict(orient='records')
        
        with open(output_file, 'w') as f:
            json.dump(output_data, f, indent=2, default=str)
    
    def create_audit_log(
        self,
        log_file: str,
        metadata: Dict[str, Any],
        results_summary: Dict[str, Any]
    ) -> None:
        """
        Create audit log file.
        
        Parameters:
        -----------
        log_file : str
            Path to log file
        metadata : Dict
            Analysis metadata
        results_summary : Dict
            Summary of results
        """
        logger.info(f"Creating audit log: {log_file}")
        
        with open(log_file, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("BENCHMARK ANALYSIS AUDIT LOG\n")
            f.write("=" * 80 + "\n\n")
            
            f.write("ANALYSIS METADATA\n")
            f.write("-" * 80 + "\n")
            for key, value in metadata.items():
                f.write(f"{key}: {value}\n")
            f.write("\n")
            
            f.write("RESULTS SUMMARY\n")
            f.write("-" * 80 + "\n")
            for key, value in results_summary.items():
                f.write(f"{key}: {value}\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write(f"Log generated: {datetime.now().isoformat()}\n")
    
    def generate_publication_workbook(
        self,
        results: Dict[str, Any],
        output_file: str,
        analysis_type: str = 'share',
        metadata: Optional[Dict[str, Any]] = None,
        fraud_in_bps: bool = True
    ) -> None:
        """
        Generate publication-ready workbook with simplified formatting.
        
        This creates a clean, stakeholder-friendly Excel file without
        debug sheets, weight details, or technical metadata.
        
        Parameters
        ----------
        results : Dict
            Analysis results containing dimension DataFrames
        output_file : str
            Output file path
        analysis_type : str
            'share' or 'rate'
        metadata : Dict, optional
            Basic metadata (entity, date, etc.)
        fraud_in_bps : bool
            If True and analysis_type='rate', convert fraud rates to basis points
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Style definitions for publication
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create Executive Summary sheet
        ws_summary = wb.create_sheet("Executive Summary")
        ws_summary['A1'] = f"Benchmark Analysis - {analysis_type.title()}"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:E1')
        
        if metadata:
            row = 3
            for key in ['entity', 'date', 'analysis_type', 'compliance_posture', 'compliance_verdict', 'acknowledgement_state']:
                if key in metadata:
                    ws_summary[f'A{row}'] = key.replace('_', ' ').title()
                    ws_summary[f'A{row}'].font = Font(bold=True)
                    ws_summary[f'B{row}'] = str(metadata[key])
                    row += 1
        
        # Create dimension sheets (publication format)
        for metric_name, result_data in results.items():
            if not isinstance(result_data, (dict, pd.DataFrame)):
                continue
            
            sheet_name = self._build_unique_sheet_name(str(metric_name), wb.sheetnames)
            ws = wb.create_sheet(sheet_name)
            
            # Sheet header
            ws['A1'] = f"{metric_name} Analysis"
            ws['A1'].font = header_font
            ws.merge_cells('A1:F1')
            
            # Get DataFrame
            if isinstance(result_data, dict) and 'data' in result_data:
                df = result_data['data']
            elif isinstance(result_data, pd.DataFrame):
                df = result_data
            else:
                continue
            
            # Apply fraud BPS conversion for rate analysis (copy to avoid mutation)
            df = df.copy(deep=True)
            if analysis_type == 'rate' and fraud_in_bps:
                is_fraud_sheet = 'fraud' in str(metric_name).lower()
                convert_all_rates = self._resolve_convert_all_rates(metadata)
                for col in df.columns:
                    if not pd.api.types.is_numeric_dtype(df[col]):
                        continue
                    if is_fraud_sheet and self._should_convert_rate_column(col, True):
                        df[col] = df[col] * 100
                        df.rename(columns={col: f"{col} (bps)"}, inplace=True)
                    elif self._should_convert_rate_column(col, convert_all_rates):
                        df[col] = df[col] * 100
            
            # Write data with formatting
            for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    if row_idx == 3:  # Header row
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        # Format numbers
                        if isinstance(value, float):
                            cell.number_format = '0.00'
        
        # Auto-adjust column widths (skip merged header cells safely)
        from openpyxl.utils import get_column_letter
        for ws in wb.worksheets:
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in ws[column_letter]:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        wb.save(output_file)
        logger.info(f"Publication workbook saved to: {output_file}")
